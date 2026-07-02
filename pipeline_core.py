#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import hashlib
import html
import json
import math
import pickle
import re
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from sklearn.decomposition import TruncatedSVD
from sklearn.ensemble import RandomForestClassifier
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.impute import SimpleImputer
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import (
    accuracy_score,
    average_precision_score,
    balanced_accuracy_score,
    brier_score_loss,
    classification_report,
    f1_score,
    roc_auc_score,
    top_k_accuracy_score,
)
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import StandardScaler

try:
    from joblib import dump as joblib_dump
except Exception:  # pragma: no cover - fallback path
    joblib_dump = None


ITALIAN_STOPWORDS = {
    "a",
    "ad",
    "al",
    "alla",
    "alle",
    "anche",
    "con",
    "da",
    "dal",
    "dalla",
    "dalle",
    "dei",
    "del",
    "della",
    "delle",
    "di",
    "e",
    "ed",
    "il",
    "in",
    "la",
    "le",
    "lo",
    "ma",
    "mentre",
    "nel",
    "nella",
    "nelle",
    "non",
    "per",
    "piu",
    "presso",
    "si",
    "sua",
    "sul",
    "sulla",
    "tra",
    "un",
    "una",
}

PROBLEM_FAMILY_LABELS = {
    "A": "Ambiente di lavoro",
    "B": "Formazione / Informazione / Addestramento",
    "C": "DPI",
    "D": "Impianti e attrezzature",
    "E": "Regolamento interno",
    "F": "5S / Ordine / Pulizia",
    "G": "Ambiente",
    "H": "Azioni e condizioni insicure",
    "I": "Idoneita sanitaria",
    "L": "Ditte esterne",
    "V": "COVID-19",
    "Z": "Comportamenti sicuri",
    "UNKNOWN": "Non classificato",
}

UNKNOWN_INJURY_DETAIL = "Altro / Non definita"

EXPERT_RISK_LABELS = {
    1: "Molto alto",
    2: "Alto",
    3: "Medio",
    4: "Basso",
    5: "Molto basso",
}

DEEPDIVE_SEVERITY_WEIGHTS = {
    1: 1.50,
    2: 1.10,
    3: 0.35,
    4: 0.12,
    5: 0.05,
}

DEFAULT_BEHAVIOR_EXPERT_RISK_RANKS = {
    "Attenzione / Buone Prassi": (4, "Deviazioni generiche a severita di norma contenuta, salvo contesto specifico."),
    "Carrelli / Transpallet / Mezzi": (1, "Interazione con mezzi e carichi mobili con potenziale di investimento o schiacciamento."),
    "Chimici / Aspirazione / Lavaggi": (1, "Esposizione a chimici, aerosol o contatti con potenziale danno elevato."),
    "Comunicazione / Messaggi": (5, "Segnale informativo, non comportamento operativo pericoloso diretto."),
    "DPI / Abbigliamento": (2, "Mancato uso o uso scorretto DPI aumenta l'esposizione al danno."),
    "DPI mani / protezioni": (2, "Protezione mani critica per tagli, urti, schiacciamenti e contatti."),
    "DPI occhi / viso": (2, "Protezione occhi e viso critica per proiezioni, schizzi e contatti."),
    "DPI piedi / abbigliamento": (2, "Protezione piedi e indumenti rilevante per urti, schiacciamenti e contatti."),
    "DPI udito": (2, "Protezione udito importante in esposizione continuativa a rumore."),
    "Da classificare": (3, "Bucket neutro provvisorio da revisionare."),
    "Emergenza / Elettrico / Vie di fuga": (1, "Rischi elettrici o di gestione emergenza con severita potenziale molto alta."),
    "Interferenze / Coordinamento": (2, "Interferenze operative tra persone o squadre con possibile perdita di controllo."),
    "Lavori Speciali / Accessi": (1, "Accessi speciali o lavori atipici con potenziale severita elevata."),
    "Macchine / Sicurezze / Attrezzaggio": (1, "Bypass sicurezze o interazione pericolosa con macchine e attrezzaggi."),
    "Manutenzione / Officina": (2, "Attivita manutentive con energia residua, utensili e parti tecniche."),
    "Movimentazione Manuale / Ergonomia": (3, "Rischio diffuso ma in media meno severo sul piano traumatico acuto."),
    "Ordine / Pulizia / 5S": (4, "Segnale importante ma in media meno severo di DPI, macchine o mezzi."),
    "Produzione / Processo": (3, "Rischi operativi presenti ma severita media molto dipendente dal contesto."),
    "Segnalazioni / Near miss / Checklist": (5, "Prompt o segnalazione, non comportamento unsafe diretto."),
    "Spostamenti / Attenzione": (2, "Include corsa, disattenzione e dinamiche di movimento con rischio caduta o urto."),
    "Spostamenti / Trasporto / Stoccaggio": (2, "Movimentazione e trasporto con rischio urto, caduta materiale o perdita controllo."),
    "Stoccaggio / Materiali": (2, "Rischio di caduta materiali, urti e schiacciamenti."),
    "Utensili / Taglio / Aria Compressa": (2, "Utensili e aria compressa con potenziale lesivo diretto."),
}

BEHAVIOR_TAXONOMY_REQUIRED_COLUMNS = {
    "comportamento_clean",
    "comportamento_canonico",
    "canonical_behavior_id",
    "comportamento_famiglia",
    "comportamento_tipo",
}

BEHAVIOR_TAXONOMY_STRING_COLUMNS = [
    "comportamento_clean",
    "comportamento_canonico",
    "canonical_behavior_id",
    "comportamento_famiglia_originale",
    "comportamento_famiglia",
    "comportamento_famiglia_v2",
    "comportamento_sottofamiglia_v2",
    "meccanismo_rischio_v2",
    "presidio_v2",
    "taxonomy_source",
    "taxonomy_confidence",
    "comportamento_tipo",
    "raw_variants",
    "old_family_conflict",
    "old_families_cluster",
    "merge_source",
    "merge_confidence",
    "merge_note",
    "review_flag",
]

BEHAVIOR_TAXONOMY_NUMERIC_COLUMNS = [
    "severity_rank_v2",
    "row_count_kept",
    "raw_variants_count",
    "canonical_cluster_size",
    "canonical_row_count_total",
]

BEHAVIOR_TAXONOMY_OUTPUT_COLUMNS = [
    "comportamento_canonico",
    "canonical_behavior_id",
    "comportamento_famiglia_originale",
    "comportamento_famiglia",
    "comportamento_famiglia_v2",
    "comportamento_sottofamiglia_v2",
    "meccanismo_rischio_v2",
    "presidio_v2",
    "severity_rank_v2",
    "taxonomy_source",
    "taxonomy_confidence",
    "canonical_cluster_size",
    "canonical_row_count_total",
    "old_family_conflict",
    "old_families_cluster",
    "merge_source",
    "merge_confidence",
    "merge_note",
    "review_flag",
]

BEHAVIOR_FAMILY_INVESTIGATION_HINTS = {
    "Attenzione / Buone Prassi": "Verificare disciplina operativa, standard di lavoro e supervisione in linea.",
    "Carrelli / Transpallet / Mezzi": "Verificare percorsi mezzi/pedoni, velocita, visibilita e stabilita dei carichi.",
    "Chimici / Aspirazione / Lavaggi": "Verificare travasi, aspirazione, contenimento, etichettatura e DPI dedicati.",
    "Comunicazione / Messaggi": "Verificare se dietro i messaggi ci sono segnali operativi non ancora classificati.",
    "DPI / Abbigliamento": "Verificare disponibilita, adeguatezza e uso corretto dei DPI richiesti.",
    "DPI mani / protezioni": "Verificare uso di guanti corretti, compatibilita con il task e momenti di mancato utilizzo.",
    "DPI occhi / viso": "Verificare occhiali/visiere richiesti, appannamento, comfort e momenti di rimozione.",
    "DPI piedi / abbigliamento": "Verificare calzature, indumenti e stato di usura nelle attivita esposte.",
    "DPI udito": "Verificare zone rumorose, uso reale dei DPI udito e eventuale sottostima dell'esposizione.",
    "Da classificare": "Riclassificare i segnali per non perdere informazione utile alla diagnosi.",
    "Emergenza / Elettrico / Vie di fuga": "Verificare quadri, cavi, segregazioni, accessibilita e gestione delle vie di fuga.",
    "Interferenze / Coordinamento": "Verificare sovrapposizioni tra squadre, ditte esterne e passaggi di consegna.",
    "Lavori Speciali / Accessi": "Verificare permessi, accessi speciali, pre-job review e barriere temporanee.",
    "Macchine / Sicurezze / Attrezzaggio": "Verificare ripari, interblocchi, attrezzaggi, bypass e gestione energia residua.",
    "Manutenzione / Officina": "Verificare isolamento energie, sequenza manutentiva, utensili e prove post-intervento.",
    "Movimentazione Manuale / Ergonomia": "Verificare posture, pesi, frequenza, ausili e punti di presa.",
    "Ordine / Pulizia / 5S": "Verificare housekeeping, passaggi liberi, contenitori, frequenza pulizia e ownership.",
    "Produzione / Processo": "Verificare deviazioni standard, pressioni di cadenza, set-up e workaround di processo.",
    "Segnalazioni / Near miss / Checklist": "Usare il segnale per aprire indagini, ma non trattarlo come driver causale diretto.",
    "Spostamenti / Attenzione": "Verificare corsa, distrazioni, uso cellulare, ostacoli e flussi pedonali.",
    "Spostamenti / Trasporto / Stoccaggio": "Verificare percorsi, movimentazione, impilamento e perdite di controllo del carico.",
    "Stoccaggio / Materiali": "Verificare impilamento, trattenimento, caduta materiali e punti di schiacciamento.",
    "Utensili / Taglio / Aria Compressa": "Verificare stato utensili, lame, protezioni, soffiaggi e modalita d'uso.",
}

DESCRIPTION_TOPIC_PATTERNS = {
    "Cadute / scivolamenti": [
        "scivol",
        "cadut",
        "inciamp",
        "perde equilibrio",
        "manca un gradino",
    ],
    "Scale / accessi": [
        "scala",
        "scale",
        "gradin",
        "maniglione",
        "ingresso",
        "uscita",
    ],
    "Pavimenti / ordine / pulizia": [
        "paviment",
        "olio",
        "sporco",
        "bagnat",
        "ghiacci",
        "condensa",
        "truciol",
        "disordine",
        "rifiut",
        "ostruit",
    ],
    "Vie di fuga / emergenza": [
        "via di esodo",
        "via d esodo",
        "via fuga",
        "uscita emergenza",
        "scala di emergenza",
        "estintor",
        "emergenza",
    ],
    "Carrelli / mezzi / AGV": [
        "carrello",
        "transpallet",
        "mulet",
        "agv",
        "mir",
        "mezzo",
    ],
    "Macchine / impianti / attrezzature": [
        "macchin",
        "impiant",
        "attrezz",
        "fresa",
        "forno",
        "pressa",
        "ripar",
        "utensil",
        "lama",
    ],
    "Taglio / urto / schiacciamento": [
        "tagli",
        "urt",
        "schiacci",
        "pizzic",
        "incastr",
        "dito",
        "dita",
        "mano",
        "pollice",
    ],
    "DPI occhi / viso": [
        "occhial",
        "occhio",
        "visiera",
        "viso",
        "schegg",
        "proiezion",
    ],
    "DPI mani / guanti": [
        "guant",
    ],
    "Chimici / lavaggi / aspirazione": [
        "chimic",
        "aspiraz",
        "solvent",
        "inchiostr",
        "vernic",
        "saponi",
        "vasca",
        "lavagg",
        "acido",
        "idropul",
    ],
    "Elettrico": [
        "elettric",
        "quadro",
        "cavo",
        "presa",
    ],
    "Movimentazione / ergonomia": [
        "sollev",
        "moviment",
        "spingev",
        "trascin",
        "peso",
        "sforz",
        "postur",
    ],
    "Materiali / stoccaggio": [
        "material",
        "stocc",
        "deposit",
        "bancal",
        "scatol",
        "impilat",
        "telai",
        "carico",
        "scarico",
    ],
    "Distrazione / comportamento": [
        "cellular",
        "fum",
        "distraz",
        "corr",
        "fretta",
    ],
    "Rumore / udito": [
        "rumor",
        "udito",
        "cuffie",
        "otoprot",
    ],
}

ACTIVITY_CANONICAL_LABEL_OVERRIDES = {
    "altri comportamenti": "Altri comportamenti",
    "burattatura": "Burattatura",
    "comportamento da inserire": "Comportamento da inserire",
    "dpi": "DPI",
    "filettatura": "Filettatura",
    "gestione sostanze chimiche": "Gestione sostanze chimiche",
    'interferenze del personale di reparto in attivita\' "extra reparto"': 'Interferenze del personale di reparto in attivita\' "extra reparto"',
    "incidenti sfiorati": "Incidenti sfiorati",
    "messaggio": "Messaggio",
    "messaggio a seguito di incidenti": "Messaggio a seguito di incidenti",
    "montaggio e registrazione": "Montaggio e registrazione",
    "ordine": "Ordine",
    "presidio sicurezza, evacuazione ed emergenza": "Presidio sicurezza, evacuazione ed emergenza",
    "pulizia": "Pulizia",
    "sfiammatura": "Sfiammatura",
    "spostamenti": "Spostamenti",
    "uso della saldatrice": "Uso della saldatrice",
    "uso ple": "Uso PLE",
}

ACTIVITY_FAMILY_OVERRIDES = {
    "altri comportamenti": "Comportamenti Generali e Buone Prassi",
    "attenzione e buone prassi": "Comportamenti Generali e Buone Prassi",
    "attrezzaggio e pulizia macchine": "Macchine Impianti e Sicurezze",
    "b - da non usare": "Da classificare",
    "comportamento da inserire": "Da classificare",
    "filettatura": "Produzione e Processo",
    'interferenze del personale di reparto in attivita\' "extra reparto"': "Interferenze e Coordinamento",
    "incidenti sfiorati": "Near Miss / Incidenti sfiorati",
    "interventi in impianto/macchina (spostamenti/nuove installazioni, messe in funzione/modifiche di macchine/impianti, interventi edili)": "Macchine Impianti e Sicurezze",
    "messaggio a seguito di incidenti": "Near Miss / Incidenti sfiorati",
    "messaggio": "Comunicazione e Sensibilizzazione",
    "movimentazione con ausili": "Spostamenti Trasporto e Stoccaggio",
    "montaggio e registrazione": "Produzione e Processo",
    "presidio sicurezza, evacuazione ed emergenza": "Emergenza / Evacuazione",
    "sfiammatura": "Produzione e Processo",
    "uso della saldatrice": "Macchine Impianti e Sicurezze",
}

ACTIVITY_FAMILY_RULES = [
    (
        "DPI e Abbigliamento",
        [
            "dpi",
            "abbigliamento da lavoro",
        ],
    ),
    (
        "Ordine Pulizia e Area di Lavoro",
        [
            "ordine",
            "pulizia",
            "area di lavoro",
            "igiene di se",
            "igiene di se'",
            "corretto smaltimento",
        ],
    ),
    (
        "Spostamenti Trasporto e Stoccaggio",
        [
            "spostamenti",
            "trasporto",
            "stoccaggio",
            "carico/scarico telai",
            "muletto",
            "transpallet",
            "deposito dei prodotti",
            "prelievo e versamento",
        ],
    ),
    (
        "Ergonomia e Movimentazione",
        [
            "ergonomia",
            "movimentazione manuale",
            "micromovimentazioni",
        ],
    ),
    (
        "Macchine Impianti e Sicurezze",
        [
            "sicurezze macchine",
            "uso macchine",
            "organi in movimento",
            "attivita' con macchine",
            "attrezzaggio macchine",
            "interventi in impianto/macchina",
            "aspirazione macchine",
            "saldatrice",
        ],
    ),
    (
        "Utensili Manuali e Aria Compressa",
        [
            "utensili manuali",
            "taglierino",
            "taglia-reggette",
            "taglia reggette",
            "aria compressa",
            "avvitatura",
        ],
    ),
    (
        "Sostanze Chimiche e Lavaggi",
        [
            "sostanze chimiche",
            "galvanica e laboratori",
            "lavaggi e rabbocchi",
            "preparazione inchiostri",
            "preparazione colori",
            "uso idropulitrice",
            "lavaggio",
        ],
    ),
    (
        "Near Miss / Incidenti sfiorati",
        [
            "incidenti sfiorati",
            "messaggio a seguito di incidenti",
        ],
    ),
    (
        "Interferenze e Coordinamento",
        [
            "interferenze del personale",
            "extra reparto",
        ],
    ),
    (
        "Emergenza Sicurezza e Buone Prassi",
        [
            "presidio sicurezza",
            "evacuazione",
            "emergenza",
            "attenzione e buone prassi",
            "messaggio",
            "altri comportamenti",
        ],
    ),
    (
        "Lavori Speciali",
        [
            "lavoro in altezza",
            "scale da lavoro",
            "sgabelli",
            "step",
            "scale basse",
            "uso ple",
            "lavori elettrici",
            "sotto tensione",
            "spazi confinati",
            "linee vita",
            "criogenici",
        ],
    ),
    (
        "Produzione e Processo",
        [
            "stampaggio",
            "burattatura",
            "saldatura",
            "puntatura",
            "lavorazioni alle ruote",
            "lavorazioni presse",
            "martellatura",
            "laser",
            "tampografia",
            "timbratura",
            "scapsulatura",
            "termoformatura",
            "pesatura",
            "attrezzaggio stampi",
            "preparazione stampi",
            "attrezzaggio e pulizia macchine",
            "verniciatura",
            "resinature",
            "incollaggi",
            "montaggio e registrazione",
            "sfiammatura",
            "filettatura",
        ],
    ),
    (
        "Manutenzione e Officina",
        [
            "manutenzione",
            "interventi meccanici",
            "carpenteria",
            "officina",
        ],
    ),
    (
        "Ditte Esterne",
        [
            "ditte esterne",
        ],
    ),
    (
        "Attivita Generiche",
        [
            "tutte le attivita",
            "durante il turno di lavoro",
            "ingresso/uscita stabilimento",
            "in mensa",
            "in area break",
            "comportamento da inserire",
        ],
    ),
]

AREA_DISPLAY_ORDER = [
    "Stamperia Plastica",
    "Stamperia Metallo",
    "Montaggio e Finitura",
    "Galvanica/Pulitura",
    "Saldatura",
    "Verniciatura",
    "Logistica e Magazzino",
    "Manutenzione",
    "Uffici",
    "Officina",
    "Ricambi",
]

BEHAVIOR_KIND_BEHAVIOR = "behavior"
BEHAVIOR_KIND_MESSAGE = "message"
BEHAVIOR_KIND_NEAR_MISS_PROMPT = "prompt_near_miss"
BEHAVIOR_KIND_OPEN_SIGNAL_PROMPT = "prompt_open_signal"
BEHAVIOR_KIND_MISSING = "missing"

BEHAVIOR_ACTIVITY_FAMILY_FALLBACK = {
    "DPI e Abbigliamento": "DPI / Abbigliamento",
    "Ordine Pulizia e Area di Lavoro": "Ordine / Pulizia / 5S",
    "Spostamenti Trasporto e Stoccaggio": "Spostamenti / Trasporto / Stoccaggio",
    "Ergonomia e Movimentazione": "Movimentazione Manuale / Ergonomia",
    "Macchine Impianti e Sicurezze": "Macchine / Sicurezze / Attrezzaggio",
    "Utensili Manuali e Aria Compressa": "Utensili / Taglio / Aria Compressa",
    "Sostanze Chimiche e Lavaggi": "Chimici / Aspirazione / Lavaggi",
    "Near Miss / Incidenti sfiorati": "Segnalazioni / Near miss / Checklist",
    "Interferenze e Coordinamento": "Interferenze / Coordinamento",
    "Emergenza / Evacuazione": "Emergenza / Elettrico / Vie di fuga",
    "Emergenza Sicurezza e Buone Prassi": "Emergenza / Elettrico / Vie di fuga",
    "Lavori Speciali": "Lavori Speciali / Accessi",
    "Produzione e Processo": "Produzione / Processo",
    "Manutenzione e Officina": "Manutenzione / Officina",
    "Ditte Esterne": "Interferenze / Coordinamento",
    "Attivita Generiche": "Attenzione / Buone Prassi",
    "Comunicazione e Sensibilizzazione": "Comunicazione / Messaggi",
    "Comportamenti Generali e Buone Prassi": "Attenzione / Buone Prassi",
    "Da classificare": "Da classificare",
}

NON_OPERATIONAL_ACTIVITY_LABELS = {
    "Altri comportamenti",
    "Incidenti sfiorati",
    "Messaggio",
    "Messaggio a seguito di incidenti",
    "Comportamento da inserire",
    "B - da NON usare",
    "Tutte le attività",
}


def normalize_whitespace(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).replace("\xa0", " ").strip()
    if not text:
        return None
    return re.sub(r"\s+", " ", text)


def strip_accents(text: str) -> str:
    return "".join(
        char
        for char in unicodedata.normalize("NFKD", text)
        if not unicodedata.combining(char)
    )


def normalize_key(value: Any) -> str | None:
    text = normalize_whitespace(value)
    if text is None:
        return None
    text = text.replace("’", "'").replace("`", "'")
    text = strip_accents(text).lower()
    text = re.sub(r"\s+", " ", text)
    return text


def coerce_null(value: Any) -> Any:
    text = normalize_whitespace(value)
    if text is None:
        return None
    if text.lower() in {"#n/a", "n/a", "na", "none", "null", "nan"}:
        return None
    return text if isinstance(value, str) else value


def to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not pd.isna(value):
        return float(value)
    text = normalize_whitespace(value)
    if text is None:
        return None
    text = text.replace(".", "").replace(",", ".") if "," in text and "." in text else text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def to_iso_date(value: Any) -> str | None:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return None
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    text = normalize_whitespace(value)
    if text is None:
        return None
    parsed = pd.to_datetime(text, errors="coerce", dayfirst=True)
    if pd.isna(parsed):
        return None
    return parsed.strftime("%Y-%m-%d")


def month_start(value: Any) -> pd.Timestamp | None:
    if value is None:
        return None
    if hasattr(value, "year") and hasattr(value, "month"):
        return pd.Timestamp(year=value.year, month=value.month, day=1)
    parsed = pd.to_datetime(value, errors="coerce", dayfirst=True)
    if pd.isna(parsed):
        return None
    return pd.Timestamp(year=parsed.year, month=parsed.month, day=1)


def snake_case(text: str) -> str:
    text = strip_accents(text)
    text = re.sub(r"[^0-9A-Za-z]+", "_", text)
    text = re.sub(r"_+", "_", text)
    return text.strip("_").lower()


def title_from_key(text: str | None) -> str | None:
    if text is None:
        return None
    return " ".join(part.capitalize() for part in text.split())


def build_canonical_behavior_id(text: Any) -> str | None:
    normalized = normalize_key(text)
    if normalized is None:
        return None
    slug = snake_case(normalized).upper() or "BEHAVIOR"
    digest = hashlib.md5(normalized.encode("utf-8")).hexdigest()[:6].upper()
    return f"BEH_{slug[:72]}_{digest}"


def read_csv_autodetect(path: Path, **kwargs: Any) -> pd.DataFrame:
    return pd.read_csv(path, sep=None, engine="python", encoding="utf-8-sig", **kwargs)


def load_behavior_taxonomy(
    path: Path | None,
) -> tuple[pd.DataFrame, pd.DataFrame, dict[str, Any]]:
    empty_profile = {
        "status": "missing",
        "path": str(path) if path is not None else None,
        "input_rows": 0,
        "lookup_rows": 0,
        "duplicate_clean_type_rows_removed": 0,
    }
    if path is None or not path.exists():
        return pd.DataFrame(), pd.DataFrame(), empty_profile

    taxonomy_df = read_csv_autodetect(path)
    missing_columns = sorted(BEHAVIOR_TAXONOMY_REQUIRED_COLUMNS - set(taxonomy_df.columns))
    if missing_columns:
        raise ValueError(
            f"Il file tassonomia comportamenti `{path}` non contiene le colonne richieste: {missing_columns}"
        )

    taxonomy_df = taxonomy_df.copy()
    for column in BEHAVIOR_TAXONOMY_STRING_COLUMNS:
        if column not in taxonomy_df.columns:
            taxonomy_df[column] = None
        taxonomy_df[column] = taxonomy_df[column].map(normalize_whitespace)

    for column in BEHAVIOR_TAXONOMY_NUMERIC_COLUMNS:
        if column not in taxonomy_df.columns:
            taxonomy_df[column] = np.nan
        taxonomy_df[column] = pd.to_numeric(taxonomy_df[column], errors="coerce")

    taxonomy_df["comportamento_famiglia_originale"] = taxonomy_df["comportamento_famiglia_originale"].combine_first(
        taxonomy_df["comportamento_famiglia"]
    )
    taxonomy_df["comportamento_famiglia_v2"] = taxonomy_df["comportamento_famiglia_v2"].combine_first(
        taxonomy_df["comportamento_famiglia"]
    )
    taxonomy_df["comportamento_famiglia"] = taxonomy_df["comportamento_famiglia_v2"].combine_first(
        taxonomy_df["comportamento_famiglia"]
    )
    taxonomy_df["comportamento_canonico"] = taxonomy_df["comportamento_canonico"].combine_first(
        taxonomy_df["comportamento_clean"]
    )
    taxonomy_df["canonical_behavior_id"] = taxonomy_df["canonical_behavior_id"].combine_first(
        taxonomy_df["comportamento_canonico"].map(build_canonical_behavior_id)
    )
    taxonomy_df["severity_rank_v2"] = taxonomy_df["severity_rank_v2"].clip(lower=1, upper=5)
    taxonomy_df["taxonomy_source"] = taxonomy_df["taxonomy_source"].fillna("external_v2")
    taxonomy_df["taxonomy_confidence"] = taxonomy_df["taxonomy_confidence"].fillna("medium")
    taxonomy_df["merge_source"] = taxonomy_df["merge_source"].fillna("external_v2")
    taxonomy_df["merge_confidence"] = taxonomy_df["merge_confidence"].fillna("medium")
    taxonomy_df["comportamento_tipo"] = taxonomy_df["comportamento_tipo"].fillna(BEHAVIOR_KIND_BEHAVIOR)
    taxonomy_df = taxonomy_df.loc[taxonomy_df["comportamento_clean"].notna()].reset_index(drop=True)

    confidence_rank = {"high": 0, "medium": 1, "low": 2}
    taxonomy_df["_merge_conf_rank"] = taxonomy_df["merge_confidence"].map(confidence_rank).fillna(3)
    taxonomy_df["_tax_conf_rank"] = taxonomy_df["taxonomy_confidence"].map(confidence_rank).fillna(3)
    taxonomy_df = taxonomy_df.sort_values(
        [
            "_merge_conf_rank",
            "_tax_conf_rank",
            "severity_rank_v2",
            "row_count_kept",
            "comportamento_clean",
            "comportamento_tipo",
        ],
        ascending=[True, True, True, False, True, True],
        na_position="last",
    ).reset_index(drop=True)

    lookup_df = taxonomy_df.drop_duplicates(["comportamento_clean", "comportamento_tipo"], keep="first").copy()
    duplicate_rows_removed = int(len(taxonomy_df) - len(lookup_df))
    taxonomy_df = taxonomy_df.drop(columns=["_merge_conf_rank", "_tax_conf_rank"])
    lookup_df = lookup_df.drop(columns=["_merge_conf_rank", "_tax_conf_rank"])

    profile = {
        "status": "loaded",
        "path": str(path),
        "input_rows": int(len(taxonomy_df)),
        "lookup_rows": int(len(lookup_df)),
        "duplicate_clean_type_rows_removed": duplicate_rows_removed,
        "unique_canonical_behaviors": int(lookup_df["canonical_behavior_id"].nunique()),
    }
    return taxonomy_df, lookup_df, profile


def fallback_behavior_taxonomy(
    behavior_clean: Any,
    behavior_kind: str,
    original_family: Any,
) -> dict[str, Any]:
    canonical_behavior = normalize_whitespace(behavior_clean)
    severity_rank = 5 if behavior_kind != BEHAVIOR_KIND_BEHAVIOR else 3
    return {
        "comportamento_canonico": canonical_behavior,
        "canonical_behavior_id": build_canonical_behavior_id(canonical_behavior),
        "comportamento_famiglia_originale": normalize_whitespace(original_family),
        "comportamento_famiglia": normalize_whitespace(original_family),
        "comportamento_famiglia_v2": normalize_whitespace(original_family),
        "comportamento_sottofamiglia_v2": None,
        "meccanismo_rischio_v2": None,
        "presidio_v2": None,
        "severity_rank_v2": severity_rank,
        "taxonomy_source": "fallback_rule",
        "taxonomy_confidence": "low",
        "canonical_cluster_size": 1,
        "canonical_row_count_total": np.nan,
        "old_family_conflict": "no",
        "old_families_cluster": normalize_whitespace(original_family),
        "merge_source": "fallback_rule",
        "merge_confidence": "low",
        "merge_note": "Comportamento non presente nel file di tassonomia v2; mantenuta classificazione corrente.",
        "review_flag": "taxonomy_missing",
    }


def resolve_behavior_taxonomy_record(
    behavior_clean: Any,
    behavior_kind: str,
    original_family: Any,
    taxonomy_lookup: dict[tuple[str, str], dict[str, Any]],
    taxonomy_clean_only_lookup: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    clean_label = normalize_whitespace(behavior_clean)
    if clean_label is None:
        return fallback_behavior_taxonomy(behavior_clean, behavior_kind, original_family)

    lookup_key = (clean_label, behavior_kind)
    taxonomy_record = taxonomy_lookup.get(lookup_key)
    if taxonomy_record is None:
        taxonomy_record = taxonomy_clean_only_lookup.get(clean_label)
    if taxonomy_record is None:
        return fallback_behavior_taxonomy(clean_label, behavior_kind, original_family)

    resolved = {column: taxonomy_record.get(column) for column in BEHAVIOR_TAXONOMY_OUTPUT_COLUMNS}
    resolved["comportamento_famiglia_originale"] = (
        resolved.get("comportamento_famiglia_originale") or normalize_whitespace(original_family)
    )
    resolved["comportamento_famiglia"] = resolved.get("comportamento_famiglia") or normalize_whitespace(original_family)
    resolved["comportamento_famiglia_v2"] = resolved.get("comportamento_famiglia_v2") or resolved["comportamento_famiglia"]
    resolved["comportamento_canonico"] = resolved.get("comportamento_canonico") or clean_label
    resolved["canonical_behavior_id"] = resolved.get("canonical_behavior_id") or build_canonical_behavior_id(
        resolved["comportamento_canonico"]
    )
    if pd.isna(resolved.get("severity_rank_v2")):
        resolved["severity_rank_v2"] = 5 if behavior_kind != BEHAVIOR_KIND_BEHAVIOR else 3
    return resolved


def build_area_from_text(raw_value: Any) -> str | None:
    raw = normalize_whitespace(raw_value)
    if raw is None:
        return None
    key = normalize_key(raw)
    if key is None:
        return None
    if "stamperia plastica" in key:
        return "Stamperia Plastica"
    if "stamperia metallo" in key or key == "cnc":
        return "Stamperia Metallo"
    if "verniciatura" in key:
        return "Verniciatura"
    if "galvanica" in key or "pulitura" in key:
        return "Galvanica/Pulitura"
    if "montaggio e finitura" in key or key == "taglio lenti":
        return "Montaggio e Finitura"
    if key.startswith("finitura -"):
        return "Montaggio e Finitura"
    if key.startswith("manutenzione"):
        return "Manutenzione"
    if "logistica" in key or "magazzino" in key or key == "confezionamento":
        return "Logistica e Magazzino"
    if key == "saldatura":
        return "Saldatura"
    if key.startswith("uffici"):
        return "Uffici"
    if key == "officina":
        return "Officina"
    if key == "ricambi":
        return "Ricambi"
    if key in {"aree comuni", "area comuni"}:
        return "Aree Comuni"
    if key == "tutta la sede":
        return "Tutta la sede"
    if key == "avanserie":
        return "Avanserie"
    if key == "prototipia":
        return "Prototipia"
    if key == "automation":
        return "Automation"
    if key == "sample":
        return "Sample"
    if key == "pianificazione":
        return "Pianificazione"
    if key == "altro":
        return "Altro"
    if "aree tecniche" in key or "tetti" in key:
        return "Aree tecniche/Tetti"
    return title_from_key(key)


def map_event_type(raw_value: Any) -> str | None:
    key = normalize_key(raw_value)
    if key is None or key == "tipologia":
        return None
    if "itiner" in key:
        return "Infortunio in itinere"
    if "smat audit" in key or "condizione insicura" in key or "azione insicura" in key or "bbs_annotazioni_strutturali" in key:
        return "SMAT Audit"
    if "primo soccorso" in key:
        return "Primo soccorso"
    if "infortunio con certificato" in key:
        return "Infortunio con certificato"
    if "incidente senza infortunio" in key:
        return "Incidente senza infortunio"
    if "malattia professionale" in key:
        return "Malattia Professionale"
    return title_from_key(key)


def map_problem_code(problem_it: Any, problem_en: Any) -> str:
    for value in (problem_it, problem_en):
        text = normalize_whitespace(value)
        if text is None:
            continue
        match = re.match(r"\s*([A-Za-z])\s*:", text)
        if match:
            return match.group(1).upper()
    return "UNKNOWN"


def map_body_part(raw_value: Any) -> str | None:
    key = normalize_key(raw_value)
    if key is None:
        return None
    replacements = {
        "occhi": "Occhi",
        "dita": "Dita",
        "mano": "Mano",
        "polso": "Polso",
        "gomito": "Gomito",
        "spalla": "Spalla",
        "collo": "Collo",
        "testa": "Testa",
        "ginocchio": "Ginocchio",
        "caviglia": "Caviglia",
        "piede": "Piede",
        "gamba": "Gamba",
        "schiena": "Schiena",
        "avambraccio": "Avambraccio",
        "viso": "Viso",
        "tronco": "Tronco",
        "braccio": "Braccio",
    }
    return replacements.get(key, title_from_key(key))


def map_body_part_macro(raw_value: Any) -> str:
    key = normalize_key(raw_value)
    if key is None or key == "altro":
        return UNKNOWN_INJURY_DETAIL
    if key in {"dita", "mano", "polso"}:
        return "Mani e dita"
    if key in {"avambraccio", "braccio", "gomito", "spalla"}:
        return "Arto superiore"
    if key in {"caviglia", "gamba", "gambe", "ginocchio", "piede", "piedi"}:
        return "Arto inferiore"
    if key in {"collo", "occhi", "testa", "viso"}:
        return "Testa / occhi / collo"
    if key in {"schiena", "torace", "tronco"}:
        return "Tronco / schiena"
    return UNKNOWN_INJURY_DETAIL


def map_injury_cause_macro(raw_value: Any) -> str:
    key = normalize_key(raw_value)
    if key is None or key == "altro":
        return UNKNOWN_INJURY_DETAIL
    if "movimenti incongrui" in key or "ergonomia" in key or "posture scorrette" in key or "posture scorrete" in key:
        return "Movimenti / postura"
    if "cadute" in key or "urti accidentali" in key:
        return "Cadute / urti"
    if (
        "presa di posizione pericolosa" in key
        or "distrazione" in key
        or "inosservanza di specifiche disposizioni" in key
        or "mancata utilizzazione di mezzi" in key
        or "lavoro non autorizzato" in key
    ):
        return "Procedure / attenzione"
    if (
        "difetto di costruzione" in key
        or "protezioni insufficienti" in key
        or "uso di attrezzi/utensili in cattivo stato" in key
        or "postazione di lavoro non idonea" in key
    ):
        return "Macchine / attrezzature"
    if "sistemazione pericolosa" in key or "piazzamento pezzi" in key or "caricamento" in key:
        return "Manipolazione / posizionamento"
    if "intervento di terze persone" in key:
        return "Interferenze / terzi"
    return UNKNOWN_INJURY_DETAIL


def clean_description(raw_value: Any) -> str | None:
    text = normalize_whitespace(raw_value)
    if text is None:
        return None
    text = text.replace("’", "'")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def description_for_tfidf(raw_value: Any) -> str:
    text = clean_description(raw_value)
    if text is None:
        return ""
    text = strip_accents(text.lower())
    text = re.sub(r"[^a-z0-9\s]", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def match_description_topics(raw_value: Any) -> list[str]:
    normalized_text = description_for_tfidf(raw_value)
    if not normalized_text:
        return []
    matched_topics = []
    for topic_name, patterns in DESCRIPTION_TOPIC_PATTERNS.items():
        if any(pattern in normalized_text for pattern in patterns):
            matched_topics.append(topic_name)
    return matched_topics


def strip_behavior_markup(raw_value: Any) -> str | None:
    text = normalize_whitespace(raw_value)
    if text is None:
        return None
    text = html.unescape(text)
    text = text.replace("’", "'").replace("`", "'").replace("^", " ")
    text = re.sub(r"<br\s*/?>", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def behavior_source_text(raw_behavior: Any, displayed_behavior: Any) -> str | None:
    return strip_behavior_markup(displayed_behavior) or strip_behavior_markup(raw_behavior)


def clean_behavior_label(raw_behavior: Any, displayed_behavior: Any) -> str | None:
    source = behavior_source_text(raw_behavior, displayed_behavior)
    if source is None:
        return None

    key = normalize_key(source) or ""
    if re.match(r"^\d{1,2}/\d{1,2}/\d{4}\s+caro osservatore", key) or key.startswith("caro osservatore") or key.startswith("caro osservatore bbs"):
        return "Messaggio osservatore"
    if key.startswith("hai visto o sei al corrente di incidenti sfiorati") or (
        key.startswith("hai visto") and "incidenti sfiorati" in key
    ):
        return "Domanda incidenti sfiorati"
    if key.startswith("hai visto altri comportamenti") or "non contenuti nella checklist" in key:
        return "Segnalazione altri comportamenti/condizioni"

    if ("quadro elettrico" in key or "quadri elettrici" in key) and ("visibili" in key or "raggiungibili" in key or "accessibili" in key):
        return "Quadri elettrici visibili e accessibili"
    if ("quadro elettrico" in key or "quadri elettrici" in key) and "chius" in key:
        return "Quadri elettrici chiusi"
    if "pavimento" in key and ("pulito" in key or "asciutto" in key):
        return "Pavimento pulito e asciutto"
    if ("postazione" in key or "scrivania" in key or "area di lavoro" in key or "banco" in key) and "pulit" in key:
        return "Area o postazione pulita"
    if ("postazione" in key or "scrivania" in key or "area di lavoro" in key or "banco" in key) and "ordine" in key:
        return "Area o postazione in ordine"
    if "rifiuti" in key and "industriali" in key and "contenitori" in key:
        return "Rifiuti industriali nei contenitori"
    if "rifiuti" in key and "urbani" in key and "contenitori" in key:
        return "Rifiuti urbani nei contenitori"
    if "rifiuti" in key and "contenitori" in key:
        return "Rifiuti nei contenitori"
    if "cammina" in key and "cuffie" in key:
        return "Cammina senza cuffie/auricolari"
    if "cammina" in key and ("non corre" in key or "senza correre" in key or "lentamente" in key):
        return "Cammina senza correre"
    if "guarda quello che fa" in key and "guarda dove mette le mani" in key:
        return "Guarda quello che fa e dove mette le mani"
    if key == "guarda quello che fa":
        return "Guarda quello che fa"
    if key == "guarda dove va":
        return "Guarda dove va"
    if key.startswith("otoprotettori") or ("inserti monouso" in key and "archetti" in key):
        return "Otoprotettori"
    if "occhiali di protezione" in key:
        return "Occhiali di protezione"
    if "scarpe antinfortunistiche" in key:
        return "Scarpe antinfortunistiche"
    if "scarpe basse e chiuse" in key:
        return "Scarpe basse e chiuse"
    if "guant" in key and ("bassa protezione al taglio" in key or ("spalmat" in key and "bassa" in key)):
        return "Guanti protezione meccanica bassa"
    if "guant" in key and ("media/alta" in key or "medio/alta" in key or "alta protezione al taglio" in key):
        return "Guanti protezione meccanica media/alta"
    if "guant" in key and ("chimic" in key or "nitrile" in key or "anti-acido" in key):
        return "Guanti protezione chimica"
    if ("guant" in key or "manicott" in key) and ("calore" in key or "alte temperature" in key or "contatto accidentale" in key or "contatto (con olii)" in key):
        return "Guanti / manicotti protezione calore"
    if "aria compressa" in key or "getto d'aria compressa" in key:
        return "Uso corretto aria compressa"
    if "taglierino" in key or "lama retrattile" in key or ("forbice" in key and "aprire i pacchi" in key):
        return "Uso corretto taglierino / lama"

    return source


def classify_behavior_kind(cleaned_behavior: Any, activity_clean: Any) -> str:
    behavior_key = normalize_key(cleaned_behavior)
    activity_key = normalize_key(activity_clean)
    if behavior_key is None:
        return BEHAVIOR_KIND_MISSING
    if cleaned_behavior == "Messaggio osservatore" or activity_key in {"messaggio", "messaggio a seguito di incidenti"}:
        return BEHAVIOR_KIND_MESSAGE
    if cleaned_behavior == "Domanda incidenti sfiorati" or activity_key == "incidenti sfiorati":
        return BEHAVIOR_KIND_NEAR_MISS_PROMPT
    if cleaned_behavior == "Segnalazione altri comportamenti/condizioni":
        return BEHAVIOR_KIND_OPEN_SIGNAL_PROMPT
    if activity_key in {"altri comportamenti", "comportamento da inserire", "b - da non usare"} and behavior_key in {
        "segnalazione altri comportamenti/condizioni",
        "comportamento da inserire",
    }:
        return BEHAVIOR_KIND_OPEN_SIGNAL_PROMPT
    return BEHAVIOR_KIND_BEHAVIOR


def map_behavior_family(cleaned_behavior: Any, behavior_kind: str, activity_family: Any) -> str:
    if behavior_kind == BEHAVIOR_KIND_MESSAGE:
        return "Comunicazione / Messaggi"
    if behavior_kind in {BEHAVIOR_KIND_NEAR_MISS_PROMPT, BEHAVIOR_KIND_OPEN_SIGNAL_PROMPT}:
        return "Segnalazioni / Near miss / Checklist"

    key = normalize_key(cleaned_behavior) or ""
    if (
        "scala" in key
        or "palchetto" in key
        or "corrimano" in key
        or "gradino" in key
        or "pioli" in key
        or "3 punti di appoggio" in key
        or "tre punti di appoggio" in key
        or "coperture" in key
        or "spazi confinati" in key
        or "linea vita" in key
        or "moschetton" in key
        or "imbracatura" in key
        or re.search(r"\bple\b", key) is not None
        or "parapetto" in key
    ):
        return "Lavori Speciali / Accessi"
    if (
        "tecnico che lavora" in key
        or "manutentore" in key
        or "attrezzista che lavora" in key
        or "piu di un metro dal" in key
        or "piu di un metro dall" in key
        or "fuori dall'area" in key
        or "sgombra da persone" in key
        or ("area di lavoro" in key and ("delimitat" in key or "segnalat" in key))
    ):
        return "Interferenze / Coordinamento"
    if "occhiali" in key or "visiera" in key:
        return "DPI occhi / viso"
    if key.startswith("otoprotettori"):
        return "DPI udito"
    if "scarpe" in key:
        return "DPI piedi / abbigliamento"
    if (
        "felpa" in key
        or "abbigli" in key
        or "capelli lunghi" in key
        or "tuta" in key
        or "abiti" in key
        or "svolazzanti" in key
        or "monili" in key
        or "pendenti" in key
        or "anelli" in key
        or "braccialetto" in key
        or "elmetto" in key
        or "maschera" in key
        or "mascherina" in key
        or "tyvek" in key
        or "autorespir" in key
    ):
        return "DPI / Abbigliamento"
    if "guant" in key or "manicott" in key:
        return "DPI mani / protezioni"
    if "vie di fuga" in key or "tagliafuoco" in key or "antincendio" in key or "quadro elettrico" in key or "quadri elettrici" in key or "pulsante di emergenza" in key:
        return "Emergenza / Elettrico / Vie di fuga"
    if (
        "rifiuti" in key
        or "pavimento" in key
        or "pulit" in key
        or "pulisc" in key
        or "ordine" in key
        or "effetti personali" in key
        or "trucioli" in key
        or "sfridi" in key
        or "scarti" in key
        or "vaschette di contenimento" in key
        or "vaschette" in key
        or "posaggio" in key
    ):
        return "Ordine / Pulizia / 5S"
    if "stocc" in key or "scaffal" in key or "tavoli" in key or "bancali" in key or "materiali" in key or "perimetro" in key or "spazi delimitati" in key or "recipienti" in key:
        return "Stoccaggio / Materiali"
    if "cammina" in key or "percor" in key or "pedon" in key or "cellulare" in key or key == "guarda dove va" or key == "guarda quello che fa" or "guarda quello che fa e dove mette le mani" in key:
        return "Spostamenti / Attenzione"
    if "carrell" in key or "transpallet" in key or "agv" in key or "mezzi in movimento" in key or "carico" in key or "scarica" in key:
        return "Carrelli / Transpallet / Mezzi"
    if (
        "schiena" in key
        or "busto" in key
        or "ginocchia" in key
        or "ruota tutto il corpo" in key
        or "movimenta" in key
        or "solleva" in key
        or "abbassa il carico" in key
        or "trasporta il carico" in key
        or "sgabello" in key
        or "seduto" in key
        or "polso" in key
        or "rotazione del pollice" in key
        or "si siede" in key
        or "sedia" in key
        or "panchina" in key
    ):
        return "Movimentazione Manuale / Ergonomia"
    if (
        "sicurezze" in key
        or "organi in movimento" in key
        or "macchina" in key
        or "attrezzista" in key
        or "attrezzaggio" in key
        or "carroponte" in key
        or "forno" in key
        or "tramoggia" in key
        or "sgm" in key
        or "schiacci" in key
        or "raggio di chiusura" in key
        or "parati" in key
        or "barriera" in key
        or "portello" in key
        or "protezion" in key
        or "plexiglass" in key
        or "incepp" in key
        or "rulliera" in key
        or "buratto" in key
        or "vassoio" in key
        or "compattatore" in key
        or "leva di fissaggio" in key
        or "fissaggio delle attrezzature" in key
        or "morsetto" in key
        or "telai" in key
        or "ganci" in key
    ):
        return "Macchine / Sicurezze / Attrezzaggio"
    if (
        "aria compressa" in key
        or "taglierino" in key
        or "lama" in key
        or "trapano" in key
        or "utensili" in key
        or "forbice" in key
        or "avvitatore" in key
        or "avvitatura" in key
        or "cacciavite" in key
        or "punteruolo" in key
        or "tronchesino" in key
        or "scapsula" in key
    ):
        return "Utensili / Taglio / Aria Compressa"
    if "chimic" in key or "aspiraz" in key or "cappa" in key or "diluenti" in key or "inchiostri" in key or "sverniciante" in key or "contenitori delle sostanze" in key or "bidoni" in key:
        return "Chimici / Aspirazione / Lavaggi"
    if (
        "pressa" in key
        or "essiccatore" in key
        or "caraffa" in key
        or "panetto" in key
        or "fil di ferro" in key
        or "granuli" in key
        or "pasta" in key
    ):
        return "Produzione / Processo"

    return BEHAVIOR_ACTIVITY_FAMILY_FALLBACK.get(normalize_whitespace(activity_family) or "", "Altro")


def clean_activity_label(raw_value: Any) -> str | None:
    raw = normalize_whitespace(raw_value)
    if raw is None:
        return None
    text = raw.replace("’", "'").replace("^", "").replace("_", " ")
    text = re.sub(r"\bAGO(?:\s+[A-Z]+)?\b", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\s*-\s*(INI|FIN|TRAS|OFF|UFF|NO INI|CNC)\b", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+", " ", text)
    text = re.sub(r"\s*-\s*-\s*", " - ", text)
    text = text.strip(" -")
    text = re.sub(r"\s+", " ", text).strip()
    normalized = normalize_key(text) or ""

    if normalized.startswith("dpi - "):
        return "DPI"
    if normalized in {"gestione sostanze chimiche", "uso sostanze chimiche"}:
        return "Gestione / uso sostanze chimiche"
    if normalized in {
        "movimentazione con carrello e transpallet",
        "movimentazione con carrello o transpallet",
        "movimentazione muletto uomo a bordo e transpallet elettrico",
        "movimentazione con transpallet pantografo",
    }:
        return "Movimentazione con ausili"
    if normalized in {
        "uso del taglierino",
        "uso del taglia-reggette",
        "uso del taglierino/taglia-reggette",
    }:
        return "Uso taglierino / taglia-reggette"

    canonical_override = ACTIVITY_CANONICAL_LABEL_OVERRIDES.get(normalized)
    return canonical_override or text


def map_activity_family(raw_activity: Any, cleaned_activity: Any) -> str:
    text = normalize_key(cleaned_activity or raw_activity) or ""
    override = ACTIVITY_FAMILY_OVERRIDES.get(text)
    if override:
        return override
    for family, keywords in ACTIVITY_FAMILY_RULES:
        if any(keyword in text for keyword in keywords):
            return family
    return "Altro"


def safe_div(numerator: float, denominator: float) -> float:
    if denominator == 0:
        return 0.0
    return numerator / denominator


def months_since_last_positive(series: pd.Series) -> pd.Series:
    result: list[float] = []
    distance: int | None = None
    positive_flags = series.gt(0).to_numpy(dtype=bool)
    previous_flags = np.concatenate(([False], positive_flags[:-1]))
    for flag in previous_flags:
        if flag:
            distance = 0
            result.append(0.0)
        elif distance is None:
            result.append(float("nan"))
        else:
            distance += 1
            result.append(float(distance))
    return pd.Series(result, index=series.index, dtype="float64")


def serialize_scalar(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (np.integer, np.floating)):
        return value.item()
    if isinstance(value, pd.Timestamp):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, (pd.Series, pd.Index)):
        return value.tolist()
    return value


def is_spurious_header_row(record: dict[str, Any], header_names: Iterable[str]) -> bool:
    matches = 0
    populated = 0
    header_lookup = {normalize_key(header) for header in header_names}
    for key, value in record.items():
        normalized = normalize_key(value)
        if normalized is None:
            continue
        populated += 1
        if normalized in header_lookup:
            matches += 1
    return populated > 0 and matches / populated >= 0.3


def column_or_empty(df: pd.DataFrame, *candidates: str) -> pd.Series:
    for candidate in candidates:
        if candidate in df.columns:
            return df[candidate]
    return pd.Series([None] * len(df), index=df.index, dtype="object")


def load_events_dataframe(path: Path) -> tuple[pd.DataFrame, dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    raw_headers = [normalize_whitespace(value) or f"col_{idx}" for idx, value in enumerate(next(rows), start=1)]
    records: list[dict[str, Any]] = []
    dropped_header_like_rows = 0
    for row in rows:
        record = {header: coerce_null(value) for header, value in zip(raw_headers, row)}
        if is_spurious_header_row(record, raw_headers):
            dropped_header_like_rows += 1
            continue
        records.append(record)
    df = pd.DataFrame.from_records(records)
    profile = {
        "input_rows": len(records) + dropped_header_like_rows,
        "header_like_rows_removed": dropped_header_like_rows,
        "input_columns": len(raw_headers),
    }
    return df, profile


def clean_events_dataframe(df: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, Any], pd.DataFrame]:
    original_columns = list(df.columns)
    all_empty_columns = [column for column in df.columns if df[column].isna().all()]
    constant_columns = [
        column
        for column in df.columns
        if column not in all_empty_columns and df[column].nunique(dropna=True) <= 1
    ]

    working = df.drop(columns=all_empty_columns + constant_columns).copy()
    working["event_date"] = pd.to_datetime(column_or_empty(working, "Data rilievo"), errors="coerce")
    working["event_month"] = working["event_date"].dt.to_period("M").dt.to_timestamp()
    working["sede"] = column_or_empty(working, "Sede").map(normalize_whitespace)
    working["stabilimento"] = column_or_empty(working, "Stabilimento").map(normalize_whitespace)
    working["societa"] = column_or_empty(working, "Società").map(normalize_whitespace)
    working["descrizione_clean"] = column_or_empty(working, "Descrizione", "Descrizione ").map(clean_description)
    working["descrizione_tfidf"] = working["descrizione_clean"].map(description_for_tfidf)
    working["descrizione_topic_labels"] = working["descrizione_clean"].map(
        lambda value: " | ".join(match_description_topics(value))
    )
    working["tipologia_raw"] = column_or_empty(working, "Tipologia").map(normalize_whitespace)
    working["tipologia_canonica"] = working["tipologia_raw"].map(map_event_type)
    working["problem_code"] = [
        map_problem_code(problem_it, problem_en)
        for problem_it, problem_en in zip(
            column_or_empty(working, "Tipologia del problema"),
            column_or_empty(working, "Tipologia problema inglese"),
        )
    ]
    working["problem_family"] = working["problem_code"].map(PROBLEM_FAMILY_LABELS).fillna(PROBLEM_FAMILY_LABELS["UNKNOWN"])
    working["area_reparto_raw"] = column_or_empty(working, "Area/Reparto").map(normalize_whitespace)
    working["area_canonica"] = working["area_reparto_raw"].map(build_area_from_text)
    working["dettaglio_area_clean"] = column_or_empty(working, "Dettaglio Area", " Dettaglio Area").map(normalize_whitespace)
    working["parte_del_corpo_clean"] = column_or_empty(working, "Parte del corpo").map(map_body_part)
    working["parte_del_corpo_macro"] = working["parte_del_corpo_clean"].map(map_body_part_macro)
    working["tipologia_causa_clean"] = column_or_empty(working, "Tipologia della causa").map(normalize_whitespace)
    working["tipologia_causa_macro"] = working["tipologia_causa_clean"].map(map_injury_cause_macro)
    working["descrizione_lesione_clean"] = column_or_empty(working, "Descrizione della lesione").map(normalize_whitespace)
    working["agente_materiale_clean"] = column_or_empty(working, "Agente materiale").map(normalize_whitespace)
    working["giorni_persi_raw"] = column_or_empty(working, "Giorni persi").map(to_float)
    working["is_covid_problem"] = working["problem_code"].eq("V")
    working["is_in_itinere"] = working["tipologia_canonica"].eq("Infortunio in itinere")

    excluded_mask = working["is_covid_problem"] | working["is_in_itinere"] | working["event_date"].isna()
    excluded_counts = {
        "covid_rows_removed": int(working["is_covid_problem"].sum()),
        "in_itinere_rows_removed": int(working["is_in_itinere"].sum()),
        "missing_event_date_removed": int(working["event_date"].isna().sum()),
    }

    cleaned = working.loc[~excluded_mask].copy()

    giorni_non_null = cleaned["giorni_persi_raw"].dropna()
    giorni_cap = float(giorni_non_null.quantile(0.99)) if not giorni_non_null.empty else None
    if giorni_cap is not None:
        cleaned["giorni_persi_capped"] = cleaned["giorni_persi_raw"].clip(upper=giorni_cap)
        cleaned["giorni_persi_outlier_flag"] = cleaned["giorni_persi_raw"].gt(giorni_cap).fillna(False)
    else:
        cleaned["giorni_persi_capped"] = cleaned["giorni_persi_raw"]
        cleaned["giorni_persi_outlier_flag"] = False
    cleaned["giorni_persi_log1p"] = cleaned["giorni_persi_capped"].fillna(0).map(lambda value: math.log1p(value))

    cleaned["is_smat_audit"] = cleaned["tipologia_canonica"].eq("SMAT Audit").astype(int)
    cleaned["is_first_aid"] = cleaned["tipologia_canonica"].eq("Primo soccorso").astype(int)
    cleaned["is_certified_injury"] = cleaned["tipologia_canonica"].eq("Infortunio con certificato").astype(int)
    cleaned["is_incident_without_injury"] = cleaned["tipologia_canonica"].eq("Incidente senza infortunio").astype(int)
    cleaned["is_occupational_disease"] = cleaned["tipologia_canonica"].eq("Malattia Professionale").astype(int)
    cleaned["is_injury_like"] = (cleaned["is_first_aid"] | cleaned["is_certified_injury"]).astype(int)

    columns_to_drop_before_rename = [
        "Sede",
        "Società",
        "Stabilimento",
        "Descrizione",
        "Descrizione ",
        "Tipologia",
        "Area/Reparto",
        "Dettaglio Area",
        " Dettaglio Area",
        "Parte del corpo",
        "Tipologia della causa",
        "Descrizione della lesione",
        "Agente materiale",
        "Giorni persi",
    ]
    cleaned = cleaned.drop(columns=[column for column in columns_to_drop_before_rename if column in cleaned.columns])

    renamed_columns = {column: snake_case(column) for column in original_columns if column in cleaned.columns}
    cleaned = cleaned.rename(columns=renamed_columns)

    selected_columns = [
        "id",
        "event_date",
        "event_month",
        "sede",
        "societa",
        "stabilimento",
        "tipologia_raw",
        "tipologia_canonica",
        "tipologia_problema_inglese",
        "tipologia_del_problema",
        "problem_code",
        "problem_family",
        "area_reparto_raw",
        "area_canonica",
        "dettaglio_area_clean",
        "mansione_attivita",
        "descrizione_clean",
        "descrizione_tfidf",
        "descrizione_topic_labels",
        "parte_del_corpo_clean",
        "parte_del_corpo_macro",
        "tipologia_causa_clean",
        "tipologia_causa_macro",
        "descrizione_lesione_clean",
        "agente_materiale_clean",
        "giorni_persi_raw",
        "giorni_persi_capped",
        "giorni_persi_log1p",
        "giorni_persi_outlier_flag",
        "is_smat_audit",
        "is_first_aid",
        "is_certified_injury",
        "is_incident_without_injury",
        "is_occupational_disease",
        "is_injury_like",
    ]
    selected_columns = [column for column in selected_columns if column in cleaned.columns]
    cleaned = cleaned[selected_columns].copy()

    profile = {
        "rows_after_cleaning": int(len(cleaned)),
        "columns_after_cleaning": int(cleaned.shape[1]),
        "all_empty_columns_removed": all_empty_columns,
        "constant_columns_removed": constant_columns,
        "giorni_persi_cap_p99": giorni_cap,
        "days_lost_outliers_flagged": int(cleaned["giorni_persi_outlier_flag"].sum()),
        **excluded_counts,
    }

    area_mapping = (
        working[["area_reparto_raw", "area_canonica"]]
        .dropna(subset=["area_reparto_raw"])
        .drop_duplicates()
        .sort_values(["area_canonica", "area_reparto_raw"])
        .reset_index(drop=True)
    )
    return cleaned, profile, area_mapping


def build_event_semantic_features(
    events_df: pd.DataFrame,
    output_dir: Path,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    semantic_base = events_df.loc[
        events_df["descrizione_tfidf"].fillna("").str.len() > 0,
        ["area_canonica", "event_month", "problem_family", "descrizione_tfidf"],
    ].copy()
    if semantic_base.empty:
        return pd.DataFrame(columns=["area_canonica", "event_month"]), pd.DataFrame()

    vectorizer = TfidfVectorizer(
        min_df=5,
        max_df=0.85,
        max_features=600,
        ngram_range=(1, 2),
        stop_words=sorted(ITALIAN_STOPWORDS),
    )
    matrix = vectorizer.fit_transform(semantic_base["descrizione_tfidf"])

    terms = np.array(vectorizer.get_feature_names_out())
    semantic_rows = []
    grouped_problem = semantic_base.groupby("problem_family").indices
    for family, positions in grouped_problem.items():
        family_slice = matrix[positions]
        if family_slice.shape[0] == 0:
            continue
        weights = np.asarray(family_slice.mean(axis=0)).ravel()
        top_indices = weights.argsort()[::-1][:10]
        for rank, index in enumerate(top_indices, start=1):
            if weights[index] <= 0:
                continue
            semantic_rows.append(
                {
                    "problem_family": family,
                    "rank": rank,
                    "term": terms[index],
                    "mean_tfidf": float(weights[index]),
                }
            )
    semantic_terms_df = pd.DataFrame(semantic_rows)
    if not semantic_terms_df.empty:
        semantic_terms_df.to_csv(output_dir / "semantic_insights_by_problem.csv", index=False)

    n_components = min(5, max(1, matrix.shape[1] - 1))
    if n_components <= 0:
        return pd.DataFrame(columns=["area_canonica", "event_month"]), semantic_terms_df
    svd = TruncatedSVD(n_components=n_components, random_state=42)
    components = svd.fit_transform(matrix)
    for index in range(n_components):
        semantic_base[f"evt_semantic_component_{index + 1}"] = components[:, index]

    semantic_columns = [column for column in semantic_base.columns if column.startswith("evt_semantic_component_")]
    area_month_semantics = (
        semantic_base.groupby(["area_canonica", "event_month"], dropna=False)[semantic_columns]
        .mean()
        .reset_index()
    )
    return area_month_semantics, semantic_terms_df


def build_event_description_topic_features(
    events_df: pd.DataFrame,
    output_dir: Path,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    topic_rows = []
    for _, row in events_df.loc[events_df["descrizione_topic_labels"].fillna("").str.len() > 0].iterrows():
        labels = [label.strip() for label in str(row["descrizione_topic_labels"]).split("|") if normalize_whitespace(label)]
        adverse_flag = int(
            row["is_first_aid"] == 1
            or row["is_certified_injury"] == 1
            or row["is_incident_without_injury"] == 1
        )
        for label in labels:
            topic_rows.append(
                {
                    "id": row["id"],
                    "area_canonica": row["area_canonica"],
                    "event_month": row["event_month"],
                    "problem_family": row["problem_family"],
                    "problem_code": row["problem_code"],
                    "tipologia_canonica": row["tipologia_canonica"],
                    "description_topic": label,
                    "is_adverse_event": adverse_flag,
                    "is_smat_audit": int(row["is_smat_audit"]),
                }
            )
    topic_event_df = pd.DataFrame(topic_rows)
    if topic_event_df.empty:
        return (
            pd.DataFrame(columns=["area_canonica", "event_month"]),
            pd.DataFrame(columns=["description_topic"]),
            pd.DataFrame(columns=["problem_family", "description_topic"]),
        )

    topic_summary_df = (
        topic_event_df.groupby("description_topic", dropna=False)
        .agg(
            tagged_event_rows=("id", "count"),
            unique_events=("id", "nunique"),
            unique_areas=("area_canonica", "nunique"),
            adverse_rows=("is_adverse_event", "sum"),
            smat_rows=("is_smat_audit", "sum"),
        )
        .reset_index()
        .sort_values(["unique_events", "description_topic"], ascending=[False, True])
        .reset_index(drop=True)
    )
    topic_summary_df.to_csv(output_dir / "description_topic_summary.csv", index=False)

    topic_by_problem_df = (
        topic_event_df.groupby(["problem_family", "description_topic"], dropna=False)
        .agg(unique_events=("id", "nunique"))
        .reset_index()
        .sort_values(["problem_family", "unique_events", "description_topic"], ascending=[True, False, True])
        .reset_index(drop=True)
    )
    topic_by_problem_df.to_csv(output_dir / "description_topic_by_problem.csv", index=False)

    area_month_frames = []
    for metric_name, value_column in [("count", "id"), ("adverse_count", "is_adverse_event")]:
        if metric_name == "count":
            grouped = (
                topic_event_df.groupby(["area_canonica", "event_month", "description_topic"], dropna=False)
                .agg(metric_value=("id", "count"))
                .reset_index()
            )
        else:
            grouped = (
                topic_event_df.groupby(["area_canonica", "event_month", "description_topic"], dropna=False)
                .agg(metric_value=("is_adverse_event", "sum"))
                .reset_index()
            )
        pivoted = grouped.pivot_table(
            index=["area_canonica", "event_month"],
            columns="description_topic",
            values="metric_value",
            fill_value=0,
        )
        pivoted.columns = [
            f"evt_desc_topic_{snake_case(str(column))}_{metric_name}"
            for column in pivoted.columns
        ]
        area_month_frames.append(pivoted)

    topic_presence_df = (
        topic_event_df.groupby(["area_canonica", "event_month"], dropna=False)
        .agg(
            evt_desc_topic_tagged_rows=("id", "count"),
            evt_desc_topic_unique_topics=("description_topic", "nunique"),
        )
        .reset_index()
    )
    area_month_topic_df = topic_presence_df.merge(
        pd.concat(area_month_frames, axis=1).reset_index(),
        on=["area_canonica", "event_month"],
        how="left",
    )
    area_month_topic_df.to_csv(output_dir / "description_topic_area_month.csv", index=False)
    return area_month_topic_df, topic_summary_df, topic_by_problem_df


def build_event_area_month_dataset(
    events_df: pd.DataFrame,
    semantic_df: pd.DataFrame,
    description_topic_df: pd.DataFrame | None = None,
) -> pd.DataFrame:
    base = (
        events_df.groupby(["area_canonica", "event_month"], dropna=False)
        .agg(
            evt_total_rows=("id", "count"),
            evt_smat_audit_count=("is_smat_audit", "sum"),
            evt_first_aid_count=("is_first_aid", "sum"),
            evt_certified_injury_count=("is_certified_injury", "sum"),
            evt_incident_without_injury_count=("is_incident_without_injury", "sum"),
            evt_occupational_disease_count=("is_occupational_disease", "sum"),
            evt_injury_like_count=("is_injury_like", "sum"),
            evt_days_lost_sum=("giorni_persi_capped", "sum"),
            evt_days_lost_avg=("giorni_persi_capped", "mean"),
            evt_days_lost_p95=("giorni_persi_capped", lambda series: series.quantile(0.95) if series.notna().any() else np.nan),
        )
        .reset_index()
    )

    problem_counts = (
        events_df.pivot_table(
            index=["area_canonica", "event_month"],
            columns="problem_code",
            values="id",
            aggfunc="count",
            fill_value=0,
        )
        .reset_index()
    )
    problem_columns = []
    for column in list(problem_counts.columns):
        if column in {"area_canonica", "event_month"}:
            continue
        new_name = f"evt_problem_{snake_case(str(column))}_count"
        problem_counts = problem_counts.rename(columns={column: new_name})
        problem_columns.append(new_name)

    merged = base.merge(problem_counts, on=["area_canonica", "event_month"], how="left")
    if not semantic_df.empty:
        merged = merged.merge(semantic_df, on=["area_canonica", "event_month"], how="left")
    if description_topic_df is not None and not description_topic_df.empty:
        merged = merged.merge(description_topic_df, on=["area_canonica", "event_month"], how="left")
    for column in merged.columns:
        if column.startswith("evt_"):
            merged[column] = merged[column].fillna(0)
    return merged.sort_values(["area_canonica", "event_month"]).reset_index(drop=True)


def stream_clean_observations(
    path: Path,
    output_path: Path,
    behavior_taxonomy_lookup_df: pd.DataFrame | None = None,
) -> tuple[pd.DataFrame, dict[str, Any], pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    headers = [normalize_whitespace(value) or f"col_{idx}" for idx, value in enumerate(next(rows), start=1)]
    index = {header: idx for idx, header in enumerate(headers)}

    taxonomy_lookup: dict[tuple[str, str], dict[str, Any]] = {}
    taxonomy_clean_only_lookup: dict[str, dict[str, Any]] = {}
    if behavior_taxonomy_lookup_df is not None and not behavior_taxonomy_lookup_df.empty:
        for record in behavior_taxonomy_lookup_df.to_dict("records"):
            clean_label = normalize_whitespace(record.get("comportamento_clean"))
            behavior_kind = normalize_whitespace(record.get("comportamento_tipo")) or BEHAVIOR_KIND_BEHAVIOR
            if clean_label is None:
                continue
            taxonomy_lookup[(clean_label, behavior_kind)] = record
        clean_counts = Counter(clean for clean, _kind in taxonomy_lookup.keys())
        for (clean_label, _behavior_kind), record in taxonomy_lookup.items():
            if clean_counts[clean_label] == 1:
                taxonomy_clean_only_lookup[clean_label] = record

    group_mapping_counter: Counter[tuple[str, str]] = Counter()
    activity_mapping_counter: Counter[tuple[str, str, str]] = Counter()
    activity_total_counter: Counter[tuple[str, str, str]] = Counter()
    activity_covid_counter: Counter[tuple[str, str, str]] = Counter()
    cleaned_activity_counter: Counter[tuple[str, str]] = Counter()
    cleaned_activity_raw_values: defaultdict[tuple[str, str], set[str]] = defaultdict(set)
    behavior_mapping_counter: Counter[tuple[Any, ...]] = Counter()
    cleaned_behavior_counter: Counter[tuple[Any, ...]] = Counter()
    cleaned_behavior_raw_values: defaultdict[tuple[Any, ...], set[str]] = defaultdict(set)
    canonical_behavior_counter: Counter[tuple[Any, ...]] = Counter()
    canonical_behavior_clean_variants: defaultdict[tuple[Any, ...], set[str]] = defaultdict(set)
    canonical_behavior_raw_variants: defaultdict[tuple[Any, ...], set[str]] = defaultdict(set)

    monthly_stats: dict[tuple[str, pd.Timestamp], dict[str, Any]] = {}
    unique_observations: dict[tuple[str, pd.Timestamp], set[Any]] = defaultdict(set)

    input_rows = 0
    covid_rows_removed = 0
    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "id_osservazione",
                "id_comportamento_osservato",
                "data",
                "event_month",
                "area_canonica",
                "gruppo_raw",
                "gruppo_clean",
                "area_osservazione",
                "attivita_raw",
                "attivita_clean",
                "attivita_famiglia",
                "comportamento",
                "comportamento_visualizzato",
                "comportamento_clean",
                "comportamento_canonico",
                "canonical_behavior_id",
                "comportamento_famiglia_originale",
                "comportamento_famiglia",
                "comportamento_famiglia_v2",
                "comportamento_sottofamiglia_v2",
                "meccanismo_rischio_v2",
                "presidio_v2",
                "severity_rank_v2",
                "taxonomy_source",
                "taxonomy_confidence",
                "canonical_cluster_size",
                "canonical_row_count_total",
                "old_family_conflict",
                "old_families_cluster",
                "merge_source",
                "merge_confidence",
                "merge_note",
                "review_flag",
                "comportamento_tipo",
                "comportamento_osservabile",
                "comportamenti_sicuri",
                "comportamenti_non_sicuri",
                "unsafe_flag",
            ],
        )
        writer.writeheader()

        for row in rows:
            input_rows += 1
            raw_activity = row[index["Attività"]] if "Attività" in index else None
            raw_group = row[index["Gruppo"]] if "Gruppo" in index else None
            raw_date = row[index["Data"]] if "Data" in index else None
            safe_value = to_float(row[index["Comportamenti sicuri"]]) if "Comportamenti sicuri" in index else 0.0
            unsafe_value = to_float(row[index["Comportamenti non sicuri"]]) if "Comportamenti non sicuri" in index else 0.0
            safe_value = 0.0 if safe_value is None else safe_value
            unsafe_value = 0.0 if unsafe_value is None else unsafe_value

            raw_activity_text = normalize_whitespace(raw_activity)
            activity_clean = clean_activity_label(raw_activity_text)
            activity_family = map_activity_family(raw_activity_text, activity_clean)
            raw_behavior_text = normalize_whitespace(row[index["Comportamento"]]) if "Comportamento" in index else None
            displayed_behavior_text = normalize_whitespace(row[index["Comportamento visualizzato"]]) if "Comportamento visualizzato" in index else None
            behavior_source = behavior_source_text(raw_behavior_text, displayed_behavior_text)
            behavior_clean = clean_behavior_label(raw_behavior_text, displayed_behavior_text)
            behavior_kind = classify_behavior_kind(behavior_clean, activity_clean)
            original_behavior_family = map_behavior_family(behavior_clean, behavior_kind, activity_family)
            behavior_taxonomy = resolve_behavior_taxonomy_record(
                behavior_clean,
                behavior_kind,
                original_behavior_family,
                taxonomy_lookup,
                taxonomy_clean_only_lookup,
            )
            behavior_family = normalize_whitespace(behavior_taxonomy.get("comportamento_famiglia")) or original_behavior_family
            if raw_activity_text:
                activity_total_counter[(raw_activity_text, activity_clean or "", activity_family)] += 1

            if raw_activity_text and "covid" in (normalize_key(raw_activity_text) or ""):
                activity_covid_counter[(raw_activity_text, activity_clean or "", activity_family)] += 1
                covid_rows_removed += 1
                continue

            group_clean = normalize_whitespace(raw_group)
            area_canonica = build_area_from_text(group_clean)
            event_month = month_start(raw_date)
            if area_canonica is None or event_month is None:
                continue

            raw_group_text = normalize_whitespace(raw_group)
            group_mapping_counter[(raw_group_text or "", area_canonica)] += 1
            activity_mapping_counter[(raw_activity_text or "", activity_clean or "", activity_family)] += 1
            cleaned_activity_counter[(activity_clean or "", activity_family)] += 1
            if raw_activity_text:
                cleaned_activity_raw_values[(activity_clean or "", activity_family)].add(raw_activity_text)
            mapping_key = (
                behavior_source or "",
                behavior_clean or "",
                behavior_taxonomy.get("comportamento_canonico") or "",
                behavior_taxonomy.get("canonical_behavior_id") or "",
                behavior_taxonomy.get("comportamento_famiglia_originale") or "",
                behavior_family,
                behavior_taxonomy.get("comportamento_famiglia_v2") or behavior_family,
                behavior_taxonomy.get("comportamento_sottofamiglia_v2") or "",
                behavior_taxonomy.get("meccanismo_rischio_v2") or "",
                behavior_taxonomy.get("presidio_v2") or "",
                behavior_taxonomy.get("severity_rank_v2"),
                behavior_taxonomy.get("taxonomy_source") or "",
                behavior_taxonomy.get("taxonomy_confidence") or "",
                behavior_taxonomy.get("canonical_cluster_size"),
                behavior_taxonomy.get("canonical_row_count_total"),
                behavior_taxonomy.get("old_family_conflict") or "",
                behavior_taxonomy.get("old_families_cluster") or "",
                behavior_taxonomy.get("merge_source") or "",
                behavior_taxonomy.get("merge_confidence") or "",
                behavior_taxonomy.get("merge_note") or "",
                behavior_taxonomy.get("review_flag") or "",
                behavior_kind,
            )
            cleaned_key = mapping_key[1:]
            canonical_key = (
                behavior_taxonomy.get("comportamento_canonico") or "",
                behavior_taxonomy.get("canonical_behavior_id") or "",
                behavior_family,
                behavior_taxonomy.get("comportamento_sottofamiglia_v2") or "",
                behavior_taxonomy.get("meccanismo_rischio_v2") or "",
                behavior_taxonomy.get("presidio_v2") or "",
                behavior_taxonomy.get("severity_rank_v2"),
                behavior_taxonomy.get("taxonomy_source") or "",
                behavior_taxonomy.get("taxonomy_confidence") or "",
                behavior_taxonomy.get("merge_source") or "",
                behavior_taxonomy.get("merge_confidence") or "",
                behavior_taxonomy.get("review_flag") or "",
                behavior_kind,
            )
            behavior_mapping_counter[mapping_key] += 1
            cleaned_behavior_counter[cleaned_key] += 1
            canonical_behavior_counter[canonical_key] += 1
            if behavior_source:
                cleaned_behavior_raw_values[cleaned_key].add(behavior_source)
                canonical_behavior_raw_variants[canonical_key].add(behavior_source)
            if behavior_clean:
                canonical_behavior_clean_variants[canonical_key].add(behavior_clean)

            row_payload = {
                "id_osservazione": row[index["ID Osservazione"]] if "ID Osservazione" in index else None,
                "id_comportamento_osservato": row[index["ID Comportamento osservato"]] if "ID Comportamento osservato" in index else None,
                "data": to_iso_date(raw_date),
                "event_month": event_month.strftime("%Y-%m-%d"),
                "area_canonica": area_canonica,
                "gruppo_raw": raw_group_text,
                "gruppo_clean": group_clean,
                "area_osservazione": normalize_whitespace(row[index["Area osservazione"]]) if "Area osservazione" in index else None,
                "attivita_raw": raw_activity_text,
                "attivita_clean": activity_clean,
                "attivita_famiglia": activity_family,
                "comportamento": raw_behavior_text,
                "comportamento_visualizzato": displayed_behavior_text,
                "comportamento_clean": behavior_clean,
                "comportamento_canonico": behavior_taxonomy.get("comportamento_canonico"),
                "canonical_behavior_id": behavior_taxonomy.get("canonical_behavior_id"),
                "comportamento_famiglia_originale": behavior_taxonomy.get("comportamento_famiglia_originale"),
                "comportamento_famiglia": behavior_family,
                "comportamento_famiglia_v2": behavior_taxonomy.get("comportamento_famiglia_v2"),
                "comportamento_sottofamiglia_v2": behavior_taxonomy.get("comportamento_sottofamiglia_v2"),
                "meccanismo_rischio_v2": behavior_taxonomy.get("meccanismo_rischio_v2"),
                "presidio_v2": behavior_taxonomy.get("presidio_v2"),
                "severity_rank_v2": behavior_taxonomy.get("severity_rank_v2"),
                "taxonomy_source": behavior_taxonomy.get("taxonomy_source"),
                "taxonomy_confidence": behavior_taxonomy.get("taxonomy_confidence"),
                "canonical_cluster_size": behavior_taxonomy.get("canonical_cluster_size"),
                "canonical_row_count_total": behavior_taxonomy.get("canonical_row_count_total"),
                "old_family_conflict": behavior_taxonomy.get("old_family_conflict"),
                "old_families_cluster": behavior_taxonomy.get("old_families_cluster"),
                "merge_source": behavior_taxonomy.get("merge_source"),
                "merge_confidence": behavior_taxonomy.get("merge_confidence"),
                "merge_note": behavior_taxonomy.get("merge_note"),
                "review_flag": behavior_taxonomy.get("review_flag"),
                "comportamento_tipo": behavior_kind,
                "comportamento_osservabile": int(behavior_kind == BEHAVIOR_KIND_BEHAVIOR),
                "comportamenti_sicuri": safe_value,
                "comportamenti_non_sicuri": unsafe_value,
                "unsafe_flag": int(unsafe_value > 0),
            }
            writer.writerow(row_payload)

            bucket_key = (area_canonica, event_month)
            if bucket_key not in monthly_stats:
                monthly_stats[bucket_key] = {
                    "area_canonica": area_canonica,
                    "event_month": event_month,
                    "obs_rows": 0.0,
                    "obs_safe_total": 0.0,
                    "obs_unsafe_total": 0.0,
                    "obs_rows_with_unsafe": 0.0,
                }
            bucket = monthly_stats[bucket_key]
            bucket["obs_rows"] += 1.0
            bucket["obs_safe_total"] += safe_value
            bucket["obs_unsafe_total"] += unsafe_value
            bucket["obs_rows_with_unsafe"] += float(unsafe_value > 0)

            family_slug = snake_case(activity_family)
            bucket[f"obs_family_{family_slug}_rows"] = bucket.get(f"obs_family_{family_slug}_rows", 0.0) + 1.0
            bucket[f"obs_family_{family_slug}_safe_total"] = bucket.get(f"obs_family_{family_slug}_safe_total", 0.0) + safe_value
            bucket[f"obs_family_{family_slug}_unsafe_total"] = bucket.get(f"obs_family_{family_slug}_unsafe_total", 0.0) + unsafe_value

            observation_id = row_payload["id_osservazione"]
            if observation_id is not None:
                unique_observations[bucket_key].add(observation_id)

    rows_for_frame = []
    for key, bucket in monthly_stats.items():
        area_canonica, event_month = key
        row_payload = dict(bucket)
        row_payload["obs_unique_observations"] = float(len(unique_observations[key]))
        total_behaviors = row_payload["obs_safe_total"] + row_payload["obs_unsafe_total"]
        row_payload["obs_unsafe_rate"] = safe_div(row_payload["obs_unsafe_total"], total_behaviors)
        row_payload["obs_unsafe_row_rate"] = safe_div(row_payload["obs_rows_with_unsafe"], row_payload["obs_rows"])
        rows_for_frame.append(row_payload)

    monthly_df = pd.DataFrame(rows_for_frame).sort_values(["area_canonica", "event_month"]).reset_index(drop=True)
    for column in list(monthly_df.columns):
        if column.startswith("obs_family_") and column.endswith("_rows"):
            prefix = column[:-5]
            safe_column = f"{prefix}_safe_total"
            unsafe_column = f"{prefix}_unsafe_total"
            rate_column = f"{prefix}_unsafe_rate"
            if safe_column in monthly_df.columns and unsafe_column in monthly_df.columns:
                total = monthly_df[safe_column].fillna(0) + monthly_df[unsafe_column].fillna(0)
                monthly_df[rate_column] = np.where(total > 0, monthly_df[unsafe_column] / total, 0.0)

    group_mapping_df = pd.DataFrame(
        [
            {"gruppo_raw": raw_group, "area_canonica": area, "row_count": count}
            for (raw_group, area), count in group_mapping_counter.items()
        ]
    ).sort_values(["area_canonica", "row_count", "gruppo_raw"], ascending=[True, False, True])
    activity_mapping_df = pd.DataFrame(
        [
            {
                "attivita_raw": raw_activity,
                "attivita_clean": clean_label,
                "attivita_famiglia": family,
                "row_count": count,
            }
            for (raw_activity, clean_label, family), count in activity_mapping_counter.items()
        ]
    ).sort_values(["attivita_famiglia", "row_count", "attivita_raw"], ascending=[True, False, True])

    activity_catalog_df = pd.DataFrame(
        [
            {
                "attivita_raw": raw_activity,
                "attivita_clean": clean_label,
                "attivita_famiglia": family,
                "row_count_total": total_count,
                "row_count_kept": activity_mapping_counter.get((raw_activity, clean_label, family), 0),
                "row_count_covid_excluded": activity_covid_counter.get((raw_activity, clean_label, family), 0),
                "is_only_covid_excluded": int(
                    activity_mapping_counter.get((raw_activity, clean_label, family), 0) == 0
                    and activity_covid_counter.get((raw_activity, clean_label, family), 0) > 0
                ),
            }
            for (raw_activity, clean_label, family), total_count in activity_total_counter.items()
        ]
    ).sort_values(
        ["is_only_covid_excluded", "attivita_famiglia", "row_count_total", "attivita_raw"],
        ascending=[False, True, False, True],
    )

    cleaned_activity_summary_df = pd.DataFrame(
        [
            {
                "attivita_clean": clean_label,
                "attivita_famiglia": family,
                "row_count_kept": count,
                "raw_variants_count": len(cleaned_activity_raw_values[(clean_label, family)]),
                "raw_variants": " | ".join(sorted(cleaned_activity_raw_values[(clean_label, family)])),
            }
            for (clean_label, family), count in cleaned_activity_counter.items()
        ]
    ).sort_values(["attivita_famiglia", "row_count_kept", "attivita_clean"], ascending=[True, False, True])

    behavior_mapping_df = pd.DataFrame(
        [
            {
                "comportamento_raw": raw_behavior,
                "comportamento_clean": clean_label,
                "comportamento_canonico": canonical_label,
                "canonical_behavior_id": canonical_behavior_id,
                "comportamento_famiglia_originale": original_family,
                "comportamento_famiglia": family,
                "comportamento_famiglia_v2": family_v2,
                "comportamento_sottofamiglia_v2": subfamily_v2,
                "meccanismo_rischio_v2": risk_mechanism_v2,
                "presidio_v2": control_v2,
                "severity_rank_v2": severity_rank_v2,
                "taxonomy_source": taxonomy_source,
                "taxonomy_confidence": taxonomy_confidence,
                "canonical_cluster_size": canonical_cluster_size,
                "canonical_row_count_total": canonical_row_count_total,
                "old_family_conflict": old_family_conflict,
                "old_families_cluster": old_families_cluster,
                "merge_source": merge_source,
                "merge_confidence": merge_confidence,
                "merge_note": merge_note,
                "review_flag": review_flag,
                "comportamento_tipo": behavior_kind,
                "row_count": count,
            }
            for (
                raw_behavior,
                clean_label,
                canonical_label,
                canonical_behavior_id,
                original_family,
                family,
                family_v2,
                subfamily_v2,
                risk_mechanism_v2,
                control_v2,
                severity_rank_v2,
                taxonomy_source,
                taxonomy_confidence,
                canonical_cluster_size,
                canonical_row_count_total,
                old_family_conflict,
                old_families_cluster,
                merge_source,
                merge_confidence,
                merge_note,
                review_flag,
                behavior_kind,
            ), count in behavior_mapping_counter.items()
        ]
    ).sort_values(
        ["comportamento_tipo", "comportamento_famiglia", "row_count", "comportamento_canonico", "comportamento_raw"],
        ascending=[True, True, False, True, True],
    )

    cleaned_behavior_summary_df = pd.DataFrame(
        [
            {
                "comportamento_clean": clean_label,
                "comportamento_canonico": canonical_label,
                "canonical_behavior_id": canonical_behavior_id,
                "comportamento_famiglia_originale": original_family,
                "comportamento_famiglia": family,
                "comportamento_famiglia_v2": family_v2,
                "comportamento_sottofamiglia_v2": subfamily_v2,
                "meccanismo_rischio_v2": risk_mechanism_v2,
                "presidio_v2": control_v2,
                "severity_rank_v2": severity_rank_v2,
                "taxonomy_source": taxonomy_source,
                "taxonomy_confidence": taxonomy_confidence,
                "canonical_cluster_size": canonical_cluster_size,
                "canonical_row_count_total": canonical_row_count_total,
                "old_family_conflict": old_family_conflict,
                "old_families_cluster": old_families_cluster,
                "merge_source": merge_source,
                "merge_confidence": merge_confidence,
                "merge_note": merge_note,
                "review_flag": review_flag,
                "comportamento_tipo": behavior_kind,
                "row_count_kept": count,
                "raw_variants_count": len(cleaned_behavior_raw_values[cleaned_key]),
                "raw_variants": " | ".join(sorted(cleaned_behavior_raw_values[cleaned_key])),
            }
            for cleaned_key, count in cleaned_behavior_counter.items()
            for (
                clean_label,
                canonical_label,
                canonical_behavior_id,
                original_family,
                family,
                family_v2,
                subfamily_v2,
                risk_mechanism_v2,
                control_v2,
                severity_rank_v2,
                taxonomy_source,
                taxonomy_confidence,
                canonical_cluster_size,
                canonical_row_count_total,
                old_family_conflict,
                old_families_cluster,
                merge_source,
                merge_confidence,
                merge_note,
                review_flag,
                behavior_kind,
            ) in [cleaned_key]
        ]
    ).sort_values(
        ["comportamento_tipo", "comportamento_famiglia", "row_count_kept", "comportamento_canonico", "comportamento_clean"],
        ascending=[True, True, False, True, True],
    )

    behavior_canonical_summary_df = pd.DataFrame(
        [
            {
                "comportamento_canonico": canonical_label,
                "canonical_behavior_id": canonical_behavior_id,
                "comportamento_famiglia": family,
                "comportamento_sottofamiglia_v2": subfamily_v2,
                "meccanismo_rischio_v2": risk_mechanism_v2,
                "presidio_v2": control_v2,
                "severity_rank_v2": severity_rank_v2,
                "taxonomy_source": taxonomy_source,
                "taxonomy_confidence": taxonomy_confidence,
                "merge_source": merge_source,
                "merge_confidence": merge_confidence,
                "review_flag": review_flag,
                "comportamento_tipo": behavior_kind,
                "row_count_kept": count,
                "clean_variants_count": len(canonical_behavior_clean_variants[canonical_key]),
                "clean_variants": " | ".join(sorted(canonical_behavior_clean_variants[canonical_key])),
                "raw_variants_count": len(canonical_behavior_raw_variants[canonical_key]),
                "raw_variants": " | ".join(sorted(canonical_behavior_raw_variants[canonical_key])),
            }
            for canonical_key, count in canonical_behavior_counter.items()
            for (
                canonical_label,
                canonical_behavior_id,
                family,
                subfamily_v2,
                risk_mechanism_v2,
                control_v2,
                severity_rank_v2,
                taxonomy_source,
                taxonomy_confidence,
                merge_source,
                merge_confidence,
                review_flag,
                behavior_kind,
            ) in [canonical_key]
        ]
    ).sort_values(
        ["comportamento_tipo", "comportamento_famiglia", "row_count_kept", "comportamento_canonico"],
        ascending=[True, True, False, True],
    )

    profile = {
        "input_rows": input_rows,
        "rows_removed_covid": covid_rows_removed,
        "rows_after_cleaning": int(monthly_df["obs_rows"].sum()) if not monthly_df.empty else 0,
        "unique_groups": int(group_mapping_df["gruppo_raw"].nunique()) if not group_mapping_df.empty else 0,
        "unique_activities": int(activity_mapping_df["attivita_raw"].nunique()) if not activity_mapping_df.empty else 0,
        "unique_activities_total_raw": int(activity_catalog_df["attivita_raw"].nunique()) if not activity_catalog_df.empty else 0,
        "unique_activities_cleaned": int(cleaned_activity_summary_df["attivita_clean"].nunique()) if not cleaned_activity_summary_df.empty else 0,
        "unique_behaviors": int(behavior_mapping_df["comportamento_raw"].nunique()) if not behavior_mapping_df.empty else 0,
        "unique_behaviors_cleaned": int(cleaned_behavior_summary_df["comportamento_clean"].nunique()) if not cleaned_behavior_summary_df.empty else 0,
        "unique_behaviors_canonical": int(behavior_canonical_summary_df["canonical_behavior_id"].nunique()) if not behavior_canonical_summary_df.empty else 0,
        "observable_behavior_rows": int(cleaned_behavior_summary_df.loc[cleaned_behavior_summary_df["comportamento_tipo"].eq(BEHAVIOR_KIND_BEHAVIOR), "row_count_kept"].sum()) if not cleaned_behavior_summary_df.empty else 0,
    }
    return (
        monthly_df,
        profile,
        group_mapping_df,
        activity_mapping_df,
        activity_catalog_df,
        cleaned_activity_summary_df,
        behavior_mapping_df,
        cleaned_behavior_summary_df,
        behavior_canonical_summary_df,
    )


def add_temporal_features(panel_df: pd.DataFrame) -> pd.DataFrame:
    panel = panel_df.sort_values(["area_canonica", "event_month"]).copy()
    feature_columns = [
        column
        for column in panel.columns
        if column.startswith("obs_") or column.startswith("evt_")
    ]
    group = panel.groupby("area_canonica", dropna=False)
    derived_columns: dict[str, pd.Series] = {}
    for column in feature_columns:
        derived_columns[f"{column}_lag1"] = group[column].shift(1)
        derived_columns[f"{column}_roll3_mean"] = (
            group[column]
            .transform(lambda series: series.shift(1).rolling(3, min_periods=1).mean())
        )

    injury_flag = panel["evt_injury_like_count"].gt(0).astype(float)
    derived_columns["months_since_last_injury"] = group["evt_injury_like_count"].transform(months_since_last_positive).fillna(999.0)
    derived_columns["obs_to_event_ratio"] = np.where(
        panel["obs_unique_observations"] > 0,
        panel["evt_total_rows"] / panel["obs_unique_observations"],
        0.0,
    )
    derived_columns["injury_flag_current_month"] = injury_flag
    panel = pd.concat([panel, pd.DataFrame(derived_columns, index=panel.index)], axis=1)
    return panel


def build_model_dataset(obs_monthly: pd.DataFrame, evt_monthly: pd.DataFrame) -> pd.DataFrame:
    obs_monthly = obs_monthly.copy()
    evt_monthly = evt_monthly.copy()
    all_areas = sorted(set(obs_monthly["area_canonica"].dropna()) | set(evt_monthly["area_canonica"].dropna()))
    start_month = min(obs_monthly["event_month"].min(), evt_monthly["event_month"].min())
    end_month = max(obs_monthly["event_month"].max(), evt_monthly["event_month"].max())
    all_months = pd.date_range(start=start_month, end=end_month, freq="MS")

    grid = pd.MultiIndex.from_product([all_areas, all_months], names=["area_canonica", "event_month"]).to_frame(index=False)
    panel = grid.merge(obs_monthly, on=["area_canonica", "event_month"], how="left")
    panel = panel.merge(evt_monthly, on=["area_canonica", "event_month"], how="left")

    numeric_columns = [column for column in panel.columns if column not in {"area_canonica", "event_month"}]
    for column in numeric_columns:
        panel[column] = panel[column].fillna(0)

    panel = add_temporal_features(panel)
    max_event_month = evt_monthly["event_month"].max()
    next_injury = panel.groupby("area_canonica", dropna=False)["evt_injury_like_count"].shift(-1)
    next_incident = panel.groupby("area_canonica", dropna=False)["evt_incident_without_injury_count"].shift(-1)
    next_total_events = panel.groupby("area_canonica", dropna=False)["evt_total_rows"].shift(-1)
    target_month = panel["event_month"] + pd.offsets.MonthBegin(1)
    target_available = target_month <= max_event_month
    target_frame = pd.DataFrame(
        {
            "target_month": target_month,
            "target_available": target_available,
            "target_injury_next_month": np.where(
                target_available,
                next_injury.fillna(0).gt(0).astype(int),
                np.nan,
            ),
            "target_adverse_next_month": np.where(
                target_available,
                (next_injury.fillna(0) + next_incident.fillna(0)).gt(0).astype(int),
                np.nan,
            ),
            "target_event_count_next_month": np.where(
                target_available,
                next_total_events.fillna(0),
                np.nan,
            ),
            "has_observation_coverage": panel["obs_rows"].gt(0).astype(int),
        },
        index=panel.index,
    )
    panel = pd.concat([panel, target_frame], axis=1)

    panel = panel.loc[panel["event_month"] >= obs_monthly["event_month"].min()].copy()
    panel = panel.sort_values(["event_month", "area_canonica"]).reset_index(drop=True)
    return panel


def build_behavior_family_month_long(observations_df: pd.DataFrame) -> pd.DataFrame:
    return build_behavior_category_month_long(
        observations_df,
        category_column="comportamento_famiglia",
        category_output_column="comportamento_famiglia",
    )


def build_behavior_category_month_long(
    observations_df: pd.DataFrame,
    category_column: str,
    category_output_column: str,
) -> pd.DataFrame:
    output_columns = [
        "area_canonica",
        "event_month",
        category_output_column,
        "unsafe_total",
        "safe_total",
        "total_behaviors",
        "unsafe_rate",
    ]
    if category_column not in observations_df.columns:
        return pd.DataFrame(columns=output_columns)

    observable_df = observations_df.loc[observations_df["comportamento_tipo"].eq(BEHAVIOR_KIND_BEHAVIOR)].copy()
    if observable_df.empty:
        return pd.DataFrame(columns=output_columns)

    observable_df[category_output_column] = observable_df[category_column].map(normalize_whitespace)
    observable_df = observable_df.loc[observable_df[category_output_column].notna()].copy()
    if observable_df.empty:
        return pd.DataFrame(columns=output_columns)

    grouped = (
        observable_df.groupby(["area_canonica", "event_month", category_output_column], dropna=False)
        .agg(
            unsafe_total=("comportamenti_non_sicuri", "sum"),
            safe_total=("comportamenti_sicuri", "sum"),
        )
        .reset_index()
    )
    grouped["total_behaviors"] = grouped["safe_total"] + grouped["unsafe_total"]
    grouped["unsafe_rate"] = np.where(
        grouped["total_behaviors"].gt(0),
        grouped["unsafe_total"] / grouped["total_behaviors"],
        0.0,
    )
    return grouped


def build_behavior_area_month_features(observations_df: pd.DataFrame) -> pd.DataFrame:
    return build_behavior_category_area_month_features(
        observations_df,
        category_column="comportamento_famiglia",
        category_output_column="comportamento_famiglia",
        feature_prefix="obs_behavior_family",
    )


def build_behavior_subfamily_area_month_features(observations_df: pd.DataFrame) -> pd.DataFrame:
    return build_behavior_category_area_month_features(
        observations_df,
        category_column="comportamento_sottofamiglia_v2",
        category_output_column="comportamento_sottofamiglia_v2",
        feature_prefix="obs_behavior_subfamily",
    )


def build_behavior_category_area_month_features(
    observations_df: pd.DataFrame,
    category_column: str,
    category_output_column: str,
    feature_prefix: str,
) -> pd.DataFrame:
    grouped = build_behavior_category_month_long(
        observations_df,
        category_column=category_column,
        category_output_column=category_output_column,
    )
    if grouped.empty:
        return pd.DataFrame(columns=["area_canonica", "event_month"])

    wide_frames = []
    for metric in ["unsafe_total", "unsafe_rate"]:
        pivoted = grouped.pivot_table(
            index=["area_canonica", "event_month"],
            columns=category_output_column,
            values=metric,
            fill_value=0,
        )
        pivoted.columns = [
            f"{feature_prefix}_{snake_case(str(column))}_{metric}"
            for column in pivoted.columns
        ]
        wide_frames.append(pivoted)
    return pd.concat(wide_frames, axis=1).reset_index()


def build_default_behavior_expert_risk_rank_frame(
    behavior_families: Iterable[str],
    observations_df: pd.DataFrame | None = None,
) -> pd.DataFrame:
    severity_defaults: dict[str, tuple[int, str]] = {}
    if observations_df is not None and {"comportamento_famiglia", "severity_rank_v2", "comportamento_tipo"}.issubset(observations_df.columns):
        severity_frame = observations_df.loc[
            observations_df["comportamento_tipo"].eq(BEHAVIOR_KIND_BEHAVIOR),
            ["comportamento_famiglia", "severity_rank_v2"],
        ].copy()
        severity_frame["comportamento_famiglia"] = severity_frame["comportamento_famiglia"].map(normalize_whitespace)
        severity_frame["severity_rank_v2"] = pd.to_numeric(severity_frame["severity_rank_v2"], errors="coerce").clip(lower=1, upper=5)
        severity_frame = severity_frame.loc[
            severity_frame["comportamento_famiglia"].notna() & severity_frame["severity_rank_v2"].notna()
        ]
        if not severity_frame.empty:
            family_severity = (
                severity_frame.groupby("comportamento_famiglia", dropna=False)["severity_rank_v2"]
                .median()
                .round()
                .clip(lower=1, upper=5)
                .astype(int)
            )
            severity_defaults = {
                family: (
                    int(rank),
                    f"Default derivato dalla mediana del severity rank v2 osservato nella famiglia (`{int(rank)}`).",
                )
                for family, rank in family_severity.items()
            }

    rows = []
    for family in sorted({normalize_whitespace(value) for value in behavior_families if normalize_whitespace(value)}):
        default_rank, rationale = severity_defaults.get(
            family,
            DEFAULT_BEHAVIOR_EXPERT_RISK_RANKS.get(
                family,
                (3, "Bozza iniziale neutra, da rivedere con HSE."),
            ),
        )
        rows.append(
            {
                "comportamento_famiglia": family,
                "risk_rank": int(default_rank),
                "risk_label": EXPERT_RISK_LABELS[int(default_rank)],
                "rationale": rationale,
            }
        )
    return pd.DataFrame(rows)


def load_behavior_expert_risk_rank(
    path: Path,
    observations_df: pd.DataFrame,
) -> pd.DataFrame:
    observed_families = sorted(
        observations_df.loc[
            observations_df["comportamento_tipo"].eq(BEHAVIOR_KIND_BEHAVIOR),
            "comportamento_famiglia",
        ].dropna().unique().tolist()
    )
    default_df = build_default_behavior_expert_risk_rank_frame(observed_families, observations_df)
    default_by_family = default_df.set_index("comportamento_famiglia")

    if path.exists():
        manual_df = pd.read_csv(path)
    else:
        manual_df = pd.DataFrame(columns=["comportamento_famiglia", "risk_rank", "risk_label", "rationale"])

    if "comportamento_famiglia" not in manual_df.columns:
        manual_df["comportamento_famiglia"] = None
    manual_df["comportamento_famiglia"] = manual_df["comportamento_famiglia"].map(normalize_whitespace)
    manual_df = manual_df.loc[manual_df["comportamento_famiglia"].notna()].drop_duplicates("comportamento_famiglia", keep="first")

    if "risk_rank" not in manual_df.columns:
        manual_df["risk_rank"] = np.nan
    manual_df["risk_rank"] = pd.to_numeric(manual_df["risk_rank"], errors="coerce").clip(lower=1, upper=5)

    if "risk_label" not in manual_df.columns:
        manual_df["risk_label"] = None
    if "rationale" not in manual_df.columns:
        manual_df["rationale"] = None
    manual_df = manual_df.set_index("comportamento_famiglia")

    resolved_df = default_by_family.copy()
    for column in ["risk_rank", "risk_label", "rationale"]:
        if column in manual_df.columns:
            resolved_df[column] = manual_df[column].combine_first(resolved_df[column])

    resolved_df["risk_rank"] = pd.to_numeric(resolved_df["risk_rank"], errors="coerce").fillna(3).clip(lower=1, upper=5).astype(int)
    resolved_df["risk_label"] = resolved_df["risk_rank"].map(EXPERT_RISK_LABELS)
    resolved_df["rationale"] = resolved_df["rationale"].fillna("Bozza iniziale neutra, da rivedere con HSE.")
    resolved_df = (
        resolved_df.reset_index()
        .sort_values(["risk_rank", "comportamento_famiglia"], ascending=[True, True])
        .reset_index(drop=True)
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    resolved_df.to_csv(path, index=False)
    return resolved_df


def add_temporal_features_for_columns(panel_df: pd.DataFrame, feature_columns: list[str]) -> pd.DataFrame:
    if not feature_columns:
        return panel_df
    panel = panel_df.sort_values(["area_canonica", "event_month"]).copy()
    grouped = panel.groupby("area_canonica", dropna=False)
    derived_columns: dict[str, pd.Series] = {}
    for column in feature_columns:
        derived_columns[f"{column}_lag1"] = grouped[column].shift(1)
        derived_columns[f"{column}_roll3_mean"] = grouped[column].transform(
            lambda series: series.shift(1).rolling(3, min_periods=1).mean()
        )
    return pd.concat([panel, pd.DataFrame(derived_columns, index=panel.index)], axis=1)


def compute_historical_feature_weights(
    panel_df: pd.DataFrame,
    feature_columns: list[str],
    target_column: str = "target_adverse_next_month",
    min_exposure_rows: int = 8,
    shrinkage_rows: int = 12,
    min_weight: float = 0.7,
    max_weight: float = 2.0,
) -> pd.DataFrame:
    if not feature_columns:
        return pd.DataFrame(
            columns=[
                "event_month",
                "feature",
                "exposure_area_months",
                "prior_base_rate",
                "prior_feature_rate",
                "raw_lift",
                "risk_weight",
            ]
        )

    modeling_df = panel_df.loc[
        panel_df["has_observation_coverage"].eq(1) & panel_df["target_available"].eq(True)
    ].sort_values("event_month").copy()
    if modeling_df.empty:
        return pd.DataFrame(
            columns=[
                "event_month",
                "feature",
                "exposure_area_months",
                "prior_base_rate",
                "prior_feature_rate",
                "raw_lift",
                "risk_weight",
            ]
        )

    default_rate = float(modeling_df[target_column].mean())
    unique_months = sorted(pd.to_datetime(modeling_df["event_month"]).drop_duplicates())
    prior_df = modeling_df.iloc[0:0].copy()
    weight_rows: list[dict[str, Any]] = []

    for month in unique_months:
        base_rate = float(prior_df[target_column].mean()) if not prior_df.empty else default_rate
        base_rate = max(base_rate, 1e-6)
        for column in feature_columns:
            if prior_df.empty:
                exposure_rows = 0
                feature_rate = base_rate
            else:
                exposure_mask = prior_df[column].fillna(0).gt(0)
                exposure_rows = int(exposure_mask.sum())
                feature_rate = (
                    float(prior_df.loc[exposure_mask, target_column].mean())
                    if exposure_rows >= min_exposure_rows
                    else base_rate
                )
            reliability = exposure_rows / (exposure_rows + shrinkage_rows) if exposure_rows > 0 else 0.0
            raw_lift = feature_rate / base_rate if base_rate > 0 else 1.0
            risk_weight = float(np.clip(1.0 + reliability * (raw_lift - 1.0), min_weight, max_weight))
            weight_rows.append(
                {
                    "event_month": pd.Timestamp(month),
                    "feature": column,
                    "exposure_area_months": exposure_rows,
                    "prior_base_rate": float(base_rate),
                    "prior_feature_rate": float(feature_rate),
                    "raw_lift": float(raw_lift),
                    "risk_weight": risk_weight,
                }
            )
        prior_df = pd.concat(
            [prior_df, modeling_df.loc[pd.to_datetime(modeling_df["event_month"]).eq(month)]],
            axis=0,
            ignore_index=True,
        )

    return pd.DataFrame(weight_rows)


def add_weighted_feature_scores(
    panel_df: pd.DataFrame,
    feature_columns: list[str],
    weights_df: pd.DataFrame,
    score_prefix: str,
) -> pd.DataFrame:
    panel = panel_df.copy()
    if not feature_columns:
        panel[f"{score_prefix}_weighted_score"] = 0.0
        panel[f"{score_prefix}_weighted_presence"] = 0.0
        return add_temporal_features_for_columns(
            panel,
            [f"{score_prefix}_weighted_score", f"{score_prefix}_weighted_presence"],
        )

    weight_matrix_df = (
        weights_df.pivot(index="event_month", columns="feature", values="risk_weight")
        .rename(columns=lambda column: f"{column}_risk_weight")
        .reset_index()
    )
    panel = panel.merge(weight_matrix_df, on="event_month", how="left")

    weight_columns = [f"{column}_risk_weight" for column in feature_columns]
    for column in weight_columns:
        if column not in panel.columns:
            panel[column] = 1.0

    exposure_matrix = panel[feature_columns].fillna(0.0).to_numpy(dtype=float)
    weight_matrix = panel[weight_columns].fillna(1.0).to_numpy(dtype=float)
    panel[f"{score_prefix}_weighted_score"] = (exposure_matrix * weight_matrix).sum(axis=1)
    panel[f"{score_prefix}_weighted_presence"] = ((exposure_matrix > 0).astype(float) * weight_matrix).sum(axis=1)
    panel = panel.drop(columns=weight_columns)
    return add_temporal_features_for_columns(
        panel,
        [f"{score_prefix}_weighted_score", f"{score_prefix}_weighted_presence"],
    )


def build_behavior_weighted_panel(
    panel_df: pd.DataFrame,
    observations_df: pd.DataFrame,
    output_dir: Path,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    behavior_month_df = build_behavior_area_month_features(observations_df)
    if behavior_month_df.empty:
        return panel_df.copy(), behavior_month_df, pd.DataFrame()

    behavior_feature_columns = [
        column
        for column in behavior_month_df.columns
        if column.startswith("obs_behavior_family_")
    ]
    behavior_unsafe_columns = [
        column
        for column in behavior_feature_columns
        if column.endswith("_unsafe_total")
    ]

    merged_panel = panel_df.merge(behavior_month_df, on=["area_canonica", "event_month"], how="left")
    for column in behavior_feature_columns:
        merged_panel[column] = merged_panel[column].fillna(0.0)
    merged_panel = add_temporal_features_for_columns(merged_panel, behavior_feature_columns)

    weights_df = compute_historical_feature_weights(merged_panel, behavior_unsafe_columns)
    weighted_panel = add_weighted_feature_scores(
        merged_panel,
        behavior_unsafe_columns,
        weights_df,
        score_prefix="obs_behavior",
    )

    behavior_month_df.to_csv(output_dir / "behavior_area_month.csv", index=False)
    weights_df.to_csv(output_dir / "behavior_family_risk_weights_by_month.csv", index=False)
    weighted_panel.to_csv(output_dir / "risk_panel_area_month_behavior_weighted.csv", index=False)
    return weighted_panel, behavior_month_df, weights_df


def build_behavior_subfamily_weighted_panel(
    panel_df: pd.DataFrame,
    observations_df: pd.DataFrame,
    output_dir: Path,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    behavior_subfamily_month_df = build_behavior_subfamily_area_month_features(observations_df)
    if behavior_subfamily_month_df.empty:
        return panel_df.copy(), behavior_subfamily_month_df, pd.DataFrame()

    behavior_subfamily_feature_columns = [
        column
        for column in behavior_subfamily_month_df.columns
        if column.startswith("obs_behavior_subfamily_")
    ]
    behavior_subfamily_unsafe_columns = [
        column
        for column in behavior_subfamily_feature_columns
        if column.endswith("_unsafe_total")
    ]

    merged_panel = panel_df.merge(behavior_subfamily_month_df, on=["area_canonica", "event_month"], how="left")
    for column in behavior_subfamily_feature_columns:
        merged_panel[column] = merged_panel[column].fillna(0.0)
    merged_panel = add_temporal_features_for_columns(merged_panel, behavior_subfamily_feature_columns)

    weights_df = compute_historical_feature_weights(merged_panel, behavior_subfamily_unsafe_columns)
    weighted_panel = add_weighted_feature_scores(
        merged_panel,
        behavior_subfamily_unsafe_columns,
        weights_df,
        score_prefix="obs_behavior_subfamily",
    )

    behavior_subfamily_month_df.to_csv(output_dir / "behavior_subfamily_area_month.csv", index=False)
    weights_df.to_csv(output_dir / "behavior_subfamily_risk_weights_by_month.csv", index=False)
    weighted_panel.to_csv(output_dir / "risk_panel_area_month_behavior_subfamily_weighted.csv", index=False)
    return weighted_panel, behavior_subfamily_month_df, weights_df


def build_behavior_expert_weighted_panel(
    panel_df: pd.DataFrame,
    observations_df: pd.DataFrame,
    risk_rank_df: pd.DataFrame,
    output_dir: Path,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    behavior_long_df = build_behavior_family_month_long(observations_df)
    if behavior_long_df.empty:
        return panel_df.copy(), pd.DataFrame()

    expert_long_df = behavior_long_df.merge(
        risk_rank_df[["comportamento_famiglia", "risk_rank", "risk_label", "rationale"]],
        on="comportamento_famiglia",
        how="left",
    )
    expert_long_df["risk_rank"] = pd.to_numeric(expert_long_df["risk_rank"], errors="coerce").fillna(3).clip(lower=1, upper=5).astype(int)
    expert_long_df["severity_weight"] = 6 - expert_long_df["risk_rank"]
    expert_long_df["unsafe_severity_score"] = expert_long_df["unsafe_total"] * expert_long_df["severity_weight"]
    expert_long_df["unsafe_presence_severity_score"] = expert_long_df["severity_weight"] * expert_long_df["unsafe_total"].gt(0).astype(float)
    expert_long_df["high_risk_unsafe_total"] = np.where(expert_long_df["risk_rank"].le(2), expert_long_df["unsafe_total"], 0.0)
    expert_long_df["critical_risk_unsafe_total"] = np.where(expert_long_df["risk_rank"].eq(1), expert_long_df["unsafe_total"], 0.0)
    expert_long_df["high_risk_family_presence"] = np.where(
        expert_long_df["risk_rank"].le(2) & expert_long_df["unsafe_total"].gt(0),
        1.0,
        0.0,
    )
    expert_long_df["critical_risk_family_presence"] = np.where(
        expert_long_df["risk_rank"].eq(1) & expert_long_df["unsafe_total"].gt(0),
        1.0,
        0.0,
    )

    expert_month_df = (
        expert_long_df.groupby(["area_canonica", "event_month"], dropna=False)
        .agg(
            obs_behavior_expert_unsafe_severity_score=("unsafe_severity_score", "sum"),
            obs_behavior_expert_presence_severity_score=("unsafe_presence_severity_score", "sum"),
            obs_behavior_expert_high_risk_unsafe_total=("high_risk_unsafe_total", "sum"),
            obs_behavior_expert_critical_risk_unsafe_total=("critical_risk_unsafe_total", "sum"),
            obs_behavior_expert_high_risk_family_presence=("high_risk_family_presence", "sum"),
            obs_behavior_expert_critical_risk_family_presence=("critical_risk_family_presence", "sum"),
            obs_behavior_expert_total_unsafe_total=("unsafe_total", "sum"),
        )
        .reset_index()
    )
    expert_month_df["obs_behavior_expert_avg_unsafe_severity"] = np.where(
        expert_month_df["obs_behavior_expert_total_unsafe_total"].gt(0),
        expert_month_df["obs_behavior_expert_unsafe_severity_score"] / expert_month_df["obs_behavior_expert_total_unsafe_total"],
        0.0,
    )

    rank_bucket_df = (
        expert_long_df.groupby(["area_canonica", "event_month", "risk_rank"], dropna=False)
        .agg(
            unsafe_total=("unsafe_total", "sum"),
            family_presence=("unsafe_total", lambda series: float(series.gt(0).sum())),
        )
        .reset_index()
    )
    rank_frames = []
    for metric in ["unsafe_total", "family_presence"]:
        pivoted = rank_bucket_df.pivot_table(
            index=["area_canonica", "event_month"],
            columns="risk_rank",
            values=metric,
            fill_value=0,
        )
        pivoted.columns = [f"obs_behavior_expert_rank_{int(column)}_{metric}" for column in pivoted.columns]
        rank_frames.append(pivoted)
    if rank_frames:
        expert_month_df = expert_month_df.merge(
            pd.concat(rank_frames, axis=1).reset_index(),
            on=["area_canonica", "event_month"],
            how="left",
        )

    expert_feature_columns = [
        column
        for column in expert_month_df.columns
        if column not in {"area_canonica", "event_month"}
    ]
    expert_panel_df = panel_df.merge(expert_month_df, on=["area_canonica", "event_month"], how="left")
    for column in expert_feature_columns:
        expert_panel_df[column] = expert_panel_df[column].fillna(0.0)
    expert_panel_df = add_temporal_features_for_columns(expert_panel_df, expert_feature_columns)

    expert_month_df.to_csv(output_dir / "behavior_expert_area_month.csv", index=False)
    expert_panel_df.to_csv(output_dir / "risk_panel_area_month_behavior_expert_weighted.csv", index=False)
    return expert_panel_df, expert_month_df


def pick_feature_columns(df: pd.DataFrame) -> list[str]:
    excluded = {
        "area_canonica",
        "event_month",
        "target_month",
        "target_available",
        "has_observation_coverage",
        "target_injury_next_month",
        "target_adverse_next_month",
        "target_event_count_next_month",
    }
    return [
        column
        for column in df.columns
        if column not in excluded and pd.api.types.is_numeric_dtype(df[column])
    ]


def artifact_name(filename: str, artifact_prefix: str = "") -> str:
    return f"{artifact_prefix}{filename}" if artifact_prefix else filename


def evaluate_classifier(model: Any, x_frame: pd.DataFrame, y_series: pd.Series) -> dict[str, float]:
    probabilities = model.predict_proba(x_frame)[:, 1]
    metrics = {
        "pr_auc": float(average_precision_score(y_series, probabilities)),
        "brier_score": float(brier_score_loss(y_series, probabilities)),
    }
    if y_series.nunique() > 1:
        metrics["roc_auc"] = float(roc_auc_score(y_series, probabilities))
    else:
        metrics["roc_auc"] = float("nan")
    return metrics


def feature_importance_frame(model: Any, feature_names: list[str]) -> pd.DataFrame:
    estimator = model.named_steps["model"] if hasattr(model, "named_steps") else model
    if hasattr(estimator, "coef_"):
        coefficients = np.abs(estimator.coef_)
        importances = coefficients.mean(axis=0) if coefficients.ndim > 1 else coefficients
    elif hasattr(estimator, "feature_importances_"):
        importances = estimator.feature_importances_
    else:
        return pd.DataFrame(columns=["feature", "importance"])
    frame = pd.DataFrame({"feature": feature_names, "importance": importances})
    return frame.sort_values("importance", ascending=False).reset_index(drop=True)


def build_baseline_candidate_models() -> dict[str, Pipeline]:
    return {
        "logistic_regression": Pipeline(
            steps=[
                ("imputer", SimpleImputer(strategy="median")),
                ("scaler", StandardScaler()),
                ("model", LogisticRegression(max_iter=3000, class_weight="balanced", random_state=42)),
            ]
        ),
        "random_forest": Pipeline(
            steps=[
                ("imputer", SimpleImputer(strategy="median")),
                ("model", RandomForestClassifier(
                    n_estimators=500,
                    min_samples_leaf=4,
                    class_weight="balanced_subsample",
                    random_state=42,
                )),
            ]
        ),
    }


def build_injury_detail_dataset(panel_df: pd.DataFrame, events_df: pd.DataFrame) -> pd.DataFrame:
    injury_events = events_df.loc[events_df["is_injury_like"].eq(1)].copy()
    injury_events = injury_events[
        [
            "id",
            "event_date",
            "event_month",
            "area_canonica",
            "problem_family",
            "problem_code",
            "tipologia_canonica",
            "tipologia_causa_clean",
            "tipologia_causa_macro",
            "parte_del_corpo_clean",
            "parte_del_corpo_macro",
            "agente_materiale_clean",
            "descrizione_clean",
            "giorni_persi_capped",
        ]
    ].copy()
    injury_events = injury_events.rename(
        columns={
            "event_month": "target_event_month",
            "event_date": "target_event_date",
            "id": "event_id",
        }
    )

    panel_features = panel_df.loc[panel_df["has_observation_coverage"].eq(1)].copy()
    panel_features = panel_features.rename(columns={"event_month": "source_month"})
    panel_features["target_event_month"] = panel_features["target_month"]

    merged = panel_features.merge(
        injury_events,
        on=["area_canonica", "target_event_month"],
        how="inner",
    )
    if merged.empty:
        return merged

    merged["target_event_month_num"] = pd.to_datetime(merged["target_event_month"]).dt.month
    merged["target_event_month_sin"] = np.sin(2 * np.pi * merged["target_event_month_num"] / 12.0)
    merged["target_event_month_cos"] = np.cos(2 * np.pi * merged["target_event_month_num"] / 12.0)

    area_dummies = pd.get_dummies(merged["area_canonica"], prefix="inj_area", dtype=float)
    merged = pd.concat([merged, area_dummies], axis=1)
    return merged.sort_values(["target_event_month", "area_canonica", "event_id"]).reset_index(drop=True)


def pick_injury_detail_feature_columns(df: pd.DataFrame) -> list[str]:
    excluded = {
        "event_id",
        "source_month",
        "target_month",
        "target_event_month",
        "target_event_date",
        "event_month",
        "area_canonica",
        "problem_family",
        "problem_code",
        "tipologia_canonica",
        "tipologia_causa_clean",
        "tipologia_causa_macro",
        "parte_del_corpo_clean",
        "parte_del_corpo_macro",
        "agente_materiale_clean",
        "descrizione_clean",
        "giorni_persi_capped",
    }
    return [
        column
        for column in df.columns
        if column not in excluded and pd.api.types.is_numeric_dtype(df[column])
    ]


def evaluate_multiclass_model(model: Any, x_frame: pd.DataFrame, y_series: pd.Series) -> dict[str, float]:
    predictions = model.predict(x_frame)
    probabilities = model.predict_proba(x_frame)
    classes = model.named_steps["model"].classes_ if hasattr(model, "named_steps") else model.classes_
    top_k = min(3, len(classes))
    return {
        "accuracy": float(accuracy_score(y_series, predictions)),
        "balanced_accuracy": float(balanced_accuracy_score(y_series, predictions)),
        "macro_f1": float(f1_score(y_series, predictions, average="macro")),
        f"top_{top_k}_accuracy": float(top_k_accuracy_score(y_series, probabilities, k=top_k, labels=classes)),
    }


def build_top_k_predictions_frame(
    base_df: pd.DataFrame,
    model: Any,
    x_frame: pd.DataFrame,
    target_column: str,
) -> pd.DataFrame:
    classes = np.array(model.named_steps["model"].classes_ if hasattr(model, "named_steps") else model.classes_)
    probabilities = model.predict_proba(x_frame)
    top_k = min(3, len(classes))
    top_indices = np.argsort(probabilities, axis=1)[:, ::-1][:, :top_k]

    output = base_df.copy()
    output[f"actual_{target_column}"] = base_df[target_column] if target_column in base_df.columns else None
    for rank in range(top_k):
        output[f"pred_{target_column}_top_{rank + 1}"] = classes[top_indices[:, rank]]
        output[f"pred_{target_column}_top_{rank + 1}_proba"] = probabilities[np.arange(len(base_df)), top_indices[:, rank]]
    return output


def collapse_rare_classes(series: pd.Series, min_count: int = 20) -> pd.Series:
    counts = series.value_counts(dropna=False)
    rare_labels = set(counts[counts < min_count].index.tolist())
    return series.map(lambda value: UNKNOWN_INJURY_DETAIL if value in rare_labels else value)


def train_injury_detail_models(
    detail_df: pd.DataFrame,
    panel_df: pd.DataFrame,
    output_dir: Path,
) -> dict[str, Any]:
    if detail_df.empty:
        return {"status": "skipped", "reason": "Nessun evento injury-like disponibile dopo il merge con le feature storiche."}

    unique_months = sorted(pd.to_datetime(detail_df["target_event_month"]).drop_duplicates())
    if len(unique_months) < 18:
        return {"status": "skipped", "reason": "Storico insufficiente per un secondo stadio temporale affidabile."}

    test_months = unique_months[-12:]
    validation_months = unique_months[-24:-12]
    train_months = unique_months[:-24]
    if len(train_months) < 6:
        train_months = unique_months[:-18]
        validation_months = unique_months[-18:-9]
        test_months = unique_months[-9:]

    feature_columns = pick_injury_detail_feature_columns(detail_df)
    candidate_models = {
        "logistic_regression": Pipeline(
            steps=[
                ("imputer", SimpleImputer(strategy="median")),
                ("scaler", StandardScaler()),
                ("model", LogisticRegression(max_iter=4000, class_weight="balanced", random_state=42)),
            ]
        ),
        "random_forest": Pipeline(
            steps=[
                ("imputer", SimpleImputer(strategy="median")),
                ("model", RandomForestClassifier(
                    n_estimators=500,
                    min_samples_leaf=3,
                    class_weight="balanced_subsample",
                    random_state=42,
                )),
            ]
        ),
    }

    report: dict[str, Any] = {
        "status": "trained",
        "n_rows": int(len(detail_df)),
        "n_features": int(len(feature_columns)),
        "train_month_range": [train_months[0].strftime("%Y-%m-%d"), train_months[-1].strftime("%Y-%m-%d")],
        "validation_month_range": [validation_months[0].strftime("%Y-%m-%d"), validation_months[-1].strftime("%Y-%m-%d")],
        "test_month_range": [test_months[0].strftime("%Y-%m-%d"), test_months[-1].strftime("%Y-%m-%d")],
        "targets": {},
    }

    latest_feature_month = panel_df.loc[
        panel_df["has_observation_coverage"].eq(1)
        & panel_df["event_month"].le(panel_df.loc[panel_df["evt_total_rows"].gt(0), "event_month"].max()),
        "event_month",
    ].max()
    latest_feature_df = panel_df.loc[
        panel_df["has_observation_coverage"].eq(1) & panel_df["event_month"].eq(latest_feature_month)
    ].copy()
    latest_feature_df["source_month"] = latest_feature_df["event_month"]
    latest_feature_df["target_event_month"] = latest_feature_df["target_month"]
    latest_feature_df["target_event_month_num"] = pd.to_datetime(latest_feature_df["target_event_month"]).dt.month
    latest_feature_df["target_event_month_sin"] = np.sin(2 * np.pi * latest_feature_df["target_event_month_num"] / 12.0)
    latest_feature_df["target_event_month_cos"] = np.cos(2 * np.pi * latest_feature_df["target_event_month_num"] / 12.0)
    latest_feature_df = pd.concat(
        [latest_feature_df, pd.get_dummies(latest_feature_df["area_canonica"], prefix="inj_area", dtype=float)],
        axis=1,
    )
    for feature in feature_columns:
        if feature not in latest_feature_df.columns:
            latest_feature_df[feature] = 0.0
    latest_feature_df = latest_feature_df.reindex(columns=list(latest_feature_df.columns) + [])

    target_configs = {
        "tipologia_causa_macro": "injury_cause",
        "parte_del_corpo_macro": "injury_body_part",
    }

    for target_column, artifact_prefix in target_configs.items():
        target_df = detail_df.loc[detail_df[target_column].notna()].copy()
        target_df[target_column] = collapse_rare_classes(target_df[target_column], min_count=20)
        train_df = target_df.loc[pd.to_datetime(target_df["target_event_month"]).isin(train_months)].copy()
        validation_df = target_df.loc[pd.to_datetime(target_df["target_event_month"]).isin(validation_months)].copy()
        test_df = target_df.loc[pd.to_datetime(target_df["target_event_month"]).isin(test_months)].copy()

        x_train = train_df[feature_columns]
        y_train = train_df[target_column]
        x_validation = validation_df[feature_columns]
        y_validation = validation_df[target_column]
        x_test = test_df[feature_columns]
        y_test = test_df[target_column]

        validation_scores: dict[str, dict[str, float]] = {}
        for name, model in candidate_models.items():
            model.fit(x_train, y_train)
            validation_scores[name] = evaluate_multiclass_model(model, x_validation, y_validation)

        selected_name = max(validation_scores, key=lambda name: validation_scores[name]["macro_f1"])
        selected_model = candidate_models[selected_name]
        train_validation_df = pd.concat([train_df, validation_df], axis=0, ignore_index=True)
        selected_model.fit(train_validation_df[feature_columns], train_validation_df[target_column])

        test_metrics = evaluate_multiclass_model(selected_model, x_test, y_test)
        importance_df = feature_importance_frame(selected_model, feature_columns)
        importance_df.to_csv(output_dir / f"{artifact_prefix}_feature_importance.csv", index=False)

        test_predictions_df = build_top_k_predictions_frame(
            test_df[["event_id", "area_canonica", "source_month", "target_event_month", target_column]].copy(),
            selected_model,
            x_test,
            target_column,
        )
        test_predictions_df.to_csv(output_dir / f"{artifact_prefix}_test_predictions.csv", index=False)

        class_report_df = pd.DataFrame(
            classification_report(y_test, selected_model.predict(x_test), output_dict=True, zero_division=0)
        ).transpose().reset_index().rename(columns={"index": "label"})
        class_report_df.to_csv(output_dir / f"{artifact_prefix}_classification_report.csv", index=False)

        latest_x = latest_feature_df[feature_columns]
        latest_predictions_df = build_top_k_predictions_frame(
            latest_feature_df[["area_canonica", "source_month", "target_event_month"]].copy(),
            selected_model,
            latest_x,
            target_column,
        )
        latest_predictions_df.to_csv(output_dir / f"{artifact_prefix}_latest_predictions.csv", index=False)

        if joblib_dump is not None:
            joblib_dump(selected_model, output_dir / f"{artifact_prefix}_model.joblib")
        else:
            with (output_dir / f"{artifact_prefix}_model.pkl").open("wb") as handle:
                pickle.dump(selected_model, handle)

        report["targets"][target_column] = {
            "selected_model": selected_name,
            "n_rows": int(len(target_df)),
            "n_classes": int(target_df[target_column].nunique()),
            "class_distribution": {
                str(label): int(count)
                for label, count in target_df[target_column].value_counts().to_dict().items()
            },
            "validation_metrics": validation_scores,
            "test_metrics": test_metrics,
        }

    return report


def train_baseline_model(
    panel_df: pd.DataFrame,
    output_dir: Path,
    artifact_prefix: str = "",
) -> tuple[dict[str, Any], pd.DataFrame]:
    modeling_df = panel_df.loc[
        panel_df["has_observation_coverage"].eq(1) & panel_df["target_available"].eq(True)
    ].copy()

    unique_months = sorted(modeling_df["event_month"].drop_duplicates())
    if len(unique_months) < 18:
        raise RuntimeError("Servono almeno 18 mesi di storico condiviso per un baseline temporale affidabile.")

    test_months = unique_months[-12:]
    validation_months = unique_months[-24:-12]
    train_months = unique_months[:-24]
    if len(train_months) < 6:
        train_months = unique_months[:-18]
        validation_months = unique_months[-18:-9]
        test_months = unique_months[-9:]

    train_df = modeling_df.loc[modeling_df["event_month"].isin(train_months)].copy()
    validation_df = modeling_df.loc[modeling_df["event_month"].isin(validation_months)].copy()
    test_df = modeling_df.loc[modeling_df["event_month"].isin(test_months)].copy()

    feature_columns = pick_feature_columns(modeling_df)
    target_column = "target_adverse_next_month"
    x_train = train_df[feature_columns]
    y_train = train_df[target_column]
    x_validation = validation_df[feature_columns]
    y_validation = validation_df[target_column]
    x_test = test_df[feature_columns]
    y_test = test_df[target_column]

    candidate_models = build_baseline_candidate_models()

    validation_scores: dict[str, dict[str, float]] = {}
    for name, model in candidate_models.items():
        model.fit(x_train, y_train)
        validation_scores[name] = evaluate_classifier(model, x_validation, y_validation)

    selected_name = max(validation_scores, key=lambda name: validation_scores[name]["pr_auc"])
    selected_model = candidate_models[selected_name]
    train_validation_df = pd.concat([train_df, validation_df], axis=0, ignore_index=True)
    selected_model.fit(train_validation_df[feature_columns], train_validation_df[target_column])

    test_metrics = evaluate_classifier(selected_model, x_test, y_test)
    probabilities = selected_model.predict_proba(x_test)[:, 1]
    test_predictions_df = test_df[["area_canonica", "event_month", "target_month", target_column, "target_injury_next_month"]].copy()
    test_predictions_df["predicted_risk_score"] = probabilities
    test_predictions_df = test_predictions_df.sort_values(["predicted_risk_score", "target_month"], ascending=[False, True])
    test_predictions_df.to_csv(output_dir / artifact_name("test_set_predictions.csv", artifact_prefix), index=False)

    importance_df = feature_importance_frame(selected_model, feature_columns)
    importance_df.to_csv(output_dir / artifact_name("feature_importance.csv", artifact_prefix), index=False)

    if joblib_dump is not None:
        joblib_dump(selected_model, output_dir / artifact_name("baseline_model.joblib", artifact_prefix))
    else:
        with (output_dir / artifact_name("baseline_model.pkl", artifact_prefix)).open("wb") as handle:
            pickle.dump(selected_model, handle)

    latest_feature_month = panel_df.loc[
        panel_df["has_observation_coverage"].eq(1) & panel_df["event_month"].le(panel_df.loc[panel_df["evt_total_rows"].gt(0), "event_month"].max())
    , "event_month"].max()
    latest_scoring_df = panel_df.loc[
        panel_df["has_observation_coverage"].eq(1) & panel_df["event_month"].eq(latest_feature_month)
    ].copy()
    latest_scoring_df["predicted_risk_score"] = selected_model.predict_proba(latest_scoring_df[feature_columns])[:, 1]
    latest_scoring_df[["area_canonica", "event_month", "target_month", "predicted_risk_score"]].sort_values(
        "predicted_risk_score", ascending=False
    ).to_csv(output_dir / artifact_name("latest_risk_scores.csv", artifact_prefix), index=False)

    model_report = {
        "target_column": target_column,
        "positive_rate_train": float(y_train.mean()),
        "positive_rate_validation": float(y_validation.mean()),
        "positive_rate_test": float(y_test.mean()),
        "selected_model": selected_name,
        "validation_metrics": validation_scores,
        "test_metrics": test_metrics,
        "train_month_range": [train_months[0].strftime("%Y-%m-%d"), train_months[-1].strftime("%Y-%m-%d")],
        "validation_month_range": [validation_months[0].strftime("%Y-%m-%d"), validation_months[-1].strftime("%Y-%m-%d")],
        "test_month_range": [test_months[0].strftime("%Y-%m-%d"), test_months[-1].strftime("%Y-%m-%d")],
        "n_train_rows": int(len(train_df)),
        "n_validation_rows": int(len(validation_df)),
        "n_test_rows": int(len(test_df)),
        "n_features": int(len(feature_columns)),
        "artifact_prefix": artifact_prefix,
    }
    return model_report, latest_scoring_df


def run_walk_forward_backtest(
    panel_df: pd.DataFrame,
    output_dir: Path,
    min_history_months: int = 18,
    validation_window_months: int = 12,
    artifact_prefix: str = "",
) -> dict[str, Any]:
    modeling_df = panel_df.loc[
        panel_df["has_observation_coverage"].eq(1) & panel_df["target_available"].eq(True)
    ].copy()
    unique_months = sorted(pd.to_datetime(modeling_df["event_month"]).drop_duplicates())
    if len(unique_months) <= min_history_months:
        return {"status": "skipped", "reason": "Storico insufficiente per eseguire un walk-forward backtest affidabile."}

    feature_columns = pick_feature_columns(modeling_df)
    target_column = "target_adverse_next_month"
    prediction_frames: list[pd.DataFrame] = []
    monthly_rows: list[dict[str, Any]] = []

    for month_index in range(min_history_months, len(unique_months)):
        test_month = unique_months[month_index]
        prior_months = unique_months[:month_index]
        validation_months = prior_months[-validation_window_months:]
        train_months = prior_months[:-len(validation_months)]
        if len(train_months) < 6 or len(validation_months) < 3:
            continue

        train_df = modeling_df.loc[modeling_df["event_month"].isin(train_months)].copy()
        validation_df = modeling_df.loc[modeling_df["event_month"].isin(validation_months)].copy()
        test_df = modeling_df.loc[modeling_df["event_month"].eq(test_month)].copy()
        if train_df.empty or validation_df.empty or test_df.empty:
            continue

        x_train = train_df[feature_columns]
        y_train = train_df[target_column]
        x_validation = validation_df[feature_columns]
        y_validation = validation_df[target_column]
        x_test = test_df[feature_columns]
        y_test = test_df[target_column]

        candidate_models = build_baseline_candidate_models()
        validation_scores: dict[str, dict[str, float]] = {}
        for name, model in candidate_models.items():
            model.fit(x_train, y_train)
            validation_scores[name] = evaluate_classifier(model, x_validation, y_validation)

        selected_name = max(validation_scores, key=lambda name: validation_scores[name]["pr_auc"])
        selected_model = candidate_models[selected_name]
        train_validation_df = pd.concat([train_df, validation_df], axis=0, ignore_index=True)
        selected_model.fit(train_validation_df[feature_columns], train_validation_df[target_column])

        test_probabilities = selected_model.predict_proba(x_test)[:, 1]
        test_predictions_df = test_df[
            [
                "area_canonica",
                "event_month",
                "target_month",
                "target_adverse_next_month",
                "target_injury_next_month",
                "target_event_count_next_month",
            ]
        ].copy()
        test_predictions_df["selected_model"] = selected_name
        test_predictions_df["validation_pr_auc_selected_model"] = validation_scores[selected_name]["pr_auc"]
        test_predictions_df["validation_pr_auc_logistic_regression"] = validation_scores["logistic_regression"]["pr_auc"]
        test_predictions_df["validation_pr_auc_random_forest"] = validation_scores["random_forest"]["pr_auc"]
        test_predictions_df["predicted_risk_score"] = test_probabilities
        test_predictions_df = test_predictions_df.sort_values("predicted_risk_score", ascending=False).reset_index(drop=True)
        test_predictions_df["predicted_rank_within_month"] = np.arange(1, len(test_predictions_df) + 1)
        prediction_frames.append(test_predictions_df)

        actual_positive_areas = int(y_test.sum())
        top_1_actual = int(test_predictions_df.iloc[0]["target_adverse_next_month"]) if not test_predictions_df.empty else 0
        top_3_df = test_predictions_df.head(min(3, len(test_predictions_df)))
        top_3_hits = int(top_3_df["target_adverse_next_month"].sum()) if not top_3_df.empty else 0
        monthly_rows.append(
            {
                "event_month": pd.Timestamp(test_month).strftime("%Y-%m-%d"),
                "target_month": pd.Timestamp(test_predictions_df["target_month"].iloc[0]).strftime("%Y-%m-%d"),
                "n_areas": int(len(test_predictions_df)),
                "actual_positive_areas": actual_positive_areas,
                "avg_predicted_risk_score": float(test_predictions_df["predicted_risk_score"].mean()),
                "selected_model": selected_name,
                "validation_pr_auc_selected_model": float(validation_scores[selected_name]["pr_auc"]),
                "top_1_area": test_predictions_df.iloc[0]["area_canonica"],
                "top_1_score": float(test_predictions_df.iloc[0]["predicted_risk_score"]),
                "top_1_actual": top_1_actual,
                "top_3_hits": top_3_hits,
                "top_3_precision": float(top_3_hits / len(top_3_df)) if len(top_3_df) > 0 else 0.0,
                "top_3_recall": float(top_3_hits / actual_positive_areas) if actual_positive_areas > 0 else 0.0,
            }
        )

    if not prediction_frames:
        return {"status": "skipped", "reason": "Nessun mese eleggibile per il walk-forward backtest."}

    backtest_predictions_df = pd.concat(prediction_frames, axis=0, ignore_index=True)
    backtest_predictions_df.to_csv(output_dir / artifact_name("walk_forward_predictions.csv", artifact_prefix), index=False)

    monthly_summary_df = pd.DataFrame(monthly_rows).sort_values("target_month").reset_index(drop=True)
    monthly_summary_df.to_csv(output_dir / artifact_name("walk_forward_monthly_summary.csv", artifact_prefix), index=False)

    overall_metrics = {
        "pr_auc": float(average_precision_score(backtest_predictions_df[target_column], backtest_predictions_df["predicted_risk_score"])),
        "brier_score": float(brier_score_loss(backtest_predictions_df[target_column], backtest_predictions_df["predicted_risk_score"])),
        "roc_auc": float(roc_auc_score(backtest_predictions_df[target_column], backtest_predictions_df["predicted_risk_score"]))
        if backtest_predictions_df[target_column].nunique() > 1
        else float("nan"),
    }

    selected_model_counts = {
        str(label): int(count)
        for label, count in monthly_summary_df["selected_model"].value_counts().to_dict().items()
    }
    report = {
        "status": "completed",
        "target_column": target_column,
        "min_history_months": int(min_history_months),
        "validation_window_months": int(validation_window_months),
        "n_prediction_rows": int(len(backtest_predictions_df)),
        "n_backtest_months": int(monthly_summary_df["target_month"].nunique()),
        "event_month_range": [
            pd.to_datetime(backtest_predictions_df["event_month"]).min().strftime("%Y-%m-%d"),
            pd.to_datetime(backtest_predictions_df["event_month"]).max().strftime("%Y-%m-%d"),
        ],
        "target_month_range": [
            pd.to_datetime(backtest_predictions_df["target_month"]).min().strftime("%Y-%m-%d"),
            pd.to_datetime(backtest_predictions_df["target_month"]).max().strftime("%Y-%m-%d"),
        ],
        "overall_metrics": overall_metrics,
        "business_metrics": {
            "top_1_hit_rate": float(monthly_summary_df["top_1_actual"].mean()),
            "top_3_precision_mean": float(monthly_summary_df["top_3_precision"].mean()),
            "top_3_recall_mean": float(monthly_summary_df["top_3_recall"].mean()),
            "mean_actual_positive_areas": float(monthly_summary_df["actual_positive_areas"].mean()),
            "mean_predicted_score": float(monthly_summary_df["avg_predicted_risk_score"].mean()),
        },
        "selected_model_counts": selected_model_counts,
        "artifact_prefix": artifact_prefix,
    }
    return report


def aggregate_observation_metrics(df: pd.DataFrame, group_columns: list[str]) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=group_columns + ["obs_rows", "unsafe_total", "safe_total", "rows_with_unsafe", "total_behaviors", "unsafe_rate", "unsafe_row_rate"])

    aggregated = (
        df.groupby(group_columns, dropna=False)
        .agg(
            obs_rows=("unsafe_flag", "size"),
            unsafe_total=("comportamenti_non_sicuri", "sum"),
            safe_total=("comportamenti_sicuri", "sum"),
            rows_with_unsafe=("unsafe_flag", "sum"),
        )
        .reset_index()
    )
    aggregated["total_behaviors"] = aggregated["unsafe_total"] + aggregated["safe_total"]
    aggregated["unsafe_rate"] = np.where(
        aggregated["total_behaviors"].gt(0),
        aggregated["unsafe_total"] / aggregated["total_behaviors"],
        0.0,
    )
    aggregated["unsafe_row_rate"] = np.where(
        aggregated["obs_rows"].gt(0),
        aggregated["rows_with_unsafe"] / aggregated["obs_rows"],
        0.0,
    )
    return aggregated


def severity_weight_from_rank(rank: Any) -> float:
    numeric_rank = pd.to_numeric(pd.Series([rank]), errors="coerce").iloc[0]
    if pd.isna(numeric_rank):
        numeric_rank = 3
    normalized_rank = int(np.clip(round(float(numeric_rank)), 1, 5))
    return float(DEEPDIVE_SEVERITY_WEIGHTS.get(normalized_rank, DEEPDIVE_SEVERITY_WEIGHTS[3]))


def add_effective_behavior_severity_rank(
    behavior_df: pd.DataFrame,
    severity_column: str = "severity_rank_v2",
    family_column: str = "comportamento_famiglia",
) -> pd.DataFrame:
    if behavior_df.empty or severity_column not in behavior_df.columns:
        return behavior_df

    enriched = behavior_df.copy()
    severity_rank = pd.to_numeric(enriched[severity_column], errors="coerce").clip(lower=1, upper=5)
    enriched["severity_rank_v2_numeric"] = severity_rank

    if family_column in enriched.columns:
        family_rank = (
            enriched.groupby(family_column, dropna=False)["severity_rank_v2_numeric"]
            .transform("median")
            .round()
            .clip(lower=1, upper=5)
        )
        # IR 1-2 resta guidato dalla sottofamiglia; IR >=3 eredita anche il profilo della famiglia madre.
        effective_rank = np.where(
            severity_rank.le(2),
            severity_rank,
            np.fmax(severity_rank.fillna(3), family_rank.fillna(3)),
        )
    else:
        effective_rank = severity_rank

    enriched["effective_severity_rank"] = (
        pd.Series(effective_rank, index=enriched.index)
        .fillna(3)
        .round()
        .clip(lower=1, upper=5)
        .astype(int)
    )
    enriched["effective_severity_weight"] = enriched["effective_severity_rank"].map(DEEPDIVE_SEVERITY_WEIGHTS).fillna(DEEPDIVE_SEVERITY_WEIGHTS[3])
    return enriched


def build_area_deepdive_table(
    window_df: pd.DataFrame,
    latest_scores_df: pd.DataFrame,
    value_column: str,
    min_obs_rows: int,
    analysis_window_start: pd.Timestamp,
    analysis_window_end: pd.Timestamp,
    recent_window_start: pd.Timestamp,
    severity_column: str | None = None,
) -> pd.DataFrame:
    if value_column not in window_df.columns:
        return pd.DataFrame()

    working = window_df.copy()
    working[value_column] = working[value_column].map(normalize_whitespace)
    working = working.loc[working[value_column].notna()].copy()
    if working.empty:
        return pd.DataFrame()

    severity_available = bool(severity_column and severity_column in working.columns)
    if severity_available:
        working[severity_column] = pd.to_numeric(working[severity_column], errors="coerce").clip(lower=1, upper=5)

    area_stats = aggregate_observation_metrics(window_df, ["area_canonica"]).rename(
        columns={
            "obs_rows": "area_obs_rows",
            "unsafe_total": "area_unsafe_total",
            "safe_total": "area_safe_total",
            "rows_with_unsafe": "area_rows_with_unsafe",
            "total_behaviors": "area_total_behaviors",
            "unsafe_rate": "area_unsafe_rate",
            "unsafe_row_rate": "area_unsafe_row_rate",
        }
    )
    company_stats = aggregate_observation_metrics(working, [value_column]).rename(
        columns={
            "obs_rows": "company_obs_rows",
            "unsafe_total": "company_unsafe_total",
            "safe_total": "company_safe_total",
            "rows_with_unsafe": "company_rows_with_unsafe",
            "total_behaviors": "company_total_behaviors",
            "unsafe_rate": "company_unsafe_rate",
            "unsafe_row_rate": "company_unsafe_row_rate",
        }
    )
    area_item_stats = aggregate_observation_metrics(working, ["area_canonica", value_column])
    active_months = (
        working.groupby(["area_canonica", value_column], dropna=False)["event_month"]
        .nunique()
        .reset_index(name="active_months_12m")
    )
    area_item_stats = area_item_stats.merge(active_months, on=["area_canonica", value_column], how="left")
    if severity_available:
        severity_stats = (
            working.groupby(["area_canonica", value_column], dropna=False)[severity_column]
            .agg(
                severity_rank="median",
                severity_rank_min="min",
                severity_rank_max="max",
            )
            .reset_index()
        )
        severity_stats["severity_rank"] = severity_stats["severity_rank"].round().clip(lower=1, upper=5)
        area_item_stats = area_item_stats.merge(severity_stats, on=["area_canonica", value_column], how="left")

    recent_df = working.loc[working["event_month"].ge(recent_window_start)].copy()
    prior_df = working.loc[working["event_month"].lt(recent_window_start)].copy()

    recent_stats = aggregate_observation_metrics(recent_df, ["area_canonica", value_column]).rename(
        columns={
            "obs_rows": "recent_obs_rows",
            "unsafe_total": "recent_unsafe_total",
            "safe_total": "recent_safe_total",
            "rows_with_unsafe": "recent_rows_with_unsafe",
            "total_behaviors": "recent_total_behaviors",
            "unsafe_rate": "recent_unsafe_rate",
            "unsafe_row_rate": "recent_unsafe_row_rate",
        }
    )
    prior_stats = aggregate_observation_metrics(prior_df, ["area_canonica", value_column]).rename(
        columns={
            "obs_rows": "prior_obs_rows",
            "unsafe_total": "prior_unsafe_total",
            "safe_total": "prior_safe_total",
            "rows_with_unsafe": "prior_rows_with_unsafe",
            "total_behaviors": "prior_total_behaviors",
            "unsafe_rate": "prior_unsafe_rate",
            "unsafe_row_rate": "prior_unsafe_row_rate",
        }
    )

    merged = (
        latest_scores_df[["area_canonica", "event_month", "target_month", "predicted_risk_score", "area_risk_rank"]]
        .merge(area_stats, on="area_canonica", how="left")
        .merge(area_item_stats, on=["area_canonica"], how="inner")
        .merge(company_stats, on=[value_column], how="left")
        .merge(recent_stats, on=["area_canonica", value_column], how="left")
        .merge(prior_stats, on=["area_canonica", value_column], how="left")
    )

    merged = merged.loc[merged["obs_rows"].ge(min_obs_rows)].copy()
    if merged.empty:
        return merged

    for column in [
        "recent_obs_rows",
        "recent_unsafe_total",
        "recent_safe_total",
        "recent_rows_with_unsafe",
        "recent_total_behaviors",
        "recent_unsafe_rate",
        "recent_unsafe_row_rate",
        "prior_obs_rows",
        "prior_unsafe_total",
        "prior_safe_total",
        "prior_rows_with_unsafe",
        "prior_total_behaviors",
        "prior_unsafe_rate",
        "prior_unsafe_row_rate",
    ]:
        if column in merged.columns:
            merged[column] = merged[column].fillna(0.0)

    merged["unsafe_share_of_area"] = np.where(
        merged["area_unsafe_total"].gt(0),
        merged["unsafe_total"] / merged["area_unsafe_total"],
        0.0,
    )
    merged["obs_share_of_area"] = np.where(
        merged["area_obs_rows"].gt(0),
        merged["obs_rows"] / merged["area_obs_rows"],
        0.0,
    )
    merged["lift_vs_area"] = np.where(
        merged["area_unsafe_rate"].gt(0),
        merged["unsafe_rate"] / merged["area_unsafe_rate"],
        0.0,
    )
    merged["lift_vs_company_same_item"] = np.where(
        merged["company_unsafe_rate"].gt(0),
        merged["unsafe_rate"] / merged["company_unsafe_rate"],
        0.0,
    )
    merged["trend_delta_unsafe_rate"] = merged["recent_unsafe_rate"] - merged["prior_unsafe_rate"]
    merged["trend_delta_unsafe_row_rate"] = merged["recent_unsafe_row_rate"] - merged["prior_unsafe_row_rate"]

    positive_trend = merged["trend_delta_unsafe_rate"].clip(lower=0.0)
    merged["priority_score_raw"] = (
        merged["unsafe_rate"]
        * np.log1p(merged["obs_rows"])
        * np.sqrt(merged["lift_vs_area"].clip(lower=0.0))
        * np.sqrt(merged["lift_vs_company_same_item"].replace([np.inf, -np.inf], np.nan).fillna(1.0).clip(lower=0.1))
        * (1.0 + positive_trend)
    )
    if severity_available and "severity_rank" in merged.columns:
        merged["severity_rank"] = (
            pd.to_numeric(merged["severity_rank"], errors="coerce")
            .fillna(3)
            .round()
            .clip(lower=1, upper=5)
            .astype(int)
        )
        merged["severity_weight"] = merged["severity_rank"].map(DEEPDIVE_SEVERITY_WEIGHTS).fillna(DEEPDIVE_SEVERITY_WEIGHTS[3])
        merged["priority_score"] = merged["priority_score_raw"] * merged["severity_weight"]
    else:
        merged["priority_score"] = merged["priority_score_raw"]
    merged["analysis_window_start"] = analysis_window_start.strftime("%Y-%m-%d")
    merged["analysis_window_end"] = analysis_window_end.strftime("%Y-%m-%d")
    merged["recent_window_start"] = recent_window_start.strftime("%Y-%m-%d")
    merged["priority_rank_within_area"] = (
        merged.groupby("area_canonica")["priority_score"].rank(method="first", ascending=False).astype(int)
    )

    ordered = merged.sort_values(
        ["area_risk_rank", "priority_score", "unsafe_share_of_area", "obs_rows"],
        ascending=[True, False, False, False],
    ).reset_index(drop=True)
    return ordered


def write_area_deepdive_report(
    output_dir: Path,
    area_context_df: pd.DataFrame,
    family_df: pd.DataFrame,
    activity_df: pd.DataFrame,
    behavior_family_df: pd.DataFrame,
    behavior_subfamily_df: pd.DataFrame,
    behavior_df: pd.DataFrame,
    analysis_window_start: pd.Timestamp,
    analysis_window_end: pd.Timestamp,
) -> None:
    lines: list[str] = []
    lines.append("# Deepdive Attivita e Comportamenti per Area")
    lines.append("")
    lines.append(
        f"Finestra osservazioni analizzata: `{analysis_window_start.strftime('%Y-%m-%d')}` -> `{analysis_window_end.strftime('%Y-%m-%d')}`."
    )
    lines.append(
        "Il ranking combina volume osservato, unsafe rate, peso sul totale unsafe dell'area, scostamento rispetto al benchmark plant e, dove disponibile, severita HSE."
    )
    lines.append("")

    top_areas = area_context_df.sort_values("predicted_risk_score", ascending=False).head(3)
    for _, area_row in top_areas.iterrows():
        area_name = area_row["area_canonica"]
        lines.append(f"## {area_name}")
        lines.append("")
        lines.append(
            f"- Score rischio mese successivo: `{area_row['predicted_risk_score']:.3f}` per `target_month={area_row['target_month']}`."
        )
        lines.append(
            f"- Unsafe rate ultimi 12 mesi: `{area_row['area_unsafe_rate_12m']:.3%}` su `{int(area_row['area_obs_rows_12m'])}` righe osservate."
        )

        family_rows = family_df.loc[family_df["area_canonica"].eq(area_name)].head(3)
        if not family_rows.empty:
            family_summary = "; ".join(
                f"{row['attivita_famiglia']} ({row['unsafe_rate']:.1%}, lift area {row['lift_vs_area']:.2f}x)"
                for _, row in family_rows.iterrows()
            )
            lines.append(f"- Famiglie da attenzionare: {family_summary}.")

        activity_rows = activity_df.loc[activity_df["area_canonica"].eq(area_name)].head(3)
        if not activity_rows.empty:
            activity_summary = "; ".join(
                f"{row['attivita_clean']} ({row['unsafe_rate']:.1%}, quota unsafe area {row['unsafe_share_of_area']:.1%})"
                for _, row in activity_rows.iterrows()
            )
            lines.append(f"- Attivita prioritarie: {activity_summary}.")

        behavior_family_rows = behavior_family_df.loc[behavior_family_df["area_canonica"].eq(area_name)].head(3)
        if not behavior_family_rows.empty:
            behavior_family_summary = "; ".join(
                f"{row['comportamento_famiglia']} ({row['unsafe_rate']:.1%}, IR {int(row['severity_rank']) if 'severity_rank' in row and pd.notna(row['severity_rank']) else 'n.d.'}, lift area {row['lift_vs_area']:.2f}x)"
                for _, row in behavior_family_rows.iterrows()
            )
            lines.append(f"- Famiglie comportamento: {behavior_family_summary}.")

        behavior_subfamily_rows = behavior_subfamily_df.loc[behavior_subfamily_df["area_canonica"].eq(area_name)].head(3)
        if not behavior_subfamily_rows.empty:
            behavior_subfamily_summary = "; ".join(
                f"{row['comportamento_sottofamiglia_v2']} ({row['unsafe_rate']:.1%}, IR {int(row['severity_rank']) if 'severity_rank' in row and pd.notna(row['severity_rank']) else 'n.d.'}, lift area {row['lift_vs_area']:.2f}x)"
                for _, row in behavior_subfamily_rows.iterrows()
            )
            lines.append(f"- Sottofamiglie comportamento: {behavior_subfamily_summary}.")

        behavior_rows = behavior_df.loc[behavior_df["area_canonica"].eq(area_name)].head(5)
        if not behavior_rows.empty:
            behavior_summary = "; ".join(
                f"{row['comportamento_focus']} ({row['unsafe_rate']:.1%}, {int(row['obs_rows'])} righe)"
                for _, row in behavior_rows.iterrows()
            )
            lines.append(f"- Comportamenti critici: {behavior_summary}.")
        lines.append("")

    lines.append("## Come usarlo")
    lines.append("")
    lines.append("- Partire da `area_risk_context_latest.csv` per ordinare le aree per score e unsafe rate recente.")
    lines.append("- Filtrare `area_activity_family_deepdive.csv`, `area_activity_deepdive.csv`, `area_behavior_family_deepdive.csv`, `area_behavior_subfamily_deepdive.csv` e `area_behavior_deepdive.csv` per `area_canonica`.")
    lines.append("- Usare `priority_rank_within_area <= 10` come shortlist iniziale per audit, coaching o interventi tecnici.")
    lines.append("")
    lines.append("## Nota")
    lines.append("")
    lines.append(
        "Questi file sono un deepdive diagnostico: non dimostrano causalita, ma aiutano a localizzare dove l'unsafe e piu concentrato o anomalo dentro l'area."
    )
    lines.append(
        "Nel ranking comportamento sono escluse le righe di messaggistica e le domande checklist (`near miss`, `altri comportamenti`) per evitare falsi segnali."
    )
    lines.append(
        "Nel ranking attivita sono escluse le etichette non operative e troppo generiche (`Messaggio`, `Incidenti sfiorati`, `Altri comportamenti`, `Tutte le attivita`)."
    )

    (output_dir / "area_risk_deepdive_latest.md").write_text("\n".join(lines).strip() + "\n", encoding="utf-8")


def build_area_risk_deepdive(
    observations_df: pd.DataFrame,
    latest_scores_df: pd.DataFrame,
    output_dir: Path,
) -> dict[str, Any]:
    if observations_df.empty or latest_scores_df.empty:
        return {"status": "skipped", "reason": "Osservazioni o latest risk scores non disponibili."}

    latest_scores = latest_scores_df.copy()
    latest_scores["event_month"] = pd.to_datetime(latest_scores["event_month"])
    latest_scores["target_month"] = pd.to_datetime(latest_scores["target_month"])
    latest_feature_month = latest_scores["event_month"].max()
    latest_scores = latest_scores.loc[latest_scores["event_month"].eq(latest_feature_month)].copy()
    latest_scores = latest_scores.sort_values("predicted_risk_score", ascending=False).reset_index(drop=True)
    latest_scores["area_risk_rank"] = np.arange(1, len(latest_scores) + 1)

    observations = observations_df.copy()
    observations["event_month"] = pd.to_datetime(observations["event_month"])
    observations = observations.loc[observations["area_canonica"].isin(latest_scores["area_canonica"])].copy()
    canonical_focus = (
        observations["comportamento_canonico"]
        if "comportamento_canonico" in observations.columns
        else pd.Series(index=observations.index, dtype="object")
    )
    observations["comportamento_focus"] = (
        canonical_focus
        .fillna(observations["comportamento_clean"])
        .fillna(observations["comportamento_visualizzato"])
        .fillna(observations["comportamento"])
    )
    observations["comportamento_focus"] = observations["comportamento_focus"].map(normalize_whitespace)

    analysis_window_end = latest_feature_month
    analysis_window_start = latest_feature_month - pd.DateOffset(months=11)
    recent_window_start = latest_feature_month - pd.DateOffset(months=5)
    window_df = observations.loc[
        observations["event_month"].between(analysis_window_start, analysis_window_end, inclusive="both")
    ].copy()
    if window_df.empty:
        return {"status": "skipped", "reason": "Nessuna osservazione disponibile nella finestra di analisi recente."}

    area_context_df = aggregate_observation_metrics(window_df, ["area_canonica"]).rename(
        columns={
            "obs_rows": "area_obs_rows_12m",
            "unsafe_total": "area_unsafe_total_12m",
            "safe_total": "area_safe_total_12m",
            "rows_with_unsafe": "area_rows_with_unsafe_12m",
            "total_behaviors": "area_total_behaviors_12m",
            "unsafe_rate": "area_unsafe_rate_12m",
            "unsafe_row_rate": "area_unsafe_row_rate_12m",
        }
    )
    active_months = (
        window_df.groupby("area_canonica", dropna=False)["event_month"]
        .nunique()
        .reset_index(name="active_months_12m")
    )
    area_context_df = (
        latest_scores[["area_canonica", "event_month", "target_month", "predicted_risk_score", "area_risk_rank"]]
        .merge(area_context_df, on="area_canonica", how="left")
        .merge(active_months, on="area_canonica", how="left")
        .sort_values(["area_risk_rank"], ascending=[True])
        .reset_index(drop=True)
    )
    area_context_df["analysis_window_start"] = analysis_window_start.strftime("%Y-%m-%d")
    area_context_df["analysis_window_end"] = analysis_window_end.strftime("%Y-%m-%d")
    area_context_df.to_csv(output_dir / "area_risk_context_latest.csv", index=False)

    family_df = build_area_deepdive_table(
        window_df,
        latest_scores,
        value_column="attivita_famiglia",
        min_obs_rows=30,
        analysis_window_start=analysis_window_start,
        analysis_window_end=analysis_window_end,
        recent_window_start=recent_window_start,
    )
    family_df.to_csv(output_dir / "area_activity_family_deepdive.csv", index=False)

    activity_window_df = window_df.loc[~window_df["attivita_clean"].isin(NON_OPERATIONAL_ACTIVITY_LABELS)].copy()
    activity_df = build_area_deepdive_table(
        activity_window_df,
        latest_scores,
        value_column="attivita_clean",
        min_obs_rows=25,
        analysis_window_start=analysis_window_start,
        analysis_window_end=analysis_window_end,
        recent_window_start=recent_window_start,
    )
    activity_df.to_csv(output_dir / "area_activity_deepdive.csv", index=False)

    behavior_window_df = window_df.loc[window_df["comportamento_tipo"].eq(BEHAVIOR_KIND_BEHAVIOR)].copy()
    behavior_window_df = add_effective_behavior_severity_rank(behavior_window_df)

    behavior_family_df = build_area_deepdive_table(
        behavior_window_df,
        latest_scores,
        value_column="comportamento_famiglia",
        min_obs_rows=25,
        analysis_window_start=analysis_window_start,
        analysis_window_end=analysis_window_end,
        recent_window_start=recent_window_start,
        severity_column="effective_severity_rank",
    )
    behavior_family_df.to_csv(output_dir / "area_behavior_family_deepdive.csv", index=False)

    if "comportamento_sottofamiglia_v2" in behavior_window_df.columns:
        behavior_subfamily_window_df = behavior_window_df.loc[
            behavior_window_df["comportamento_sottofamiglia_v2"].notna()
        ].copy()
    else:
        behavior_subfamily_window_df = behavior_window_df.iloc[0:0].copy()
    behavior_subfamily_df = build_area_deepdive_table(
        behavior_subfamily_window_df,
        latest_scores,
        value_column="comportamento_sottofamiglia_v2",
        min_obs_rows=20,
        analysis_window_start=analysis_window_start,
        analysis_window_end=analysis_window_end,
        recent_window_start=recent_window_start,
        severity_column="effective_severity_rank",
    )
    behavior_subfamily_df.to_csv(output_dir / "area_behavior_subfamily_deepdive.csv", index=False)

    behavior_df = build_area_deepdive_table(
        behavior_window_df,
        latest_scores,
        value_column="comportamento_focus",
        min_obs_rows=20,
        analysis_window_start=analysis_window_start,
        analysis_window_end=analysis_window_end,
        recent_window_start=recent_window_start,
        severity_column="effective_severity_rank",
    )
    behavior_df.to_csv(output_dir / "area_behavior_deepdive.csv", index=False)

    write_area_deepdive_report(
        output_dir=output_dir,
        area_context_df=area_context_df,
        family_df=family_df,
        activity_df=activity_df,
        behavior_family_df=behavior_family_df,
        behavior_subfamily_df=behavior_subfamily_df,
        behavior_df=behavior_df,
        analysis_window_start=analysis_window_start,
        analysis_window_end=analysis_window_end,
    )

    return {
        "status": "built",
        "analysis_window_start": analysis_window_start.strftime("%Y-%m-%d"),
        "analysis_window_end": analysis_window_end.strftime("%Y-%m-%d"),
        "n_areas": int(area_context_df["area_canonica"].nunique()),
        "family_rows": int(len(family_df)),
        "activity_rows": int(len(activity_df)),
        "behavior_family_rows": int(len(behavior_family_df)),
        "behavior_subfamily_rows": int(len(behavior_subfamily_df)),
        "behavior_rows": int(len(behavior_df)),
    }


def shorten_text(value: Any, limit: int = 120) -> str:
    text = normalize_whitespace(value) or ""
    if len(text) <= limit:
        return text
    return text[: max(0, limit - 3)].rstrip() + "..."


def area_top_rows(df: pd.DataFrame, area_name: str, limit: int = 3) -> pd.DataFrame:
    if df.empty or "area_canonica" not in df.columns:
        return df.iloc[0:0].copy()
    return df.loc[df["area_canonica"].eq(area_name)].head(limit).copy()


def format_named_rows(rows_df: pd.DataFrame, label_column: str, detail_formatter: Any, limit: int = 3) -> str:
    if rows_df.empty:
        return ""
    chunks = []
    for _, row in rows_df.head(limit).iterrows():
        label = normalize_whitespace(row.get(label_column)) or "-"
        detail = detail_formatter(row)
        chunks.append(f"{label} ({detail})" if detail else label)
    return "; ".join(chunks)


def split_pipe_labels(value: Any) -> list[str]:
    text = normalize_whitespace(value)
    if text is None:
        return []
    return [label for label in (normalize_whitespace(chunk) for chunk in str(text).split("|")) if label]


def parse_count_summary_labels(value: Any) -> list[str]:
    text = normalize_whitespace(value)
    if text is None:
        return []
    labels = []
    for chunk in str(text).split("|"):
        label = normalize_whitespace(str(chunk).split(":", 1)[0])
        if label:
            labels.append(label)
    return labels


def build_event_similarity_document(row: pd.Series) -> str:
    topic_labels = split_pipe_labels(row.get("descrizione_topic_labels"))
    parts = [
        normalize_whitespace(row.get("descrizione_clean")),
        " ".join(topic_labels) if topic_labels else None,
        normalize_whitespace(row.get("problem_family")),
        normalize_whitespace(row.get("problem_code")),
        normalize_whitespace(row.get("tipologia_canonica")),
        normalize_whitespace(row.get("dettaglio_area_clean")),
        normalize_whitespace(row.get("mansione_attivita")),
        normalize_whitespace(row.get("tipologia_causa_macro")),
        normalize_whitespace(row.get("parte_del_corpo_macro")),
    ]
    raw_text = " ".join(part for part in parts if part)
    return description_for_tfidf(raw_text)


def build_area_similarity_query(
    area_name: str,
    area_row: pd.Series,
    area_window_df: pd.DataFrame,
    top_behavior_family_rows: pd.DataFrame,
    top_behavior_rows: pd.DataFrame,
) -> dict[str, Any]:
    query_parts: list[str] = []
    query_strategy_parts: list[str] = []
    query_topic_labels: list[str] = []
    source_event_ids: list[Any] = []
    behavior_context_parts: list[str] = []

    adverse_source_df = area_window_df.loc[
        (
            area_window_df["is_first_aid"].eq(1)
            | area_window_df["is_certified_injury"].eq(1)
            | area_window_df["is_incident_without_injury"].eq(1)
        )
        & area_window_df["descrizione_clean"].notna()
    ].sort_values("event_date", ascending=False)
    if not adverse_source_df.empty:
        source_df = adverse_source_df.head(3).copy()
        query_parts.extend(source_df["descrizione_clean"].dropna().astype(str).tolist())
        source_event_ids = source_df["id"].tolist()
        query_strategy_parts.append("recent_adverse_descriptions")
    else:
        generic_source_df = area_window_df.loc[area_window_df["descrizione_clean"].notna()].sort_values("event_date", ascending=False)
        if not generic_source_df.empty:
            source_df = generic_source_df.head(3).copy()
            query_parts.extend(source_df["descrizione_clean"].dropna().astype(str).tolist())
            source_event_ids = source_df["id"].tolist()
            query_strategy_parts.append("recent_area_event_descriptions")

    topic_sources = (
        parse_count_summary_labels(area_row.get("top_adverse_description_topics_12m"))
        or parse_count_summary_labels(area_row.get("top_description_topics_12m"))
    )
    if topic_sources:
        query_topic_labels.extend(topic_sources[:3])
        query_parts.extend(topic_sources[:3])
        query_strategy_parts.append("description_topics")

    top_problem_codes = parse_count_summary_labels(area_row.get("top_problem_codes_12m"))
    if top_problem_codes:
        query_parts.extend(PROBLEM_FAMILY_LABELS.get(code, code) for code in top_problem_codes[:3])
        query_strategy_parts.append("problem_history")

    for value in top_behavior_family_rows["comportamento_famiglia"].head(2).tolist():
        label = normalize_whitespace(value)
        if label:
            behavior_context_parts.append(label)
    for value in top_behavior_rows["comportamento_focus"].head(3).tolist():
        label = normalize_whitespace(value)
        if label:
            behavior_context_parts.append(label)
    if behavior_context_parts:
        query_parts.extend(behavior_context_parts)
        query_strategy_parts.append("behavior_context")

    query_text_raw = " ".join(query_parts)
    query_text = description_for_tfidf(query_text_raw)
    return {
        "area_canonica": area_name,
        "query_strategy": " + ".join(dict.fromkeys(query_strategy_parts)) if query_strategy_parts else "empty",
        "query_text_raw": normalize_whitespace(query_text_raw) or "",
        "query_text_preview": shorten_text(query_text_raw, limit=320),
        "query_text_char_count": len(query_text_raw),
        "query_text": query_text,
        "query_topics": " | ".join(query_topic_labels),
        "source_event_count": len(source_event_ids),
        "source_event_ids": " | ".join(str(value) for value in source_event_ids),
    }


def build_area_similarity_retrieval_artifacts(
    output_dir: Path,
    events_df: pd.DataFrame,
    merged_area_context_df: pd.DataFrame,
    behavior_family_expert_df: pd.DataFrame,
    behavior_df: pd.DataFrame,
    analysis_window_start: pd.Timestamp,
    analysis_window_end: pd.Timestamp,
    top_n_cases: int = 3,
) -> tuple[dict[str, Any], pd.DataFrame]:
    query_output_path = output_dir / "area_similarity_queries_latest.csv"
    match_output_path = output_dir / "area_similar_historical_events_latest.csv"

    adverse_events_df = events_df.loc[
        events_df["is_first_aid"].eq(1)
        | events_df["is_certified_injury"].eq(1)
        | events_df["is_incident_without_injury"].eq(1)
    ].copy()
    adverse_events_df["event_date"] = pd.to_datetime(adverse_events_df["event_date"])
    adverse_events_df = adverse_events_df.loc[adverse_events_df["event_date"].le(analysis_window_end)].copy()
    adverse_events_df["similarity_document"] = adverse_events_df.apply(build_event_similarity_document, axis=1)
    adverse_events_df = adverse_events_df.loc[adverse_events_df["similarity_document"].str.len() > 0].copy()

    empty_summary_df = pd.DataFrame(
        columns=[
            "area_canonica",
            "similar_historical_events_summary",
            "similar_historical_event_patterns",
            "similar_historical_event_1",
            "similar_historical_event_2",
            "similar_historical_event_3",
        ]
    )
    if adverse_events_df.empty:
        pd.DataFrame(
            columns=[
                "area_canonica",
                "query_strategy",
                "query_topics",
                "source_event_count",
                "source_event_ids",
                "query_text_preview",
                "query_text_char_count",
            ]
        ).to_csv(query_output_path, index=False)
        pd.DataFrame(
            columns=[
                "area_canonica",
                "similarity_rank",
                "similarity_score",
                "shared_topic_count",
                "shared_description_topics",
                "same_area_flag",
                "matched_event_id",
                "matched_event_date",
                "matched_area_canonica",
                "matched_tipologia_canonica",
                "matched_problem_code",
                "matched_problem_family",
                "matched_description_topics",
                "matched_giorni_persi",
                "matched_event_summary",
            ]
        ).to_csv(match_output_path, index=False)
        return {
            "status": "skipped",
            "reason": "Nessun evento avverso descrivibile disponibile per il retrieval testuale.",
        }, empty_summary_df

    vectorizer = TfidfVectorizer(
        min_df=2,
        max_df=0.9,
        max_features=1200,
        ngram_range=(1, 2),
        stop_words=sorted(ITALIAN_STOPWORDS),
    )
    candidate_matrix = vectorizer.fit_transform(adverse_events_df["similarity_document"])

    window_df = events_df.copy()
    window_df["event_month"] = pd.to_datetime(window_df["event_month"])
    window_df["event_date"] = pd.to_datetime(window_df["event_date"])
    window_df = window_df.loc[
        window_df["event_month"].between(analysis_window_start, analysis_window_end, inclusive="both")
    ].copy()

    query_rows = []
    match_rows = []
    summary_rows = []
    areas_with_matches = 0

    for _, area_row in merged_area_context_df.iterrows():
        area_name = area_row["area_canonica"]
        area_window_df = window_df.loc[window_df["area_canonica"].eq(area_name)].copy()
        top_behavior_family_rows = area_top_rows(behavior_family_expert_df, area_name, limit=3)
        top_behavior_rows = area_top_rows(behavior_df, area_name, limit=3)

        query_payload = build_area_similarity_query(
            area_name=area_name,
            area_row=area_row,
            area_window_df=area_window_df,
            top_behavior_family_rows=top_behavior_family_rows,
            top_behavior_rows=top_behavior_rows,
        )
        query_rows.append(query_payload)

        if not query_payload["query_text"]:
            summary_rows.append(
                {
                    "area_canonica": area_name,
                    "similar_historical_events_summary": "",
                    "similar_historical_event_patterns": "",
                    "similar_historical_event_1": "",
                    "similar_historical_event_2": "",
                    "similar_historical_event_3": "",
                }
            )
            continue

        query_vector = vectorizer.transform([query_payload["query_text"]])
        similarity_scores = query_vector.dot(candidate_matrix.T).toarray().ravel()
        candidate_slice_df = adverse_events_df.copy()
        candidate_slice_df["similarity_score"] = similarity_scores
        candidate_slice_df = candidate_slice_df.loc[candidate_slice_df["similarity_score"] > 0].copy()

        source_event_ids = {str(value) for value in split_pipe_labels(query_payload["source_event_ids"])}
        if source_event_ids:
            candidate_slice_df = candidate_slice_df.loc[
                ~candidate_slice_df["id"].astype(str).isin(source_event_ids)
            ].copy()

        query_topics = set(split_pipe_labels(query_payload["query_topics"]))
        candidate_slice_df["matched_description_topics"] = candidate_slice_df["descrizione_topic_labels"].fillna("")
        candidate_slice_df["shared_description_topics"] = candidate_slice_df["matched_description_topics"].map(
            lambda value: " | ".join(
                [topic for topic in split_pipe_labels(value) if not query_topics or topic in query_topics]
            )
        )
        candidate_slice_df["shared_topic_count"] = candidate_slice_df["shared_description_topics"].map(
            lambda value: len(split_pipe_labels(value))
        )
        candidate_slice_df["same_area_flag"] = candidate_slice_df["area_canonica"].eq(area_name).astype(int)
        candidate_slice_df["matched_giorni_persi"] = candidate_slice_df["giorni_persi_capped"].fillna(0.0)
        candidate_slice_df = candidate_slice_df.sort_values(
            [
                "similarity_score",
                "shared_topic_count",
                "same_area_flag",
                "is_certified_injury",
                "matched_giorni_persi",
                "event_date",
            ],
            ascending=[False, False, False, False, False, False],
        ).reset_index(drop=True)

        top_matches_df = candidate_slice_df.head(top_n_cases).copy()
        if top_matches_df.empty:
            summary_rows.append(
                {
                    "area_canonica": area_name,
                    "similar_historical_events_summary": "",
                    "similar_historical_event_patterns": "",
                    "similar_historical_event_1": "",
                    "similar_historical_event_2": "",
                    "similar_historical_event_3": "",
                }
            )
            continue

        areas_with_matches += 1
        case_summaries = []
        topic_counter: Counter[str] = Counter()
        summary_payload = {
            "area_canonica": area_name,
            "similar_historical_events_summary": "",
            "similar_historical_event_patterns": "",
            "similar_historical_event_1": "",
            "similar_historical_event_2": "",
            "similar_historical_event_3": "",
        }

        for rank, (_, match_row) in enumerate(top_matches_df.iterrows(), start=1):
            shared_topics = split_pipe_labels(match_row["shared_description_topics"])
            topic_counter.update(shared_topics or split_pipe_labels(match_row["matched_description_topics"]))
            match_summary = " - ".join(
                [
                    pd.to_datetime(match_row["event_date"]).strftime("%Y-%m-%d"),
                    normalize_whitespace(match_row["area_canonica"]) or "Area non definita",
                    normalize_whitespace(match_row["tipologia_canonica"]) or "Evento",
                    normalize_whitespace(match_row["problem_code"]) or "UNKNOWN",
                    f"sim {float(match_row['similarity_score']):.2f}",
                    shorten_text(match_row["descrizione_clean"], limit=95),
                ]
            )
            case_summaries.append(match_summary)
            summary_payload[f"similar_historical_event_{rank}"] = match_summary
            match_rows.append(
                {
                    "area_canonica": area_name,
                    "similarity_rank": rank,
                    "similarity_score": float(match_row["similarity_score"]),
                    "shared_topic_count": int(match_row["shared_topic_count"]),
                    "shared_description_topics": " | ".join(shared_topics),
                    "same_area_flag": int(match_row["same_area_flag"]),
                    "matched_event_id": match_row["id"],
                    "matched_event_date": pd.to_datetime(match_row["event_date"]).strftime("%Y-%m-%d"),
                    "matched_area_canonica": normalize_whitespace(match_row["area_canonica"]),
                    "matched_tipologia_canonica": normalize_whitespace(match_row["tipologia_canonica"]),
                    "matched_problem_code": normalize_whitespace(match_row["problem_code"]),
                    "matched_problem_family": normalize_whitespace(match_row["problem_family"]),
                    "matched_description_topics": normalize_whitespace(match_row["matched_description_topics"]),
                    "matched_giorni_persi": float(match_row["matched_giorni_persi"]),
                    "matched_event_summary": match_summary,
                }
            )

        summary_payload["similar_historical_events_summary"] = " | ".join(case_summaries)
        summary_payload["similar_historical_event_patterns"] = " | ".join(
            f"{topic}:{count}" for topic, count in topic_counter.most_common(3)
        )
        summary_rows.append(summary_payload)

    pd.DataFrame(query_rows)[
        [
            "area_canonica",
            "query_strategy",
            "query_topics",
            "source_event_count",
            "source_event_ids",
            "query_text_preview",
            "query_text_char_count",
        ]
    ].to_csv(query_output_path, index=False)
    pd.DataFrame(match_rows).to_csv(match_output_path, index=False)
    summary_df = pd.DataFrame(summary_rows)
    return {
        "status": "built",
        "query_rows": int(len(query_rows)),
        "match_rows": int(len(match_rows)),
        "areas_with_matches": int(areas_with_matches),
        "top_n_cases_per_area": int(top_n_cases),
    }, summary_df


def build_recent_event_context(
    events_df: pd.DataFrame,
    analysis_window_start: pd.Timestamp,
    analysis_window_end: pd.Timestamp,
) -> pd.DataFrame:
    if events_df.empty:
        return pd.DataFrame(columns=["area_canonica"])

    window_df = events_df.copy()
    window_df["event_month"] = pd.to_datetime(window_df["event_month"])
    window_df["event_date"] = pd.to_datetime(window_df["event_date"])
    window_df = window_df.loc[
        window_df["event_month"].between(analysis_window_start, analysis_window_end, inclusive="both")
    ].copy()
    if window_df.empty:
        return pd.DataFrame(columns=["area_canonica"])

    base_df = (
        window_df.groupby("area_canonica", dropna=False)
        .agg(
            evt_total_rows_12m=("id", "count"),
            evt_smat_audit_count_12m=("is_smat_audit", "sum"),
            evt_first_aid_count_12m=("is_first_aid", "sum"),
            evt_certified_injury_count_12m=("is_certified_injury", "sum"),
            evt_incident_without_injury_count_12m=("is_incident_without_injury", "sum"),
            evt_injury_like_count_12m=("is_injury_like", "sum"),
            evt_days_lost_sum_12m=("giorni_persi_capped", "sum"),
        )
        .reset_index()
    )

    top_problem_rows_df = (
        window_df.groupby(["area_canonica", "problem_code"], dropna=False)
        .size()
        .reset_index(name="row_count")
        .sort_values(["area_canonica", "row_count", "problem_code"], ascending=[True, False, True])
        .groupby("area_canonica", dropna=False)
        .head(3)
    )
    problem_summary_rows = []
    for area_name, frame in top_problem_rows_df.groupby("area_canonica", dropna=False):
        summary = " | ".join(
            f"{normalize_whitespace(row['problem_code']) or 'UNKNOWN'}:{int(row['row_count'])}"
            for _, row in frame.iterrows()
        )
        problem_summary_rows.append({"area_canonica": area_name, "top_problem_codes_12m": summary})
    problem_summary_df = pd.DataFrame(problem_summary_rows)

    topic_rows = []
    for _, row in window_df.loc[window_df["descrizione_topic_labels"].fillna("").str.len() > 0].iterrows():
        labels = [label.strip() for label in str(row["descrizione_topic_labels"]).split("|") if normalize_whitespace(label)]
        for label in labels:
            topic_rows.append(
                {
                    "area_canonica": row["area_canonica"],
                    "description_topic": label,
                    "is_adverse_event": int(
                        row["is_first_aid"] == 1
                        or row["is_certified_injury"] == 1
                        or row["is_incident_without_injury"] == 1
                    ),
                }
            )
    if topic_rows:
        topic_context_df = pd.DataFrame(topic_rows)
        top_topic_rows_df = (
            topic_context_df.groupby(["area_canonica", "description_topic"], dropna=False)
            .agg(row_count=("description_topic", "count"))
            .reset_index()
            .sort_values(["area_canonica", "row_count", "description_topic"], ascending=[True, False, True])
            .groupby("area_canonica", dropna=False)
            .head(3)
        )
        top_topic_summary_rows = []
        for area_name, frame in top_topic_rows_df.groupby("area_canonica", dropna=False):
            summary = " | ".join(
                f"{normalize_whitespace(row['description_topic'])}:{int(row['row_count'])}"
                for _, row in frame.iterrows()
            )
            top_topic_summary_rows.append({"area_canonica": area_name, "top_description_topics_12m": summary})
        top_topic_summary_df = pd.DataFrame(top_topic_summary_rows)

        adverse_topic_rows_df = (
            topic_context_df.loc[topic_context_df["is_adverse_event"].eq(1)]
            .groupby(["area_canonica", "description_topic"], dropna=False)
            .agg(row_count=("description_topic", "count"))
            .reset_index()
            .sort_values(["area_canonica", "row_count", "description_topic"], ascending=[True, False, True])
            .groupby("area_canonica", dropna=False)
            .head(3)
        )
        adverse_topic_summary_rows = []
        for area_name, frame in adverse_topic_rows_df.groupby("area_canonica", dropna=False):
            summary = " | ".join(
                f"{normalize_whitespace(row['description_topic'])}:{int(row['row_count'])}"
                for _, row in frame.iterrows()
            )
            adverse_topic_summary_rows.append({"area_canonica": area_name, "top_adverse_description_topics_12m": summary})
        adverse_topic_summary_df = pd.DataFrame(adverse_topic_summary_rows)
    else:
        top_topic_summary_df = pd.DataFrame(columns=["area_canonica", "top_description_topics_12m"])
        adverse_topic_summary_df = pd.DataFrame(columns=["area_canonica", "top_adverse_description_topics_12m"])

    adverse_df = window_df.loc[
        window_df["is_first_aid"].eq(1)
        | window_df["is_certified_injury"].eq(1)
        | window_df["is_incident_without_injury"].eq(1)
    ].copy()
    if adverse_df.empty:
        latest_adverse_df = pd.DataFrame(
            columns=[
                "area_canonica",
                "latest_adverse_event_date",
                "latest_adverse_event_type",
                "latest_adverse_event_problem",
                "latest_adverse_event_summary",
                "recent_adverse_events_12m",
            ]
        )
    else:
        adverse_df = adverse_df.sort_values(["area_canonica", "event_date"], ascending=[True, False])
        latest_adverse_df = (
            adverse_df.groupby("area_canonica", dropna=False)
            .head(1)
            .assign(
                latest_adverse_event_date=lambda frame: frame["event_date"].dt.strftime("%Y-%m-%d"),
                latest_adverse_event_type=lambda frame: frame["tipologia_canonica"].map(normalize_whitespace),
                latest_adverse_event_problem=lambda frame: frame["problem_code"].map(normalize_whitespace),
            )[
                [
                    "area_canonica",
                    "latest_adverse_event_date",
                    "latest_adverse_event_type",
                    "latest_adverse_event_problem",
                ]
            ]
            .copy()
        )
        latest_adverse_df["latest_adverse_event_summary"] = latest_adverse_df.apply(
            lambda row: " - ".join(
                [
                    value
                    for value in [
                        normalize_whitespace(row["latest_adverse_event_date"]),
                        normalize_whitespace(row["latest_adverse_event_type"]),
                        normalize_whitespace(row["latest_adverse_event_problem"]),
                    ]
                    if value
                ]
            ),
            axis=1,
        )
        recent_adverse_summary_df = (
            adverse_df.groupby("area_canonica", dropna=False)
            .head(3)
            .assign(
                summary=lambda frame: frame.apply(
                    lambda row: " - ".join(
                        [
                            value
                            for value in [
                                pd.to_datetime(row["event_date"]).strftime("%Y-%m-%d") if pd.notna(row["event_date"]) else "",
                                normalize_whitespace(row["tipologia_canonica"]) or "",
                                normalize_whitespace(row["problem_code"]) or "",
                                shorten_text(row["descrizione_clean"], limit=70),
                            ]
                            if value
                        ]
                    ),
                    axis=1,
                )
            )
            .groupby("area_canonica", dropna=False)["summary"]
            .apply(lambda series: " | ".join(series.tolist()))
            .reset_index(name="recent_adverse_events_12m")
        )
        latest_adverse_df = latest_adverse_df.merge(recent_adverse_summary_df, on="area_canonica", how="left")

    return (
        base_df
        .merge(problem_summary_df, on="area_canonica", how="left")
        .merge(top_topic_summary_df, on="area_canonica", how="left")
        .merge(adverse_topic_summary_df, on="area_canonica", how="left")
        .merge(latest_adverse_df, on="area_canonica", how="left")
    )


def compose_area_investigation_hypothesis(
    behavior_family_rows: pd.DataFrame,
    event_context_row: pd.Series | None,
) -> str:
    behavior_family = normalize_whitespace(behavior_family_rows.iloc[0]["comportamento_famiglia"]) if not behavior_family_rows.empty else None
    second_behavior_family = normalize_whitespace(behavior_family_rows.iloc[1]["comportamento_famiglia"]) if len(behavior_family_rows) > 1 else None
    top_problem_codes = normalize_whitespace(event_context_row.get("top_problem_codes_12m")) if event_context_row is not None else None

    chunks = []
    if behavior_family:
        chunks.append(f"Rischio plausibilmente trainato da `{behavior_family}`.")

    if second_behavior_family:
        chunks.append(f"Segnale secondario su `{second_behavior_family}`.")
    if top_problem_codes:
        chunks.append(f"Storico eventi recente coerente su `{top_problem_codes}`.")
    return " ".join(chunks).strip()


def compose_area_field_checks(behavior_family_rows: pd.DataFrame) -> str:
    if behavior_family_rows.empty:
        return ""
    hints = []
    seen = set()
    for _, row in behavior_family_rows.head(3).iterrows():
        family = normalize_whitespace(row.get("comportamento_famiglia")) or ""
        hint = BEHAVIOR_FAMILY_INVESTIGATION_HINTS.get(
            family,
            "Verificare condizioni operative, standard e barriere collegati alla famiglia comportamento.",
        )
        if hint not in seen:
            hints.append(hint)
            seen.add(hint)
    return " | ".join(hints)


def build_area_investigation_pack(
    output_dir: Path,
    events_df: pd.DataFrame,
    expert_latest_scores_df: pd.DataFrame,
    behavior_risk_rank_df: pd.DataFrame,
    top_n_report: int = 5,
) -> dict[str, Any]:
    area_context_path = output_dir / "area_risk_context_latest.csv"
    behavior_family_path = output_dir / "area_behavior_family_deepdive.csv"
    behavior_subfamily_path = output_dir / "area_behavior_subfamily_deepdive.csv"
    behavior_path = output_dir / "area_behavior_deepdive.csv"
    required_paths = [
        area_context_path,
        behavior_family_path,
        behavior_path,
    ]
    if any(not path.exists() for path in required_paths):
        return {"status": "skipped", "reason": "Artifact deepdive mancanti per costruire l'investigation pack."}

    area_context_df = pd.read_csv(area_context_path, parse_dates=["event_month", "target_month"])
    behavior_family_df = pd.read_csv(behavior_family_path)
    behavior_subfamily_df = pd.read_csv(behavior_subfamily_path) if behavior_subfamily_path.exists() else pd.DataFrame()
    behavior_df = pd.read_csv(behavior_path)
    if area_context_df.empty:
        return {"status": "skipped", "reason": "Area context vuoto."}

    expert_scores_df = expert_latest_scores_df.copy()
    expert_scores_df["event_month"] = pd.to_datetime(expert_scores_df["event_month"])
    expert_scores_df["target_month"] = pd.to_datetime(expert_scores_df["target_month"])
    latest_expert_month = expert_scores_df["event_month"].max()
    expert_scores_df = expert_scores_df.loc[expert_scores_df["event_month"].eq(latest_expert_month)].copy()
    expert_scores_df = expert_scores_df.sort_values("predicted_risk_score", ascending=False).reset_index(drop=True)
    expert_scores_df["expert_area_risk_rank"] = np.arange(1, len(expert_scores_df) + 1)
    expert_scores_df = expert_scores_df.rename(columns={"predicted_risk_score": "expert_predicted_risk_score"})

    behavior_family_expert_df = behavior_family_df.merge(
        behavior_risk_rank_df[["comportamento_famiglia", "risk_rank", "risk_label", "rationale"]],
        on="comportamento_famiglia",
        how="left",
    )
    behavior_family_expert_df["risk_rank"] = pd.to_numeric(behavior_family_expert_df["risk_rank"], errors="coerce").fillna(3).astype(int)
    behavior_family_expert_df["risk_label"] = behavior_family_expert_df["risk_label"].fillna(behavior_family_expert_df["risk_rank"].map(EXPERT_RISK_LABELS))
    behavior_family_expert_df["severity_weight"] = 6 - behavior_family_expert_df["risk_rank"]
    behavior_family_expert_df["expert_priority_score"] = behavior_family_expert_df["priority_score"] * behavior_family_expert_df["severity_weight"]
    behavior_family_expert_df = behavior_family_expert_df.sort_values(
        ["area_risk_rank", "expert_priority_score", "priority_score", "risk_rank", "unsafe_total"],
        ascending=[True, False, False, True, False],
    ).reset_index(drop=True)
    behavior_family_expert_df.to_csv(output_dir / "area_behavior_family_expert_deepdive.csv", index=False)

    analysis_window_start = pd.to_datetime(area_context_df["analysis_window_start"].iloc[0])
    analysis_window_end = pd.to_datetime(area_context_df["analysis_window_end"].iloc[0])
    event_context_df = build_recent_event_context(events_df, analysis_window_start, analysis_window_end)

    pack_rows = []
    merged_area_context_df = (
        area_context_df.merge(
            expert_scores_df[["area_canonica", "expert_predicted_risk_score", "expert_area_risk_rank"]],
            on="area_canonica",
            how="left",
        )
        .merge(event_context_df, on="area_canonica", how="left")
        .sort_values("area_risk_rank")
        .reset_index(drop=True)
    )
    merged_area_context_df["expert_minus_baseline_risk_score"] = (
        merged_area_context_df["expert_predicted_risk_score"] - merged_area_context_df["predicted_risk_score"]
    )
    similarity_report, similar_event_summary_df = build_area_similarity_retrieval_artifacts(
        output_dir=output_dir,
        events_df=events_df,
        merged_area_context_df=merged_area_context_df,
        behavior_family_expert_df=behavior_family_expert_df,
        behavior_df=behavior_df,
        analysis_window_start=analysis_window_start,
        analysis_window_end=analysis_window_end,
    )
    merged_area_context_df = merged_area_context_df.merge(similar_event_summary_df, on="area_canonica", how="left")

    for _, area_row in merged_area_context_df.iterrows():
        area_name = area_row["area_canonica"]
        top_behavior_family_rows = area_top_rows(behavior_family_expert_df, area_name, limit=3)
        top_behavior_subfamily_rows = area_top_rows(behavior_subfamily_df, area_name, limit=3)
        top_behavior_rows = area_top_rows(behavior_df, area_name, limit=3)

        hypothesis = compose_area_investigation_hypothesis(
            top_behavior_family_rows,
            area_row,
        )
        field_checks = compose_area_field_checks(top_behavior_family_rows)

        pack_row = {
            "area_canonica": area_name,
            "event_month": pd.to_datetime(area_row["event_month"]).strftime("%Y-%m-%d"),
            "target_month": pd.to_datetime(area_row["target_month"]).strftime("%Y-%m-%d"),
            "baseline_predicted_risk_score": float(area_row["predicted_risk_score"]),
            "baseline_area_risk_rank": int(area_row["area_risk_rank"]),
            "expert_predicted_risk_score": float(area_row["expert_predicted_risk_score"]) if pd.notna(area_row["expert_predicted_risk_score"]) else np.nan,
            "expert_area_risk_rank": int(area_row["expert_area_risk_rank"]) if pd.notna(area_row["expert_area_risk_rank"]) else np.nan,
            "expert_minus_baseline_risk_score": float(area_row["expert_minus_baseline_risk_score"]) if pd.notna(area_row["expert_minus_baseline_risk_score"]) else np.nan,
            "area_unsafe_rate_12m": float(area_row["area_unsafe_rate_12m"]),
            "area_unsafe_row_rate_12m": float(area_row["area_unsafe_row_rate_12m"]),
            "area_obs_rows_12m": int(area_row["area_obs_rows_12m"]),
            "evt_total_rows_12m": int(area_row["evt_total_rows_12m"]) if pd.notna(area_row.get("evt_total_rows_12m")) else 0,
            "evt_smat_audit_count_12m": int(area_row["evt_smat_audit_count_12m"]) if pd.notna(area_row.get("evt_smat_audit_count_12m")) else 0,
            "evt_first_aid_count_12m": int(area_row["evt_first_aid_count_12m"]) if pd.notna(area_row.get("evt_first_aid_count_12m")) else 0,
            "evt_certified_injury_count_12m": int(area_row["evt_certified_injury_count_12m"]) if pd.notna(area_row.get("evt_certified_injury_count_12m")) else 0,
            "evt_incident_without_injury_count_12m": int(area_row["evt_incident_without_injury_count_12m"]) if pd.notna(area_row.get("evt_incident_without_injury_count_12m")) else 0,
            "evt_days_lost_sum_12m": float(area_row["evt_days_lost_sum_12m"]) if pd.notna(area_row.get("evt_days_lost_sum_12m")) else 0.0,
            "top_problem_codes_12m": normalize_whitespace(area_row.get("top_problem_codes_12m")),
            "top_description_topics_12m": normalize_whitespace(area_row.get("top_description_topics_12m")),
            "top_adverse_description_topics_12m": normalize_whitespace(area_row.get("top_adverse_description_topics_12m")),
            "latest_adverse_event_summary": normalize_whitespace(area_row.get("latest_adverse_event_summary")),
            "recent_adverse_events_12m": normalize_whitespace(area_row.get("recent_adverse_events_12m")),
            "similar_historical_events_summary": normalize_whitespace(area_row.get("similar_historical_events_summary")),
            "similar_historical_event_patterns": normalize_whitespace(area_row.get("similar_historical_event_patterns")),
            "top_behavior_families_expert_summary": format_named_rows(
                top_behavior_family_rows,
                "comportamento_famiglia",
                lambda row: f"IR {int(row['risk_rank'])}, unsafe {row['unsafe_rate']:.1%}, expert prio {row['expert_priority_score']:.2f}",
            ),
            "top_behavior_subfamilies_summary": format_named_rows(
                top_behavior_subfamily_rows,
                "comportamento_sottofamiglia_v2",
                lambda row: f"IR {int(row['severity_rank']) if pd.notna(row.get('severity_rank')) else 'n.d.'}, unsafe {row['unsafe_rate']:.1%}, prio {row['priority_score']:.2f}",
            ),
            "top_behavior_focus_summary": format_named_rows(
                top_behavior_rows,
                "comportamento_focus",
                lambda row: f"unsafe {row['unsafe_rate']:.1%}, {int(row['obs_rows'])} righe",
            ),
            "investigation_hypothesis": hypothesis,
            "field_checks": field_checks,
        }

        for index, (_, row) in enumerate(top_behavior_family_rows.head(3).iterrows(), start=1):
            pack_row[f"top_behavior_family_expert_{index}"] = normalize_whitespace(row["comportamento_famiglia"])
            pack_row[f"top_behavior_family_expert_{index}_risk_rank"] = int(row["risk_rank"])
            pack_row[f"top_behavior_family_expert_{index}_risk_label"] = normalize_whitespace(row["risk_label"])
        for index, (_, row) in enumerate(top_behavior_subfamily_rows.head(3).iterrows(), start=1):
            pack_row[f"top_behavior_subfamily_{index}"] = normalize_whitespace(row["comportamento_sottofamiglia_v2"])
            if pd.notna(row.get("severity_rank")):
                pack_row[f"top_behavior_subfamily_{index}_severity_rank"] = int(row["severity_rank"])
        for index, (_, row) in enumerate(top_behavior_rows.head(3).iterrows(), start=1):
            pack_row[f"top_behavior_focus_{index}"] = normalize_whitespace(row["comportamento_focus"])
        for index in range(1, 4):
            pack_row[f"similar_historical_event_{index}"] = normalize_whitespace(area_row.get(f"similar_historical_event_{index}"))

        pack_rows.append(pack_row)

    pack_df = pd.DataFrame(pack_rows).sort_values("baseline_area_risk_rank").reset_index(drop=True)
    pack_df.to_csv(output_dir / "area_investigation_pack_latest.csv", index=False)

    lines = [
        "# Investigation Pack Aree Prioritarie",
        "",
        f"Mese target: `{pd.to_datetime(pack_df['target_month'].iloc[0]).strftime('%Y-%m-%d')}`.",
        "Questo pack unisce ranking area-mese, driver comportamentali e contesto eventi recente.",
        "",
        "## Come usarlo",
        "",
        "- Partire dal ranking baseline per scegliere le prime aree da investigare.",
        "- Usare il punteggio expert come overlay sui comportamenti piu severi.",
        "- Verificare sul campo i driver indicati: non sono prove causali, ma ipotesi operative prioritarie.",
        "",
    ]
    for _, row in pack_df.head(top_n_report).iterrows():
        lines.append(f"## {int(row['baseline_area_risk_rank'])}. {row['area_canonica']}")
        lines.append("")
        lines.append(
            f"- Score baseline `{row['baseline_predicted_risk_score']:.3f}` (rank `{int(row['baseline_area_risk_rank'])}`), "
            f"score expert `{row['expert_predicted_risk_score']:.3f}` (rank `{int(row['expert_area_risk_rank'])}`) per `target_month={row['target_month']}`."
        )
        lines.append(
            f"- Unsafe ultimi 12 mesi: `{row['area_unsafe_rate_12m']:.3%}` su `{int(row['area_obs_rows_12m'])}` righe osservate."
        )
        lines.append(
            f"- Eventi ultimi 12 mesi: `SMAT={int(row['evt_smat_audit_count_12m'])}`, "
            f"`primo soccorso={int(row['evt_first_aid_count_12m'])}`, "
            f"`certificati={int(row['evt_certified_injury_count_12m'])}`, "
            f"`incidenti senza infortunio={int(row['evt_incident_without_injury_count_12m'])}`."
        )
        if normalize_whitespace(row.get("top_problem_codes_12m")):
            lines.append(f"- Problem code ricorrenti: {row['top_problem_codes_12m']}.")
        if normalize_whitespace(row.get("top_description_topics_12m")):
            lines.append(f"- Topic descrizione ricorrenti: {row['top_description_topics_12m']}.")
        if normalize_whitespace(row.get("top_adverse_description_topics_12m")):
            lines.append(f"- Topic descrizione sugli eventi avversi: {row['top_adverse_description_topics_12m']}.")
        if normalize_whitespace(row.get("similar_historical_event_patterns")):
            lines.append(f"- Pattern ricorrenti dai casi simili: {row['similar_historical_event_patterns']}.")
        if normalize_whitespace(row.get("latest_adverse_event_summary")):
            lines.append(f"- Ultimo evento avverso noto: {row['latest_adverse_event_summary']}.")
        if normalize_whitespace(row.get("top_behavior_families_expert_summary")):
            lines.append(f"- Famiglie comportamento piu critiche: {row['top_behavior_families_expert_summary']}.")
        if normalize_whitespace(row.get("top_behavior_subfamilies_summary")):
            lines.append(f"- Sottofamiglie comportamento piu critiche: {row['top_behavior_subfamilies_summary']}.")
        if normalize_whitespace(row.get("top_behavior_focus_summary")):
            lines.append(f"- Comportamenti da osservare: {row['top_behavior_focus_summary']}.")
        if normalize_whitespace(row.get("investigation_hypothesis")):
            lines.append(f"- Ipotesi di lavoro: {row['investigation_hypothesis']}")
        if normalize_whitespace(row.get("field_checks")):
            lines.append(f"- Checklist campo: {row['field_checks']}")
        if normalize_whitespace(row.get("recent_adverse_events_12m")):
            lines.append(f"- Storico eventi recente: {row['recent_adverse_events_12m']}")
        if normalize_whitespace(row.get("similar_historical_events_summary")):
            lines.append(f"- Casi storici simili: {row['similar_historical_events_summary']}")
        lines.append("")

    lines.append("## Nota")
    lines.append("")
    lines.append(
        "Il pack serve per orientare audit e sopralluoghi: suggerisce dove cercare i driver del rischio, ma non dimostra causalita diretta."
    )
    (output_dir / "area_investigation_pack_latest.md").write_text("\n".join(lines).strip() + "\n", encoding="utf-8")

    return {
        "status": "built",
        "rows": int(len(pack_df)),
        "report_top_areas": int(min(top_n_report, len(pack_df))),
        "analysis_window_start": analysis_window_start.strftime("%Y-%m-%d"),
        "analysis_window_end": analysis_window_end.strftime("%Y-%m-%d"),
        "similarity_report": similarity_report,
    }


def write_json(path: Path, payload: dict[str, Any]) -> None:
    serializable = json.loads(json.dumps(payload, default=serialize_scalar))
    path.write_text(json.dumps(serializable, indent=2, ensure_ascii=False), encoding="utf-8")


def summarize_model_variant(
    variant_name: str,
    model_report: dict[str, Any],
    walk_forward_report: dict[str, Any],
) -> dict[str, Any]:
    test_metrics = model_report.get("test_metrics", {}) if model_report else {}
    walk_metrics = walk_forward_report.get("overall_metrics", {}) if walk_forward_report else {}
    business_metrics = walk_forward_report.get("business_metrics", {}) if walk_forward_report else {}
    return {
        "variant": variant_name,
        "selected_model_holdout": model_report.get("selected_model"),
        "n_features": model_report.get("n_features"),
        "holdout_pr_auc": test_metrics.get("pr_auc"),
        "holdout_roc_auc": test_metrics.get("roc_auc"),
        "holdout_brier": test_metrics.get("brier_score"),
        "walk_forward_pr_auc": walk_metrics.get("pr_auc"),
        "walk_forward_roc_auc": walk_metrics.get("roc_auc"),
        "walk_forward_brier": walk_metrics.get("brier_score"),
        "walk_forward_top_1_hit_rate": business_metrics.get("top_1_hit_rate"),
        "walk_forward_top_3_precision_mean": business_metrics.get("top_3_precision_mean"),
        "walk_forward_top_3_recall_mean": business_metrics.get("top_3_recall_mean"),
    }


def write_business_report(
    output_dir: Path,
    cleaning_report: dict[str, Any],
    model_report: dict[str, Any],
    injury_detail_report: dict[str, Any],
) -> None:
    risk_scores_path = output_dir / "latest_risk_scores.csv"
    cause_latest_path = output_dir / "injury_cause_latest_predictions.csv"
    body_latest_path = output_dir / "injury_body_part_latest_predictions.csv"
    feature_importance_path = output_dir / "feature_importance.csv"

    risk_scores = pd.read_csv(risk_scores_path) if risk_scores_path.exists() else pd.DataFrame()
    cause_latest = pd.read_csv(cause_latest_path) if cause_latest_path.exists() else pd.DataFrame()
    body_latest = pd.read_csv(body_latest_path) if body_latest_path.exists() else pd.DataFrame()
    feature_importance = pd.read_csv(feature_importance_path) if feature_importance_path.exists() else pd.DataFrame()

    risk_test = model_report.get("test_metrics", {})
    cause_target = injury_detail_report.get("targets", {}).get("tipologia_causa_macro", {})
    body_target = injury_detail_report.get("targets", {}).get("parte_del_corpo_macro", {})

    lines: list[str] = []
    lines.append("# Recap Modello Rischio Operativo")
    lines.append("")
    lines.append("## Executive Summary")
    lines.append("")
    lines.append(
        "Il modello principale area-mese è abbastanza solido per fare prioritizzazione operativa e early warning, "
        "mentre i modelli di dettaglio su causa e parte del corpo sono da usare come supporto interpretativo e non come previsione deterministica."
    )
    lines.append("")
    lines.append("## Cosa È Stato Fatto")
    lines.append("")
    lines.append(
        f"- File eventi pulito da `101` a `{cleaning_report['events']['columns_after_cleaning']}` colonne, "
        f"con `{cleaning_report['events']['rows_after_cleaning']}` righe finali dopo l'esclusione di COVID e in itinere."
    )
    lines.append(
        f"- File osservazioni ridotto a `{cleaning_report['observations']['rows_after_cleaning']}` righe utili, "
        f"con `{cleaning_report['observations']['unique_activities_cleaned']}` attività clean e tassonomia attività consolidata."
    )
    behavior_taxonomy = cleaning_report.get("behavior_taxonomy", {})
    if behavior_taxonomy.get("status") == "loaded":
        lines.append(
            f"- Integrata tassonomia comportamenti v2 con `{behavior_taxonomy.get('lookup_rows')}` righe lookup e "
            f"`{cleaning_report['observations'].get('unique_behaviors_canonical', 0)}` comportamenti canonici finali."
        )
    lines.append(
        f"- Costruito panel `area-mese` con `{cleaning_report['panel']['rows']}` righe su periodo "
        f"`{cleaning_report['panel']['month_min']}` -> `{cleaning_report['panel']['month_max']}`."
    )
    lines.append(
        "- Addestrato un modello principale per predire il rischio di evento avverso del mese successivo "
        "(primo soccorso, infortunio con certificato, incidente senza infortunio)."
    )
    if injury_detail_report.get("status") == "trained":
        lines.append(
            "- Addestrato un secondo stadio sugli injury-like events per stimare macro-causa e macro-parte-del-corpo più probabili."
        )
    lines.append("")
    lines.append("## Risultati Principali")
    lines.append("")
    lines.append("### Modello Principale")
    lines.append("")
    lines.append(
        f"- Split temporale: train `{model_report['train_month_range'][0]}` -> `{model_report['train_month_range'][1]}`, "
        f"validation `{model_report['validation_month_range'][0]}` -> `{model_report['validation_month_range'][1]}`, "
        f"test `{model_report['test_month_range'][0]}` -> `{model_report['test_month_range'][1]}`."
    )
    lines.append(f"- Modello selezionato: `{model_report['selected_model']}`.")
    lines.append(
        f"- Performance test: `PR-AUC {risk_test.get('pr_auc', float('nan')):.3f}`, "
        f"`ROC-AUC {risk_test.get('roc_auc', float('nan')):.3f}`, "
        f"`Brier {risk_test.get('brier_score', float('nan')):.3f}`."
    )
    lines.append(
        "- Interpretazione: buono per ranking del rischio tra aree e per supportare la priorità di intervento, "
        "non da leggere come probabilità puntuale perfetta di singolo infortunio."
    )
    lines.append("")

    if not risk_scores.empty:
        lines.append("### Aree Più a Rischio Nell'Ultimo Scoring")
        lines.append("")
        for _, row in risk_scores.head(8).iterrows():
            lines.append(
                f"- `{row['area_canonica']}`: score `{row['predicted_risk_score']:.3f}` "
                f"per target month `{row['target_month']}`."
            )
        lines.append("")

    if injury_detail_report.get("status") == "trained":
        lines.append("### Secondo Stadio: Dettaglio Injury")
        lines.append("")
        lines.append(
            f"- Dataset injury-like: `{injury_detail_report['n_rows']}` righe, `{injury_detail_report['n_features']}` feature."
        )
        if cause_target:
            cause_test = cause_target.get("test_metrics", {})
            lines.append(
                f"- `Causa macro`: `accuracy {cause_test.get('accuracy', float('nan')):.3f}`, "
                f"`macro-F1 {cause_test.get('macro_f1', float('nan')):.3f}`, "
                f"`top-3 accuracy {cause_test.get('top_3_accuracy', float('nan')):.3f}`."
            )
        if body_target:
            body_test = body_target.get("test_metrics", {})
            lines.append(
                f"- `Parte del corpo macro`: `accuracy {body_test.get('accuracy', float('nan')):.3f}`, "
                f"`macro-F1 {body_test.get('macro_f1', float('nan')):.3f}`, "
                f"`top-3 accuracy {body_test.get('top_3_accuracy', float('nan')):.3f}`."
            )
        lines.append(
            "- Interpretazione: la `causa` può essere usata come shortlist top-3 da investigare; "
            "la `parte del corpo` oggi è ancora troppo debole per uso decisionale autonomo."
        )
        lines.append("")

    if not cause_latest.empty:
        lines.append("### Cause Più Probabili Nell'Ultimo Scoring")
        lines.append("")
        for _, row in cause_latest.iterrows():
            lines.append(
                f"- `{row['area_canonica']}`: top cause `{row['pred_tipologia_causa_macro_top_1']}` "
                f"({row['pred_tipologia_causa_macro_top_1_proba']:.3f}), "
                f"poi `{row['pred_tipologia_causa_macro_top_2']}` e `{row['pred_tipologia_causa_macro_top_3']}`."
            )
        lines.append("")

    if not body_latest.empty:
        lines.append("### Parti Del Corpo Più Probabili Nell'Ultimo Scoring")
        lines.append("")
        for _, row in body_latest.iterrows():
            lines.append(
                f"- `{row['area_canonica']}`: top body part `{row['pred_parte_del_corpo_macro_top_1']}` "
                f"({row['pred_parte_del_corpo_macro_top_1_proba']:.3f}), "
                f"poi `{row['pred_parte_del_corpo_macro_top_2']}` e `{row['pred_parte_del_corpo_macro_top_3']}`."
            )
        lines.append("")

    lines.append("## Come Usarlo Operativamente")
    lines.append("")
    lines.append("- Ogni mese ordinare le aree per score in `latest_risk_scores.csv` e focalizzare il review HSE sulle prime 3-5.")
    lines.append("- Per le aree ad alto rischio leggere le top-3 cause del secondo stadio come ipotesi di lavoro, non come verità certa.")
    lines.append("- Collegare le top cause alle famiglie attività/unsafe dominanti dell'area per decidere audit, coaching o interventi tecnici.")
    lines.append("- Usare il modello per prioritizzare, non per chiudere l'analisi: il sopralluogo resta indispensabile.")
    lines.append("")

    lines.append("## Cosa Predice Bene")
    lines.append("")
    lines.append("- Il ranking del rischio relativo tra aree e periodi.")
    lines.append("- L'emersione di aree cronicamente esposte dove unsafe e storico eventi si rafforzano a vicenda.")
    lines.append("- Una shortlist ragionevole di macro-cause possibili quando il rischio injury è elevato.")
    lines.append("")

    lines.append("## Cosa Non Predice Bene")
    lines.append("")
    lines.append("- Il singolo infortunio individuale.")
    lines.append("- La parte del corpo precisa con affidabilità alta.")
    lines.append("- Le classi molto rare o residuali (`Altro / Non definita`, casi eccezionali, dinamiche poco frequenti).")
    lines.append("")

    if not feature_importance.empty:
        lines.append("## Segnali Più Informativi Nel Modello Principale")
        lines.append("")
        for _, row in feature_importance.head(10).iterrows():
            lines.append(f"- `{row['feature']}`: importanza `{row['importance']:.3f}`.")
        lines.append("")

    lines.append("## Miglioramenti Consigliati")
    lines.append("")
    lines.append("- Aggiungere dati di esposizione: ore lavorate, headcount, turni, volumi produttivi.")
    lines.append("- Raffinare ulteriormente le macro-classi di causa con review HSE.")
    lines.append("- Costruire una tassonomia dedicata anche per i `comportamenti` osservati, non solo per le attività.")
    lines.append("- Valutare panel `area-settimana` se i volumi reggono.")
    lines.append("- Misurare performance business-oriented: precision@k, lift sulle top aree, stabilità mese su mese.")
    lines.append("")

    recap_path = output_dir / "business_model_recap.md"
    auto_recap_path = output_dir / "business_model_recap_auto.md"
    target_path = auto_recap_path if recap_path.exists() else recap_path
    target_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main(argv: list[str] | None = None) -> None:
    parser = argparse.ArgumentParser(description="Cleaning e baseline modeling per rischio operativo Agordo.")
    parser.add_argument(
        "--events",
        default="SAMPLE 2 - Estratto_eventi_ITA_2016_2026WK16.xlsx",
        help="Percorso del file eventi Excel.",
    )
    parser.add_argument(
        "--observations",
        default="Osservazioni_Agordo_cf09b6dda9dc4bfbb40d2fee5335bf5d.xlsx",
        help="Percorso del file osservazioni Excel.",
    )
    parser.add_argument(
        "--output-dir",
        default="artifacts",
        help="Cartella di output per dataset puliti, report e modello.",
    )
    parser.add_argument(
        "--behavior-expert-ranks",
        default="artifacts/behavior_family_expert_risk_rank.csv",
        help="CSV manuale con ranking esperto IR-like per famiglia comportamento (`1` alto, `5` basso).",
    )
    parser.add_argument(
        "--behavior-taxonomy",
        default="artifacts/behavior_taxonomy_pipeline_v2.csv",
        help="CSV opzionale con tassonomia comportamento v2, canonicalizzazione e gerarchia HSE.",
    )
    args = parser.parse_args(argv)

    base_dir = Path.cwd()
    events_path = (base_dir / args.events).resolve()
    observations_path = (base_dir / args.observations).resolve()
    output_dir = (base_dir / args.output_dir).resolve()
    behavior_expert_ranks_path = (base_dir / args.behavior_expert_ranks).resolve()
    behavior_taxonomy_path = Path(args.behavior_taxonomy).expanduser()
    if not behavior_taxonomy_path.is_absolute():
        behavior_taxonomy_path = (base_dir / behavior_taxonomy_path).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    behavior_taxonomy_source_df, behavior_taxonomy_lookup_df, behavior_taxonomy_profile = load_behavior_taxonomy(
        behavior_taxonomy_path
    )
    if not behavior_taxonomy_source_df.empty:
        behavior_taxonomy_source_df.to_csv(output_dir / "behavior_taxonomy_pipeline_v2.csv", index=False)

    raw_events_df, raw_events_profile = load_events_dataframe(events_path)
    cleaned_events_df, events_profile, event_area_mapping_df = clean_events_dataframe(raw_events_df)
    event_area_mapping_df.to_csv(output_dir / "event_area_mapping.csv", index=False)

    semantic_df, semantic_terms_df = build_event_semantic_features(cleaned_events_df, output_dir)
    description_topic_df, description_topic_summary_df, description_topic_by_problem_df = build_event_description_topic_features(
        cleaned_events_df,
        output_dir,
    )
    cleaned_events_df.to_csv(output_dir / "events_cleaned.csv", index=False)

    event_month_df = build_event_area_month_dataset(cleaned_events_df, semantic_df, description_topic_df)
    event_month_df.to_csv(output_dir / "events_area_month.csv", index=False)

    (
        obs_monthly_df,
        obs_profile,
        group_mapping_df,
        activity_mapping_df,
        activity_catalog_df,
        cleaned_activity_summary_df,
        behavior_mapping_df,
        cleaned_behavior_summary_df,
        behavior_canonical_summary_df,
    ) = stream_clean_observations(
        observations_path,
        output_dir / "observations_cleaned.csv",
        behavior_taxonomy_lookup_df=behavior_taxonomy_lookup_df,
    )
    obs_monthly_df.to_csv(output_dir / "observations_area_month.csv", index=False)
    group_mapping_df.to_csv(output_dir / "group_area_mapping.csv", index=False)
    activity_mapping_df.to_csv(output_dir / "activity_mapping.csv", index=False)
    activity_catalog_df.to_csv(output_dir / "activity_catalog_full.csv", index=False)
    cleaned_activity_summary_df.to_csv(output_dir / "activity_clean_summary.csv", index=False)
    behavior_mapping_df.to_csv(output_dir / "behavior_mapping.csv", index=False)
    cleaned_behavior_summary_df.to_csv(output_dir / "behavior_clean_summary.csv", index=False)
    behavior_canonical_summary_df.to_csv(output_dir / "behavior_canonical_summary.csv", index=False)
    observations_cleaned_df = pd.read_csv(output_dir / "observations_cleaned.csv", parse_dates=["event_month"])
    behavior_expert_risk_rank_df = load_behavior_expert_risk_rank(behavior_expert_ranks_path, observations_cleaned_df)

    panel_df = build_model_dataset(obs_monthly_df, event_month_df)
    panel_df.to_csv(output_dir / "risk_panel_area_month.csv", index=False)

    model_report, latest_scores_df = train_baseline_model(panel_df, output_dir)
    walk_forward_report = run_walk_forward_backtest(panel_df, output_dir)
    (
        behavior_weighted_panel_df,
        behavior_month_df,
        behavior_weight_df,
    ) = build_behavior_weighted_panel(panel_df, observations_cleaned_df, output_dir)
    behavior_weighted_model_report, _ = train_baseline_model(
        behavior_weighted_panel_df,
        output_dir,
        artifact_prefix="behavior_weighted_",
    )
    behavior_weighted_walk_forward_report = run_walk_forward_backtest(
        behavior_weighted_panel_df,
        output_dir,
        artifact_prefix="behavior_weighted_",
    )
    (
        behavior_subfamily_weighted_panel_df,
        behavior_subfamily_month_df,
        behavior_subfamily_weight_df,
    ) = build_behavior_subfamily_weighted_panel(panel_df, observations_cleaned_df, output_dir)
    behavior_subfamily_weighted_model_report, _ = train_baseline_model(
        behavior_subfamily_weighted_panel_df,
        output_dir,
        artifact_prefix="behavior_subfamily_weighted_",
    )
    behavior_subfamily_weighted_walk_forward_report = run_walk_forward_backtest(
        behavior_subfamily_weighted_panel_df,
        output_dir,
        artifact_prefix="behavior_subfamily_weighted_",
    )
    behavior_expert_panel_df, behavior_expert_month_df = build_behavior_expert_weighted_panel(
        panel_df,
        observations_cleaned_df,
        behavior_expert_risk_rank_df,
        output_dir,
    )
    behavior_expert_model_report, behavior_expert_latest_scores_df = train_baseline_model(
        behavior_expert_panel_df,
        output_dir,
        artifact_prefix="behavior_expert_",
    )
    behavior_expert_walk_forward_report = run_walk_forward_backtest(
        behavior_expert_panel_df,
        output_dir,
        artifact_prefix="behavior_expert_",
    )
    variant_comparison_df = pd.DataFrame(
        [
            summarize_model_variant("baseline", model_report, walk_forward_report),
            summarize_model_variant(
                "behavior_weighted_experiment",
                behavior_weighted_model_report,
                behavior_weighted_walk_forward_report,
            ),
            summarize_model_variant(
                "behavior_subfamily_weighted_experiment",
                behavior_subfamily_weighted_model_report,
                behavior_subfamily_weighted_walk_forward_report,
            ),
            summarize_model_variant(
                "behavior_expert_weighted_experiment",
                behavior_expert_model_report,
                behavior_expert_walk_forward_report,
            ),
        ]
    )
    variant_comparison_df.to_csv(output_dir / "main_model_variant_comparison.csv", index=False)
    area_deepdive_report = build_area_risk_deepdive(observations_cleaned_df, latest_scores_df, output_dir)
    investigation_pack_report = build_area_investigation_pack(
        output_dir,
        cleaned_events_df,
        behavior_expert_latest_scores_df,
        behavior_expert_risk_rank_df,
    )
    injury_detail_df = build_injury_detail_dataset(panel_df, cleaned_events_df)
    injury_detail_df.to_csv(output_dir / "injury_detail_dataset.csv", index=False)
    injury_detail_report = train_injury_detail_models(injury_detail_df, panel_df, output_dir)

    cleaning_report = {
        "events": {**raw_events_profile, **events_profile},
        "observations": obs_profile,
        "behavior_taxonomy": behavior_taxonomy_profile,
        "semantic": {
            "semantic_terms_rows": int(len(semantic_terms_df)),
            "semantic_component_rows": int(len(semantic_df)),
        },
        "description_topics": {
            "topic_area_month_rows": int(len(description_topic_df)),
            "topic_summary_rows": int(len(description_topic_summary_df)),
            "topic_by_problem_rows": int(len(description_topic_by_problem_df)),
        },
        "injury_detail": {
            "rows": int(len(injury_detail_df)),
            "cause_macro_distribution": {
                str(label): int(count)
                for label, count in injury_detail_df["tipologia_causa_macro"].value_counts(dropna=False).to_dict().items()
            } if not injury_detail_df.empty else {},
            "body_part_macro_distribution": {
                str(label): int(count)
                for label, count in injury_detail_df["parte_del_corpo_macro"].value_counts(dropna=False).to_dict().items()
            } if not injury_detail_df.empty else {},
        },
        "area_deepdive": area_deepdive_report,
        "investigation_pack": investigation_pack_report,
        "panel": {
            "rows": int(len(panel_df)),
            "areas": sorted(panel_df["area_canonica"].dropna().unique().tolist()),
            "month_min": panel_df["event_month"].min(),
            "month_max": panel_df["event_month"].max(),
            "target_adverse_positive_rate": float(panel_df["target_adverse_next_month"].mean()),
            "target_injury_positive_rate": float(panel_df["target_injury_next_month"].mean()),
        },
        "behavior_weighted_experiment": {
            "behavior_area_month_rows": int(len(behavior_month_df)),
            "behavior_weight_rows": int(len(behavior_weight_df)),
            "panel_rows": int(len(behavior_weighted_panel_df)),
            "comparison_file": "main_model_variant_comparison.csv",
        },
        "behavior_subfamily_weighted_experiment": {
            "behavior_subfamily_area_month_rows": int(len(behavior_subfamily_month_df)),
            "behavior_subfamily_weight_rows": int(len(behavior_subfamily_weight_df)),
            "panel_rows": int(len(behavior_subfamily_weighted_panel_df)),
            "comparison_file": "main_model_variant_comparison.csv",
        },
        "behavior_expert_experiment": {
            "risk_rank_file": str(behavior_expert_ranks_path),
            "risk_rank_rows": int(len(behavior_expert_risk_rank_df)),
            "behavior_expert_area_month_rows": int(len(behavior_expert_month_df)),
            "panel_rows": int(len(behavior_expert_panel_df)),
            "comparison_file": "main_model_variant_comparison.csv",
        },
    }
    write_json(output_dir / "cleaning_report.json", cleaning_report)
    write_json(output_dir / "model_report.json", model_report)
    write_json(output_dir / "walk_forward_report.json", walk_forward_report)
    write_json(output_dir / "behavior_weighted_model_report.json", behavior_weighted_model_report)
    write_json(output_dir / "behavior_weighted_walk_forward_report.json", behavior_weighted_walk_forward_report)
    write_json(output_dir / "behavior_subfamily_weighted_model_report.json", behavior_subfamily_weighted_model_report)
    write_json(output_dir / "behavior_subfamily_weighted_walk_forward_report.json", behavior_subfamily_weighted_walk_forward_report)
    write_json(output_dir / "behavior_expert_model_report.json", behavior_expert_model_report)
    write_json(output_dir / "behavior_expert_walk_forward_report.json", behavior_expert_walk_forward_report)
    write_json(output_dir / "injury_detail_model_report.json", injury_detail_report)
    write_business_report(output_dir, cleaning_report, model_report, injury_detail_report)

    print(f"Pipeline completata. Output disponibili in: {output_dir}")


if __name__ == "__main__":
    main()
